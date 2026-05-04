from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from .forms import ImportExcelForm
from .models import Assignment, Principal, School, Supervisor
from .utils.import_schools_with_supervisors import (
    build_schools_with_supervisors_template,
    commit_schools_with_supervisors_import,
    parse_schools_with_supervisors_workbook,
)


# ============================================================================
# Stats
# ============================================================================
@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0


# ============================================================================
# Rejected (Session Keys)
# ============================================================================
SESSION_REJ_HEADERS = "import_rejected_headers"
SESSION_REJ_ROWS = "import_rejected_rows"
MAX_REJECTED_IN_SESSION = 3000
SCHOOLS_WITH_SUPERVISORS_IMPORT_SESSION_KEY = "schools_with_supervisors_import_preview"


def _rej_add(rejected: list[dict], row: dict, reason: str, importer: str) -> None:
    if len(rejected) >= MAX_REJECTED_IN_SESSION:
        return
    x = dict(row or {})
    x["_reason"] = reason
    x["_importer"] = importer
    rejected.append(x)


# ============================================================================
# Helpers
# ============================================================================
def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _digits(v: Any) -> str:
    """
    يرجع الأرقام فقط من أي قيمة:
    - 1020103717 -> 1020103717
    - "1020103717 " -> 1020103717
    - 70228.0 -> 70228
    - "70228-1" -> 702281
    """
    s = _norm(v)
    if not s:
        return ""
    s = s.replace(".0", "").strip()
    return re.sub(r"\D+", "", s)


def _code(v: Any) -> str:
    """
    كود المدرسة / الرقم الإحصائي:
    - يحافظ على الحروف مثل M3964353
    - يحول 70228.0 إلى 70228
    - يحذف الفراغات فقط
    """
    s = _norm(v)
    if not s:
        return ""
    s = s.replace(".0", "").strip()
    s = s.replace(" ", "")
    return s


def _to_bool(v: Any) -> bool:
    s = _norm(v).lower()
    if s in {"1", "true", "yes", "y", "نعم"}:
        return True
    if s in {"0", "false", "no", "n", "لا"}:
        return False
    return True


def _canon_header(h: str) -> str:
    """
    تحويل الهيدر (عربي/إنجليزي) إلى مفاتيح قياسية موحدة.
    """
    x = _norm(h).lower()
    x = x.replace("ـ", "").replace("_", " ").strip()

    # ---------- Schools ----------
    if x in {"stat_code", "stat code", "الرقم الإحصائي", "الرقم الاحصائي", "رقم احصائي", "code"}:
        return "stat_code"
    if x in {"name", "اسم المدرسة", "المدرسة", "schoolname"}:
        return "name"
    if x in {"gender", "الجنس"}:
        return "gender"
    if x in {"is_active", "active", "نشط"}:
        return "is_active"

    # ---------- Principals ----------
    if x in {"school_stat_code", "school stat code", "رقم المدرسة", "رقم احصائي المدرسة"}:
        return "school_stat_code"
    if x in {"full_name", "full name", "الاسم", "اسم القائد", "اسم القائدة", "اسم المدير", "اسم المديرة"}:
        return "full_name"
    if x in {"mobile", "الجوال", "رقم الجوال", "الهاتف", "phone"}:
        return "mobile"

    # ---------- Supervisors ----------
    if x in {"national_id", "national id", "السجل المدني", "رقم الهوية", "الهوية", "nid"}:
        return "national_id"
    if x in {"supervisor_national_id", "supervisor national id", "رقم هوية المشرف"}:
        return "supervisor_national_id"
    if x in {"supervisor_name", "supervisor name", "اسم المشرف", "المشرف"}:
        return "supervisor_name"

    # ---------- Assignments ----------
    if x in {"school", "school_stat_code", "school stat code"}:
        return "school_stat_code"
    if x in {"supervisor", "supervisor_national_id"}:
        return "supervisor_national_id"

    return _norm(h)


def _sheet_rows(file) -> list[dict]:
    """
    يقرأ الشيت ويعيد list[dict] بحيث مفاتيح الأعمدة تكون:
    - المفاتيح الأصلية
    - والمفاتيح القياسية canonical
    """
    wb = load_workbook(filename=file, data_only=True)
    ws = wb.active

    out: list[dict] = []
    headers_raw: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers_raw = [_norm(x) for x in row]
            continue

        if not any(x is not None and _norm(x) != "" for x in row):
            continue

        rec: dict[str, Any] = {}

        for j in range(len(headers_raw)):
            key = headers_raw[j] if j < len(headers_raw) else f"col_{j}"
            rec[key] = row[j] if j < len(row) else None

        for j in range(len(headers_raw)):
            canon = _canon_header(headers_raw[j])
            if canon and canon not in rec:
                rec[canon] = row[j] if j < len(row) else None

        out.append(rec)

    return out




# ============================================================================
# Schools + Supervisors Assignment Import Helpers
# ============================================================================
def _import_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in ("1", "true", "yes", "on")


def _schools_supervisors_default_options() -> dict[str, bool]:
    return {
        "create_missing_schools": False,
        "update_schools": True,
        "create_missing_supervisors": True,
        "update_supervisors": True,
        "create_missing_sectors": False,
        "allow_transfer": False,
        "deactivate_when_supervisor_blank": False,
    }


def _schools_supervisors_options_from_post(request: HttpRequest) -> dict[str, bool]:
    return {
        "create_missing_schools": _import_bool(request.POST.get("create_missing_schools"), False),
        "update_schools": _import_bool(request.POST.get("update_schools"), True),
        "create_missing_supervisors": _import_bool(request.POST.get("create_missing_supervisors"), True),
        "update_supervisors": _import_bool(request.POST.get("update_supervisors"), True),
        "create_missing_sectors": _import_bool(request.POST.get("create_missing_sectors"), False),
        "allow_transfer": _import_bool(request.POST.get("allow_transfer"), False),
        "deactivate_when_supervisor_blank": _import_bool(
            request.POST.get("deactivate_when_supervisor_blank"),
            False,
        ),
    }


def _uploaded_excel_file(request: HttpRequest):
    return (
        request.FILES.get("file")
        or request.FILES.get("excel_file")
        or request.FILES.get("upload")
    )


def _store_rejected_errors_for_download(request: HttpRequest, errors: list[dict]) -> None:
    if not errors:
        request.session.pop(SESSION_REJ_HEADERS, None)
        request.session.pop(SESSION_REJ_ROWS, None)
        request.session.modified = True
        return

    normalized_rows: list[dict] = []
    for err in errors[:MAX_REJECTED_IN_SESSION]:
        normalized_rows.append(
            {
                "_importer": "schools_with_supervisors",
                "_reason": err.get("message", ""),
                "row_no": err.get("row_no", ""),
                "school_stat_code": err.get("school_stat_code", ""),
                "school_name": err.get("school_name", ""),
                "supervisor": err.get("supervisor", ""),
            }
        )

    request.session[SESSION_REJ_HEADERS] = [
        "_importer",
        "_reason",
        "row_no",
        "school_stat_code",
        "school_name",
        "supervisor",
    ]
    request.session[SESSION_REJ_ROWS] = normalized_rows
    request.session.modified = True


# ============================================================================
# Importers
# ============================================================================
def _import_schools(file, gender: str, rejected: list[dict]) -> ImportStats:
    """
    الأعمدة المتوقعة:
      stat_code | name | is_active (اختياري)
    """
    st = ImportStats()
    rows = _sheet_rows(file)

    for r in rows:
        stat_code = _code(r.get("stat_code"))
        name = _norm(r.get("name"))

        if not stat_code or not name:
            st.skipped += 1
            _rej_add(rejected, r, "نقص الرقم الإحصائي أو الاسم", "schools")
            continue

        defaults = {
            "name": name,
            "gender": gender,
            "is_active": _to_bool(r.get("is_active")) if "is_active" in r else True,
        }

        _, created = School.objects.update_or_create(
            stat_code=stat_code,
            defaults=defaults,
        )
        if created:
            st.created += 1
        else:
            st.updated += 1

    return st


def _import_principals(file, rejected: list[dict]) -> ImportStats:
    """
    الأعمدة المتوقعة:
      school_stat_code | full_name | mobile (اختياري)

    يدعم كذلك الملفات التي يكون فيها العمود باسم:
      الرقم الإحصائي
    """
    st = ImportStats()
    rows = _sheet_rows(file)

    for r in rows:
        school_stat_code = _code(r.get("school_stat_code")) or _code(r.get("stat_code"))
        full_name = _norm(r.get("full_name"))

        if not school_stat_code or not full_name:
            st.skipped += 1
            _rej_add(rejected, r, "نقص الرقم الإحصائي للمدرسة أو اسم المدير", "principals")
            continue

        school = School.objects.filter(stat_code=school_stat_code).first()
        if not school:
            st.skipped += 1
            _rej_add(rejected, r, f"المدرسة غير موجودة: {school_stat_code}", "principals")
            continue

        defaults = {
            "full_name": full_name,
            "mobile": _digits(r.get("mobile")) or None,
        }

        _, created = Principal.objects.update_or_create(
            school=school,
            defaults=defaults,
        )
        if created:
            st.created += 1
        else:
            st.updated += 1

    return st


def _import_supervisors(file, rejected: list[dict]) -> ImportStats:
    """
    الأعمدة المتوقعة:
      national_id | full_name | mobile (اختياري) | is_active (اختياري)
    """
    st = ImportStats()
    rows = _sheet_rows(file)

    for r in rows:
        national_id = _digits(r.get("national_id")) or _digits(r.get("supervisor_national_id"))
        full_name = _norm(r.get("full_name")) or _norm(r.get("supervisor_name"))

        if not national_id or not full_name:
            st.skipped += 1
            _rej_add(rejected, r, "نقص رقم الهوية أو اسم المشرف", "supervisors")
            continue

        defaults = {
            "full_name": full_name,
            "mobile": _digits(r.get("mobile")) or None,
            "is_active": _to_bool(r.get("is_active")) if "is_active" in r else True,
        }

        _, created = Supervisor.objects.update_or_create(
            national_id=national_id,
            defaults=defaults,
        )
        if created:
            st.created += 1
        else:
            st.updated += 1

    return st


def _import_assignments(file, rejected: list[dict]) -> ImportStats:
    """
    يدعم:
      supervisor_national_id | school_stat_code | is_active (اختياري)
    أو الصيغ العربية الشائعة.
    """
    st = ImportStats()
    rows = _sheet_rows(file)

    for r in rows:
        sup_nid = _digits(r.get("supervisor_national_id")) or _digits(r.get("national_id"))

        if not sup_nid:
            sup_nid = (
                _digits(r.get("supervisor_name"))
                or _digits(r.get("اسم المشرف"))
                or _digits(r.get("المشرف"))
            )

        school_stat_code = (
            _code(r.get("school_stat_code"))
            or _code(r.get("stat_code"))
            or _code(r.get("الرقم الإحصائي"))
            or _code(r.get("الرقم الاحصائي"))
        )

        if not sup_nid or not school_stat_code:
            st.skipped += 1
            _rej_add(rejected, r, "نقص هوية المشرف أو الرقم الإحصائي للمدرسة", "assignments")
            continue

        supervisor = Supervisor.objects.filter(national_id=sup_nid).first()
        if not supervisor:
            st.skipped += 1
            _rej_add(rejected, r, f"المشرف غير موجود: {sup_nid}", "assignments")
            continue

        school = School.objects.filter(stat_code=school_stat_code).first()
        if not school:
            st.skipped += 1
            _rej_add(rejected, r, f"المدرسة غير موجودة: {school_stat_code}", "assignments")
            continue

        defaults = {
            "is_active": _to_bool(r.get("is_active")) if "is_active" in r else True,
        }

        _, created = Assignment.objects.update_or_create(
            supervisor=supervisor,
            school=school,
            defaults=defaults,
        )

        if created:
            st.created += 1
        else:
            st.updated += 1

    return st


# ============================================================================
# View: Manager Import
# ============================================================================
@staff_member_required
def manager_import_view(request: HttpRequest) -> HttpResponse:
    results: dict[str, ImportStats] = {}
    rejected: list[dict] = []

    if request.method == "POST":
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    if form.cleaned_data.get("schools_boys"):
                        results["المدارس (بنين)"] = _import_schools(
                            form.cleaned_data["schools_boys"],
                            "boys",
                            rejected,
                        )

                    if form.cleaned_data.get("schools_girls"):
                        results["المدارس (بنات)"] = _import_schools(
                            form.cleaned_data["schools_girls"],
                            "girls",
                            rejected,
                        )

                    if form.cleaned_data.get("principals"):
                        results["مديرو المدارس"] = _import_principals(
                            form.cleaned_data["principals"],
                            rejected,
                        )

                    if form.cleaned_data.get("supervisors"):
                        results["المشرفون"] = _import_supervisors(
                            form.cleaned_data["supervisors"],
                            rejected,
                        )

                    if form.cleaned_data.get("assignments"):
                        results["الإسنادات"] = _import_assignments(
                            form.cleaned_data["assignments"],
                            rejected,
                        )

                if rejected:
                    keys = set()
                    for rr in rejected:
                        keys |= set(rr.keys())

                    fixed = ["_importer", "_reason"]
                    headers = fixed + sorted([k for k in keys if k not in fixed])

                    request.session[SESSION_REJ_HEADERS] = headers
                    request.session[SESSION_REJ_ROWS] = rejected
                else:
                    request.session.pop(SESSION_REJ_HEADERS, None)
                    request.session.pop(SESSION_REJ_ROWS, None)

                request.session.modified = True

                if rejected and len(rejected) >= MAX_REJECTED_IN_SESSION:
                    messages.success(
                        request,
                        f"تمت عملية الاستيراد بنجاح ✅ مع وجود مرفوضات كثيرة. "
                        f"تم حفظ أول {MAX_REJECTED_IN_SESSION} سجل مرفوض فقط للتنزيل."
                    )
                elif rejected:
                    messages.success(
                        request,
                        f"تمت عملية الاستيراد بنجاح ✅ (مرفوض: {len(rejected)})"
                    )
                else:
                    messages.success(request, "تمت عملية الاستيراد بنجاح ✅")

            except Exception as e:
                messages.error(request, f"فشل الاستيراد: {e}")
        else:
            messages.error(request, "تحقق من الملفات المرفوعة (ارفع ملفًا واحدًا على الأقل).")
    else:
        form = ImportExcelForm()
        request.session.pop(SESSION_REJ_HEADERS, None)
        request.session.pop(SESSION_REJ_ROWS, None)
        request.session.modified = True

    return render(
        request,
        "visits/manager_import.html",
        {
            "form": form,
            "results": {k: asdict(v) for k, v in results.items()},
        },
    )




# ============================================================================
# View: Import Schools With Assigned Supervisors
# ============================================================================
@staff_member_required
def admin_schools_with_supervisors_import_view(request: HttpRequest) -> HttpResponse:
    """
    صفحة استيراد المدارس وإسنادها للمشرفين.

    آلية العمل:
    1. رفع ملف Excel.
    2. فحص الملف دون حفظ.
    3. عند نجاح الفحص، تحفظ البيانات مؤقتًا في session.
    4. الضغط على تنفيذ الاستيراد يحفظ المدارس/المشرفين/الإسنادات حسب الخيارات.
    """
    context = {
        "result": None,
        "commit_result": None,
        "options": _schools_supervisors_default_options(),
    }

    if request.method == "POST":
        action = (request.POST.get("action") or "preview").strip()

        if action == "commit":
            payload = request.session.get(SCHOOLS_WITH_SUPERVISORS_IMPORT_SESSION_KEY)

            if not payload or not payload.get("rows"):
                messages.error(request, "لا توجد بيانات جاهزة للتنفيذ. ارفع الملف وافحصه أولًا.")
                return render(request, "visits/admin_schools_with_supervisors_import.html", context)

            rows = payload["rows"]
            options = payload.get("options", _schools_supervisors_default_options())

            try:
                commit_result = commit_schools_with_supervisors_import(
                    rows,
                    create_missing_schools=bool(options.get("create_missing_schools", False)),
                    update_schools=bool(options.get("update_schools", True)),
                    create_missing_supervisors=bool(options.get("create_missing_supervisors", True)),
                    update_supervisors=bool(options.get("update_supervisors", True)),
                    create_missing_sectors=bool(options.get("create_missing_sectors", False)),
                    allow_transfer=bool(options.get("allow_transfer", False)),
                    deactivate_when_supervisor_blank=bool(options.get("deactivate_when_supervisor_blank", False)),
                )

                request.session.pop(SCHOOLS_WITH_SUPERVISORS_IMPORT_SESSION_KEY, None)
                request.session.pop(SESSION_REJ_HEADERS, None)
                request.session.pop(SESSION_REJ_ROWS, None)
                request.session.modified = True

                messages.success(request, "تم تنفيذ استيراد المدارس وإسنادها للمشرفين بنجاح ✅")

                context["commit_result"] = commit_result
                context["options"] = options

            except Exception as e:
                messages.error(request, f"فشل تنفيذ الاستيراد: {e}")

            return render(request, "visits/admin_schools_with_supervisors_import.html", context)

        options = _schools_supervisors_options_from_post(request)
        uploaded_file = _uploaded_excel_file(request)

        context["options"] = options

        if not uploaded_file:
            messages.error(request, "فضلاً اختر ملف Excel.")
            return render(request, "visits/admin_schools_with_supervisors_import.html", context)

        try:
            result = parse_schools_with_supervisors_workbook(
                uploaded_file,
                create_missing_schools=options["create_missing_schools"],
                update_schools=options["update_schools"],
                create_missing_supervisors=options["create_missing_supervisors"],
                update_supervisors=options["update_supervisors"],
                create_missing_sectors=options["create_missing_sectors"],
                allow_transfer=options["allow_transfer"],
                deactivate_when_supervisor_blank=options["deactivate_when_supervisor_blank"],
            )
        except Exception as e:
            result = {
                "ok": False,
                "rows": [],
                "errors": [
                    {
                        "row_no": "-",
                        "school_stat_code": "",
                        "school_name": "",
                        "supervisor": "",
                        "message": f"تعذر فحص الملف: {e}",
                    }
                ],
                "summary": {},
            }

        context["result"] = result

        if result.get("ok"):
            request.session[SCHOOLS_WITH_SUPERVISORS_IMPORT_SESSION_KEY] = {
                "rows": result.get("rows", []),
                "options": options,
            }
            request.session.pop(SESSION_REJ_HEADERS, None)
            request.session.pop(SESSION_REJ_ROWS, None)
            request.session.modified = True

            messages.success(request, "تم فحص الملف بنجاح، ويمكنك الآن تنفيذ الاستيراد.")
        else:
            request.session.pop(SCHOOLS_WITH_SUPERVISORS_IMPORT_SESSION_KEY, None)
            _store_rejected_errors_for_download(request, result.get("errors", []))

            messages.error(request, "يوجد أخطاء في الملف. صحح الأخطاء ثم أعد الرفع.")

        return render(request, "visits/admin_schools_with_supervisors_import.html", context)

    return render(request, "visits/admin_schools_with_supervisors_import.html", context)


@staff_member_required
def admin_schools_with_supervisors_import_template_view(request: HttpRequest) -> HttpResponse:
    """
    تحميل قالب Excel الخاص باستيراد المدارس وإسنادها للمشرفين.
    """
    wb = build_schools_with_supervisors_template()

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="schools_with_supervisors_template.xlsx"'
    return resp




# ============================================================================
# View: Export Schools With Assigned Supervisors
# ============================================================================
def _export_gender_label(value: Any) -> str:
    v = _norm(value).lower()
    if v in {"boys", "male", "m", "بنين"}:
        return "بنين"
    if v in {"girls", "female", "f", "بنات"}:
        return "بنات"
    return _norm(value) or ""


def _export_bool_label(value: Any, true_label: str = "نشط", false_label: str = "غير نشط") -> str:
    return true_label if bool(value) else false_label


def _school_sector_name(school: School) -> str:
    sector = getattr(school, "sector", None)
    return getattr(sector, "name", "") or ""


def _supervisor_value(supervisor: Supervisor | None, field_name: str) -> str:
    if not supervisor:
        return ""
    return _norm(getattr(supervisor, field_name, "") or "")


def _build_schools_with_supervisors_export_workbook(*, scope: str = "all") -> Workbook:
    """
    تصدير المدارس بالمشرفين بنفس منطق قالب الاستيراد.

    scope:
    - all: جميع المدارس
    - assigned: المدارس المسندة فقط
    - unassigned: المدارس غير المسندة فقط
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "المدارس والمشرفون"
    ws.sheet_view.rightToLeft = True

    headers = [
        "الرقم الإحصائي",
        "اسم المدرسة",
        "جنس المدرسة",
        "القطاع",
        "حالة المدرسة",
        "سجل المشرف",
        "اسم المشرف المسند",
        "جوال المشرف",
        "بريد المشرف",
        "جنس المشرف",
        "حالة الإسناد",
        "ملاحظات",
    ]
    ws.append(headers)

    active_assignments = list(
        Assignment.objects.filter(is_active=True)
        .select_related("school", "supervisor")
        .order_by("school_id", "id")
    )

    assignment_by_school: dict[int, Assignment] = {}
    assignment_count_by_school: dict[int, int] = {}

    for assignment in active_assignments:
        school_id = assignment.school_id
        assignment_count_by_school[school_id] = assignment_count_by_school.get(school_id, 0) + 1

        # أول إسناد نشط هو الذي سيظهر في ملف التصدير،
        # مع وضع ملاحظة إذا وجدت إسنادات نشطة متعددة لنفس المدرسة.
        if school_id not in assignment_by_school:
            assignment_by_school[school_id] = assignment

    assigned_school_ids = set(assignment_by_school.keys())

    schools_qs = School.objects.all().order_by("name")

    if scope == "assigned":
        schools_qs = schools_qs.filter(id__in=assigned_school_ids)
    elif scope == "unassigned":
        schools_qs = schools_qs.exclude(id__in=assigned_school_ids)

    for school in schools_qs:
        assignment = assignment_by_school.get(school.id)
        supervisor = assignment.supervisor if assignment else None

        notes: list[str] = []

        if assignment_count_by_school.get(school.id, 0) > 1:
            notes.append("تنبيه: توجد أكثر من عملية إسناد نشطة لهذه المدرسة.")

        if supervisor and hasattr(supervisor, "is_active") and not getattr(supervisor, "is_active", True):
            notes.append("تنبيه: المشرف المسند غير نشط.")

        if not assignment:
            assignment_status = "غير مسندة"
        else:
            assignment_status = "نشط" if getattr(assignment, "is_active", False) else "غير نشط"

        ws.append(
            [
                _norm(getattr(school, "stat_code", "") or ""),
                _norm(getattr(school, "name", "") or ""),
                _export_gender_label(getattr(school, "gender", "") or ""),
                _school_sector_name(school),
                _export_bool_label(getattr(school, "is_active", True), "نشطة", "غير نشطة"),
                _supervisor_value(supervisor, "national_id"),
                _supervisor_value(supervisor, "full_name"),
                _supervisor_value(supervisor, "mobile"),
                _supervisor_value(supervisor, "email"),
                _export_gender_label(getattr(supervisor, "gender", "") if supervisor else ""),
                assignment_status,
                " ".join(notes),
            ]
        )

    widths = {
        "A": 18,
        "B": 42,
        "C": 14,
        "D": 24,
        "E": 16,
        "F": 18,
        "G": 32,
        "H": 18,
        "I": 30,
        "J": 14,
        "K": 16,
        "L": 55,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    return wb


@staff_member_required
def admin_schools_with_supervisors_export_view(request: HttpRequest) -> HttpResponse:
    """
    تصدير الوضع الحالي للمدارس والمشرفين إلى Excel.

    يمكن استخدام الرابط مع:
    ?scope=all        جميع المدارس
    ?scope=assigned   المدارس المسندة فقط
    ?scope=unassigned المدارس غير المسندة فقط
    """
    scope = (request.GET.get("scope") or "all").strip().lower()

    if scope not in {"all", "assigned", "unassigned"}:
        scope = "all"

    wb = _build_schools_with_supervisors_export_workbook(scope=scope)

    filename = f"schools_with_supervisors_{scope}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ============================================================================
# View: Download Rejected Excel
# ============================================================================
@staff_member_required
def download_rejected_view(request: HttpRequest) -> HttpResponse:
    headers: list[str] = request.session.get(SESSION_REJ_HEADERS) or []
    rows: list[dict] = request.session.get(SESSION_REJ_ROWS) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "rejected"

    if not rows:
        ws.append(["لا توجد سجلات مرفوضة حالياً ✅"])
    else:
        if not headers:
            keys = set()
            for r in rows:
                keys |= set(r.keys())
            headers = ["_importer", "_reason"] + sorted(
                [k for k in keys if k not in {"_importer", "_reason"}]
            )

        ws.append(headers)

        for r in rows:
            ws.append([_norm(r.get(h)) for h in headers])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"rejected_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp