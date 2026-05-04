from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import Assignment, School, Sector, Supervisor


# =============================================================================
# Helpers
# =============================================================================
def _model_has_field(model, field_name: str) -> bool:
    return any(field.name == field_name for field in model._meta.get_fields())


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _digits(value: Any) -> str:
    return "".join(ch for ch in _cell(value) if ch.isdigit())


def _compact(value: Any) -> str:
    text = _cell(value).lower().replace("ـ", "")
    return re.sub(r"[\s_\-:/\\|().،,]+", "", text)


def _normalize_mobile(value: Any) -> str:
    d = _digits(value)
    if len(d) == 9 and d.startswith("5"):
        d = "0" + d
    return d


def _normalize_email(value: Any) -> str:
    return _cell(value).lower()


def _normalize_gender(value: Any) -> str:
    v = _compact(value)
    if not v:
        return ""

    boys = {"بنين", "ذكر", "ذكور", "boys", "boy", "male", "m"}
    girls = {"بنات", "انثى", "أنثى", "girls", "girl", "female", "f"}

    if v in {_compact(x) for x in boys}:
        return "boys"

    if v in {_compact(x) for x in girls}:
        return "girls"

    return ""


def _parse_bool(value: Any, *, default: bool = True) -> bool:
    v = _compact(value)
    if not v:
        return default

    true_values = {
        "1", "true", "yes", "on",
        "نشط", "مفعل", "فعال", "نعم",
    }

    false_values = {
        "0", "false", "no", "off",
        "غيرنشط", "غيرمفعل", "معطل", "لا",
    }

    if v in {_compact(x) for x in true_values}:
        return True

    if v in {_compact(x) for x in false_values}:
        return False

    return default


def _email_error(email: str) -> str:
    if not email:
        return ""

    try:
        validate_email(email)
    except ValidationError:
        return "صيغة البريد الإلكتروني غير صحيحة."

    return ""


# =============================================================================
# Headers
# =============================================================================
HEADER_ALIASES = {
    "school_stat_code": [
        "الرقم الإحصائي",
        "الرقم الاحصائي",
        "رقم المدرسة",
        "كود المدرسة",
        "stat_code",
        "school_code",
        "school_stat_code",
    ],
    "school_name": [
        "اسم المدرسة",
        "المدرسة",
        "school",
        "school_name",
    ],
    "school_gender": [
        "جنس المدرسة",
        "نوع المدرسة",
        "school_gender",
    ],
    "school_sector": [
        "قطاع المدرسة",
        "القطاع",
        "اسم القطاع",
        "sector",
    ],
    "school_active": [
        "حالة المدرسة",
        "نشاط المدرسة",
        "school_active",
        "school_is_active",
    ],
    "supervisor_national_id": [
        "سجل المشرف",
        "السجل المدني للمشرف",
        "هوية المشرف",
        "هوية المشرف المسند",
        "رقم هوية المشرف",
        "supervisor_national_id",
        "supervisor_nid",
        "national_id",
        "nid",
    ],
    "supervisor_name": [
        "اسم المشرف",
        "اسم المشرف المسند",
        "المشرف",
        "المشرف المسند",
        "supervisor",
        "supervisor_name",
    ],
    "supervisor_mobile": [
        "جوال المشرف",
        "رقم جوال المشرف",
        "الجوال",
        "mobile",
        "phone",
    ],
    "supervisor_email": [
        "بريد المشرف",
        "البريد الإلكتروني للمشرف",
        "البريد",
        "email",
    ],
    "supervisor_gender": [
        "جنس المشرف",
        "نوع المشرف",
        "supervisor_gender",
    ],
    "assignment_active": [
        "حالة الإسناد",
        "الإسناد",
        "assignment_active",
        "is_active",
    ],
    "notes": [
        "ملاحظات",
        "ملاحظة",
        "notes",
        "note",
    ],
}

NORMALIZED_HEADERS: dict[str, str] = {}
for key, aliases in HEADER_ALIASES.items():
    for alias in aliases:
        NORMALIZED_HEADERS[_compact(alias)] = key


def _detect_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}

    for idx, raw_header in enumerate(header_row):
        key = NORMALIZED_HEADERS.get(_compact(raw_header))
        if key and key not in header_map:
            header_map[key] = idx

    return header_map


def _get(row: tuple[Any, ...], header_map: dict[str, int], key: str) -> str:
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
        return ""
    return _cell(row[idx])


# =============================================================================
# Lookups
# =============================================================================
def _find_sector(value: str) -> Sector | None:
    value = _cell(value)

    if not value:
        return None

    if value.isdigit():
        return Sector.objects.filter(id=int(value)).first()

    return Sector.objects.filter(name__iexact=value).first()


def _get_or_create_sector(value: str) -> Sector | None:
    value = _cell(value)

    if not value:
        return None

    found = _find_sector(value)
    if found:
        return found

    if value.isdigit():
        return None

    defaults = {}
    if _model_has_field(Sector, "is_active"):
        defaults["is_active"] = True

    sector, _ = Sector.objects.get_or_create(name=value, defaults=defaults)
    return sector


def _find_school(
    *,
    stat_code: str,
    school_name: str,
    gender: str = "",
    sector_id: int | None = None,
) -> School | None:
    qs = School.objects.all()

    if stat_code and _model_has_field(School, "stat_code"):
        school = qs.filter(stat_code=stat_code).first()
        if school:
            return school

    if not school_name:
        return None

    qs = qs.filter(name__iexact=school_name)

    if gender in ("boys", "girls") and _model_has_field(School, "gender"):
        qs = qs.filter(gender=gender)

    if sector_id and _model_has_field(School, "sector"):
        qs = qs.filter(sector_id=sector_id)

    return qs.first()


def _find_supervisor_by_name(name: str) -> tuple[Supervisor | None, str]:
    name = _cell(name)

    if not name:
        return None, ""

    qs = Supervisor.objects.filter(full_name__iexact=name)

    count = qs.count()
    if count == 0:
        return None, ""

    if count > 1:
        return None, "اسم المشرف مكرر في قاعدة البيانات؛ يلزم إدخال سجل المشرف."

    return qs.first(), ""


def _school_unique_key(stat_code: str, school_name: str) -> str:
    if stat_code:
        return f"code:{stat_code}"
    return f"name:{_compact(school_name)}"


def _active_assignment_for_school(school: School) -> Assignment | None:
    return (
        Assignment.objects
        .filter(school=school, is_active=True)
        .select_related("supervisor")
        .first()
    )


# =============================================================================
# Dataclass
# =============================================================================
@dataclass
class SchoolSupervisorImportRow:
    row_no: int

    school_id: int | None
    school_stat_code: str
    school_name: str
    school_gender: str
    school_sector_id: int | None
    school_sector_name: str
    school_active: bool
    school_exists: bool

    supervisor_id: int | None
    supervisor_national_id: str
    supervisor_name: str
    supervisor_mobile: str
    supervisor_email: str
    supervisor_gender: str
    supervisor_exists: bool

    assignment_active: bool
    current_supervisor_name: str
    current_supervisor_national_id: str
    assigned_elsewhere: bool
    assignment_exists: bool

    action: str
    notes: str


def _empty_summary() -> dict[str, int]:
    return {
        "total_rows": 0,
        "valid_count": 0,
        "error_count": 0,
        "new_schools": 0,
        "updated_schools": 0,
        "new_supervisors": 0,
        "updated_supervisors": 0,
        "new_assignments": 0,
        "updated_assignments": 0,
        "transfer_assignments": 0,
        "blank_supervisor_rows": 0,
        "estimated_deactivations": 0,
        "conflicts": 0,
        "duplicates": 0,
    }


# =============================================================================
# Preview
# =============================================================================
def parse_schools_with_supervisors_workbook(
    uploaded_file,
    *,
    create_missing_schools: bool = False,
    update_schools: bool = True,
    create_missing_supervisors: bool = True,
    update_supervisors: bool = True,
    create_missing_sectors: bool = False,
    allow_transfer: bool = False,
    deactivate_when_supervisor_blank: bool = False,
) -> dict[str, Any]:
    try:
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception:
        return {
            "ok": False,
            "rows": [],
            "errors": [
                {
                    "row_no": "-",
                    "school_stat_code": "",
                    "school_name": "",
                    "supervisor": "",
                    "message": "تعذر قراءة ملف Excel. تأكد أن الملف بصيغة xlsx صحيحة.",
                }
            ],
            "summary": _empty_summary(),
        }

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {
            "ok": False,
            "rows": [],
            "errors": [
                {
                    "row_no": "-",
                    "school_stat_code": "",
                    "school_name": "",
                    "supervisor": "",
                    "message": "الملف فارغ.",
                }
            ],
            "summary": _empty_summary(),
        }

    header_map = _detect_header_map(header_row)

    missing_headers = []

    if "school_stat_code" not in header_map and "school_name" not in header_map:
        missing_headers.append("الرقم الإحصائي أو اسم المدرسة")

    if "supervisor_national_id" not in header_map and "supervisor_name" not in header_map:
        missing_headers.append("سجل المشرف أو اسم المشرف المسند")

    if missing_headers:
        return {
            "ok": False,
            "rows": [],
            "errors": [
                {
                    "row_no": 1,
                    "school_stat_code": "",
                    "school_name": "",
                    "supervisor": "",
                    "message": "الأعمدة الأساسية غير موجودة: " + "، ".join(missing_headers),
                }
            ],
            "summary": _empty_summary(),
        }

    errors: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []

    seen_schools: dict[str, int] = {}
    total_rows = 0

    for excel_row_no, row in enumerate(rows_iter, start=2):
        if not row or all(_cell(cell) == "" for cell in row):
            continue

        total_rows += 1
        row_errors: list[str] = []

        school_stat_code = _digits(_get(row, header_map, "school_stat_code"))
        school_name = _get(row, header_map, "school_name")
        school_gender = _normalize_gender(_get(row, header_map, "school_gender"))
        school_sector_value = _get(row, header_map, "school_sector")
        school_active = _parse_bool(_get(row, header_map, "school_active"), default=True)

        supervisor_national_id = _digits(_get(row, header_map, "supervisor_national_id"))
        supervisor_name = _get(row, header_map, "supervisor_name")
        supervisor_mobile = _normalize_mobile(_get(row, header_map, "supervisor_mobile"))
        supervisor_email = _normalize_email(_get(row, header_map, "supervisor_email"))
        supervisor_gender = _normalize_gender(_get(row, header_map, "supervisor_gender"))

        assignment_active = _parse_bool(_get(row, header_map, "assignment_active"), default=True)
        notes = _get(row, header_map, "notes")

        if not school_stat_code and not school_name:
            row_errors.append("يجب إدخال الرقم الإحصائي أو اسم المدرسة.")

        school_key = _school_unique_key(school_stat_code, school_name)
        previous_row = seen_schools.get(school_key)

        if previous_row:
            row_errors.append(f"المدرسة مكررة داخل الملف، سبق ظهورها في الصف {previous_row}.")
        else:
            seen_schools[school_key] = excel_row_no

        email_problem = _email_error(supervisor_email)
        if email_problem:
            row_errors.append(email_problem)

        if supervisor_national_id and len(supervisor_national_id) != 10:
            row_errors.append("سجل المشرف يجب أن يكون 10 أرقام.")

        sector = None
        sector_id = None
        sector_name = ""

        if school_sector_value:
            sector = _find_sector(school_sector_value)
            if sector:
                sector_id = sector.id
                sector_name = sector.name
            elif create_missing_sectors and not _cell(school_sector_value).isdigit():
                sector_name = _cell(school_sector_value)
            else:
                row_errors.append("القطاع غير موجود.")

        final_gender = school_gender or supervisor_gender

        school = _find_school(
            stat_code=school_stat_code,
            school_name=school_name,
            gender=final_gender,
            sector_id=sector_id,
        )

        if not school and not create_missing_schools:
            row_errors.append("المدرسة غير موجودة. استورد المدارس أولًا أو فعّل خيار إنشاء المدارس غير الموجودة.")

        if not school and create_missing_schools and not school_name:
            row_errors.append("اسم المدرسة مطلوب عند إنشاء مدرسة جديدة.")

        if school:
            if not school_name:
                school_name = getattr(school, "name", "") or ""

            if not school_stat_code and _model_has_field(School, "stat_code"):
                school_stat_code = getattr(school, "stat_code", "") or ""

            if not school_gender and _model_has_field(School, "gender"):
                school_gender = getattr(school, "gender", "") or ""

            if not sector_id and getattr(school, "sector_id", None):
                sector_id = school.sector_id
                sector_name = getattr(getattr(school, "sector", None), "name", "") or sector_name

        supervisor = None
        supervisor_lookup_error = ""

        has_supervisor_data = bool(supervisor_national_id or supervisor_name)

        if supervisor_national_id:
            supervisor = Supervisor.objects.filter(national_id=supervisor_national_id).first()
        elif supervisor_name:
            supervisor, supervisor_lookup_error = _find_supervisor_by_name(supervisor_name)

        if supervisor_lookup_error:
            row_errors.append(supervisor_lookup_error)

        if has_supervisor_data and not supervisor and not create_missing_supervisors:
            row_errors.append("المشرف غير موجود. فعّل خيار إنشاء المشرفين غير الموجودين أو صحح البيانات.")

        if has_supervisor_data and not supervisor and create_missing_supervisors:
            if not supervisor_national_id:
                row_errors.append("سجل المشرف مطلوب عند إنشاء مشرف جديد.")
            if not supervisor_name:
                row_errors.append("اسم المشرف مطلوب عند إنشاء مشرف جديد.")

        if supervisor:
            if not supervisor_name:
                supervisor_name = supervisor.full_name

            if not supervisor_national_id:
                supervisor_national_id = getattr(supervisor, "national_id", "") or ""

            if not supervisor_gender and _model_has_field(Supervisor, "gender"):
                supervisor_gender = getattr(supervisor, "gender", "") or ""

            if not supervisor_mobile and _model_has_field(Supervisor, "mobile"):
                supervisor_mobile = getattr(supervisor, "mobile", "") or ""

            if not supervisor_email and _model_has_field(Supervisor, "email"):
                supervisor_email = getattr(supervisor, "email", "") or ""

        if school and has_supervisor_data:
            db_school_gender = getattr(school, "gender", "") or ""
            check_supervisor_gender = supervisor_gender or school_gender

            if db_school_gender in ("boys", "girls") and check_supervisor_gender in ("boys", "girls"):
                if db_school_gender != check_supervisor_gender:
                    row_errors.append("جنس المدرسة لا يطابق جنس المشرف.")

        current_supervisor_name = ""
        current_supervisor_national_id = ""
        assigned_elsewhere = False
        assignment_exists = False
        estimated_deactivation = 0

        if school:
            current_assignment = _active_assignment_for_school(school)

            if current_assignment:
                current_supervisor = current_assignment.supervisor
                current_supervisor_name = getattr(current_supervisor, "full_name", "") or ""
                current_supervisor_national_id = getattr(current_supervisor, "national_id", "") or ""

            if supervisor:
                assignment_exists = Assignment.objects.filter(
                    school=school,
                    supervisor=supervisor,
                ).exists()

                assigned_elsewhere = (
                    Assignment.objects
                    .filter(school=school, is_active=True)
                    .exclude(supervisor=supervisor)
                    .exists()
                )

                if assigned_elsewhere and not allow_transfer:
                    row_errors.append(
                        "المدرسة مسندة حاليًا لمشرف آخر. فعّل خيار السماح بنقل الإسناد إذا رغبت بنقلها."
                    )

            if not has_supervisor_data and deactivate_when_supervisor_blank:
                estimated_deactivation = Assignment.objects.filter(school=school, is_active=True).count()

        if row_errors:
            errors.append(
                {
                    "row_no": excel_row_no,
                    "school_stat_code": school_stat_code,
                    "school_name": school_name,
                    "supervisor": supervisor_national_id or supervisor_name,
                    "message": " ".join(row_errors),
                }
            )
            continue

        if not has_supervisor_data:
            action = "deactivate_assignment" if deactivate_when_supervisor_blank else "no_assignment"
        elif not school:
            action = "create_school_and_assignment"
        elif not supervisor:
            action = "create_supervisor_and_assignment"
        elif assigned_elsewhere:
            action = "transfer_assignment"
        elif assignment_exists:
            action = "update_assignment"
        else:
            action = "create_assignment"

        item = SchoolSupervisorImportRow(
            row_no=excel_row_no,

            school_id=school.id if school else None,
            school_stat_code=school_stat_code,
            school_name=school_name,
            school_gender=school_gender or final_gender,
            school_sector_id=sector_id,
            school_sector_name=sector_name,
            school_active=school_active,
            school_exists=bool(school),

            supervisor_id=supervisor.id if supervisor else None,
            supervisor_national_id=supervisor_national_id,
            supervisor_name=supervisor_name,
            supervisor_mobile=supervisor_mobile,
            supervisor_email=supervisor_email,
            supervisor_gender=supervisor_gender or school_gender,
            supervisor_exists=bool(supervisor),

            assignment_active=assignment_active,
            current_supervisor_name=current_supervisor_name,
            current_supervisor_national_id=current_supervisor_national_id,
            assigned_elsewhere=assigned_elsewhere,
            assignment_exists=assignment_exists,

            action=action,
            notes=notes,
        )

        row_dict = asdict(item)
        row_dict["estimated_deactivation"] = estimated_deactivation
        valid_rows.append(row_dict)

    summary = {
        "total_rows": total_rows,
        "valid_count": len(valid_rows),
        "error_count": len(errors),
        "new_schools": sum(1 for r in valid_rows if not r["school_exists"]),
        "updated_schools": sum(1 for r in valid_rows if r["school_exists"]),
        "new_supervisors": sum(1 for r in valid_rows if r["supervisor_name"] and not r["supervisor_exists"]),
        "updated_supervisors": sum(1 for r in valid_rows if r["supervisor_exists"]),
        "new_assignments": sum(
            1 for r in valid_rows
            if r["action"] in ("create_assignment", "create_supervisor_and_assignment", "create_school_and_assignment")
        ),
        "updated_assignments": sum(1 for r in valid_rows if r["action"] == "update_assignment"),
        "transfer_assignments": sum(1 for r in valid_rows if r["action"] == "transfer_assignment"),
        "blank_supervisor_rows": sum(1 for r in valid_rows if r["action"] in ("no_assignment", "deactivate_assignment")),
        "estimated_deactivations": sum(int(r.get("estimated_deactivation") or 0) for r in valid_rows),
        "conflicts": sum(1 for e in errors if "مسندة حاليًا" in e["message"] or "مكرر" in e["message"]),
        "duplicates": sum(1 for e in errors if "مكررة" in e["message"]),
    }

    return {
        "ok": len(errors) == 0,
        "rows": valid_rows,
        "errors": errors,
        "summary": summary,
    }


# =============================================================================
# Commit
# =============================================================================
@transaction.atomic
def commit_schools_with_supervisors_import(
    rows: list[dict[str, Any]],
    *,
    create_missing_schools: bool = False,
    update_schools: bool = True,
    create_missing_supervisors: bool = True,
    update_supervisors: bool = True,
    create_missing_sectors: bool = False,
    allow_transfer: bool = False,
    deactivate_when_supervisor_blank: bool = False,
) -> dict[str, int]:
    created_schools = 0
    updated_schools = 0
    created_supervisors = 0
    updated_supervisors = 0
    created_assignments = 0
    updated_assignments = 0
    transferred_assignments = 0
    deactivated_assignments = 0
    no_action_rows = 0

    for row in rows:
        sector = None

        if row.get("school_sector_id"):
            sector = Sector.objects.filter(id=row["school_sector_id"]).first()

        if not sector and row.get("school_sector_name") and create_missing_sectors:
            sector = _get_or_create_sector(row["school_sector_name"])

        school = None

        if row.get("school_id"):
            school = School.objects.filter(id=row["school_id"]).first()

        if not school:
            school = _find_school(
                stat_code=row.get("school_stat_code", ""),
                school_name=row.get("school_name", ""),
                gender=row.get("school_gender", ""),
                sector_id=sector.id if sector else None,
            )

        if not school:
            if not create_missing_schools:
                no_action_rows += 1
                continue

            school_data = {
                "name": row["school_name"],
            }

            if _model_has_field(School, "stat_code"):
                school_data["stat_code"] = row.get("school_stat_code", "")

            if _model_has_field(School, "gender"):
                school_data["gender"] = row.get("school_gender") or "boys"

            if _model_has_field(School, "sector") and sector:
                school_data["sector"] = sector

            if _model_has_field(School, "is_active"):
                school_data["is_active"] = bool(row.get("school_active", True))

            school = School.objects.create(**school_data)
            created_schools += 1

        elif update_schools:
            changed_fields: list[str] = []

            if row.get("school_name") and getattr(school, "name", "") != row["school_name"]:
                school.name = row["school_name"]
                changed_fields.append("name")

            if _model_has_field(School, "stat_code") and row.get("school_stat_code"):
                if getattr(school, "stat_code", "") != row["school_stat_code"]:
                    school.stat_code = row["school_stat_code"]
                    changed_fields.append("stat_code")

            if _model_has_field(School, "gender") and row.get("school_gender"):
                if getattr(school, "gender", "") != row["school_gender"]:
                    school.gender = row["school_gender"]
                    changed_fields.append("gender")

            if _model_has_field(School, "sector") and sector:
                if getattr(school, "sector_id", None) != sector.id:
                    school.sector = sector
                    changed_fields.append("sector")

            if _model_has_field(School, "is_active"):
                new_active = bool(row.get("school_active", True))
                if getattr(school, "is_active", True) != new_active:
                    school.is_active = new_active
                    changed_fields.append("is_active")

            if changed_fields:
                school.save(update_fields=changed_fields)
                updated_schools += 1

        has_supervisor_data = bool(row.get("supervisor_national_id") or row.get("supervisor_name"))

        if not has_supervisor_data:
            if deactivate_when_supervisor_blank:
                count = Assignment.objects.filter(school=school, is_active=True).count()
                Assignment.objects.filter(school=school, is_active=True).update(is_active=False)
                deactivated_assignments += count
            else:
                no_action_rows += 1
            continue

        supervisor = None

        if row.get("supervisor_id"):
            supervisor = Supervisor.objects.filter(id=row["supervisor_id"]).first()

        if not supervisor and row.get("supervisor_national_id"):
            supervisor = Supervisor.objects.filter(national_id=row["supervisor_national_id"]).first()

        if not supervisor and row.get("supervisor_name"):
            supervisor, _ = _find_supervisor_by_name(row["supervisor_name"])

        if not supervisor:
            if not create_missing_supervisors:
                no_action_rows += 1
                continue

            supervisor = Supervisor(
                national_id=row["supervisor_national_id"],
                full_name=row["supervisor_name"],
            )
            created_supervisors += 1

        elif update_supervisors:
            updated_supervisors += 1

        if update_supervisors or created_supervisors:
            if row.get("supervisor_name"):
                supervisor.full_name = row["supervisor_name"]

            if _model_has_field(Supervisor, "mobile") and row.get("supervisor_mobile"):
                supervisor.mobile = row["supervisor_mobile"]

            if _model_has_field(Supervisor, "email") and row.get("supervisor_email"):
                old_email = getattr(supervisor, "email", "") or ""
                new_email = row["supervisor_email"]
                supervisor.email = new_email

                if old_email != new_email and _model_has_field(Supervisor, "email_verified"):
                    supervisor.email_verified = False

            if _model_has_field(Supervisor, "gender") and row.get("supervisor_gender"):
                supervisor.gender = row["supervisor_gender"]

            if _model_has_field(Supervisor, "sector") and sector:
                supervisor.sector = sector

            if _model_has_field(Supervisor, "is_active"):
                supervisor.is_active = True

            supervisor.save()

        if allow_transfer:
            old_assignments = (
                Assignment.objects
                .filter(school=school, is_active=True)
                .exclude(supervisor=supervisor)
            )
            count = old_assignments.count()
            if count:
                old_assignments.update(is_active=False)
                transferred_assignments += count

        assignment, created = Assignment.objects.get_or_create(
            school=school,
            supervisor=supervisor,
            defaults={"is_active": bool(row.get("assignment_active", True))},
        )

        if created:
            created_assignments += 1
        else:
            assignment.is_active = bool(row.get("assignment_active", True))
            assignment.save(update_fields=["is_active"])
            updated_assignments += 1

    return {
        "created_schools": created_schools,
        "updated_schools": updated_schools,
        "created_supervisors": created_supervisors,
        "updated_supervisors": updated_supervisors,
        "created_assignments": created_assignments,
        "updated_assignments": updated_assignments,
        "transferred_assignments": transferred_assignments,
        "deactivated_assignments": deactivated_assignments,
        "no_action_rows": no_action_rows,
        "total_assignment_actions": created_assignments + updated_assignments + transferred_assignments + deactivated_assignments,
    }


# =============================================================================
# Template workbook
# =============================================================================
def build_schools_with_supervisors_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "المدارس والإسناد"
    ws.sheet_view.rightToLeft = True

    headers = [
        "الرقم الإحصائي",
        "اسم المدرسة",
        "جنس المدرسة",
        "القطاع",
        "حالة المدرسة",
        "سجل المشرف",
        "اسم المشرف المسند",
        "جوال المشرف",
        "بريد المشرف",
        "جنس المشرف",
        "حالة الإسناد",
        "ملاحظات",
    ]

    example = [
        "70228",
        "ابتدائية أبو أيوب الأنصاري بمحايل عسير",
        "بنين",
        "محايل عسير",
        "نشط",
        "1020103717",
        "أحمد محمد عسيري",
        "0550000000",
        "example@moe.gov.sa",
        "بنين",
        "نشط",
        "",
    ]

    title_font = Font(name="Cairo", bold=True, size=14)
    header_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="E8F5E9")
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    example_fill = PatternFill("solid", fgColor="FFFDF5")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value="قالب استيراد المدارس وإسنادها للمشرفين")
    cell.font = title_font
    cell.alignment = center
    cell.fill = title_fill

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = normal_font
        cell.alignment = center if col in (1, 3, 5, 6, 8, 10, 11) else right
        cell.fill = example_fill
        cell.border = border

    widths = {
        1: 18,
        2: 42,
        3: 14,
        4: 24,
        5: 14,
        6: 18,
        7: 30,
        8: 18,
        9: 28,
        10: 14,
        11: 16,
        12: 30,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"

    return wb