from __future__ import annotations

from datetime import datetime
from typing import Any

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Assignment, Plan, PlanConfig, PlanDay, School, Supervisor
from .utils_dates import week_rows


def _get_config() -> PlanConfig:
    cfg = PlanConfig.objects.order_by("-id").first()
    if not cfg:
        raise RuntimeError("PlanConfig غير موجود. أنشئ سجل واحد من لوحة الإدارة.")
    return cfg


def _get_supervisor(request: HttpRequest) -> Supervisor:
    """
    حالياً: نربط المشرف بالمستخدم عبر إدخال رقم سجله المدني في querystring:
    /plan/?sup=1036...&week=1

    لاحقاً نعمل Login خاص بالمشرف.
    """
    sup_id = (request.GET.get("sup") or "").strip()
    if not sup_id:
        raise RuntimeError("أضف ?sup=السجل_المدني_للمشرف في الرابط مؤقتاً.")
    sup = Supervisor.objects.filter(national_id=sup_id, is_active=True).first()
    if not sup:
        raise RuntimeError("المشرف غير موجود أو غير نشط.")
    return sup


@transaction.atomic
def plan_view(request: HttpRequest) -> HttpResponse:
    cfg = _get_config()
    week_no = int(request.GET.get("week", "1") or "1")
    week_no = max(1, min(cfg.weeks_count, week_no))

    # مؤقتاً جلب المشرف من sup=
    try:
        supervisor = _get_supervisor(request)
    except RuntimeError as e:
        return HttpResponse(f"<h3 dir='rtl'>{e}</h3>")

    plan, _ = Plan.objects.get_or_create(supervisor=supervisor, week_no=week_no)

    # ضمان وجود 5 سجلات للأيام
    existing = {d.weekday: d for d in plan.days.all()}
    for i in range(5):
        if i not in existing:
            PlanDay.objects.create(plan=plan, weekday=i)

    plan.refresh_from_db()
    days_map = {d.weekday: d for d in plan.days.all()}

    # مدارس المشرف المسندة
    assigned_school_ids = Assignment.objects.filter(supervisor=supervisor, is_active=True).values_list("school_id", flat=True)
    schools = School.objects.filter(id__in=assigned_school_ids, is_active=True).order_by("name")

    rows = week_rows(cfg.start_week1_sunday, week_no)

    if request.method == "POST":
        action = request.POST.get("action", "save")
        if plan.status == Plan.STATUS_APPROVED:
            messages.error(request, "الخطة معتمدة ولا يمكن تعديلها. اطلب فك الاعتماد أولاً.")
            return redirect(f"/plan/?sup={supervisor.national_id}&week={week_no}")

        # حفظ اختيار المدرسة لكل يوم
        for r in rows:
            key = f"school_{r['weekday']}"
            val = (request.POST.get(key) or "").strip()
            day_obj = days_map[r["weekday"]]

            if not val:
                day_obj.school = None
                day_obj.save(update_fields=["school"])
                continue

            # تحقق المدرسة ضمن مدارس المشرف
            school = School.objects.filter(id=val, id__in=assigned_school_ids).first()
            if not school:
                messages.error(request, f"اختيار مدرسة غير صحيح لليوم {r['weekday_name']}.")
                return redirect(f"/plan/?sup={supervisor.national_id}&week={week_no}")

            day_obj.school = school
            day_obj.save(update_fields=["school"])

        plan.saved_at = timezone.now()
        plan.status = Plan.STATUS_DRAFT
        plan.save(update_fields=["saved_at", "status"])

        if action == "approve":
            # شرط الاعتماد: تعبئة جميع الأيام
            missing = [r["weekday_name"] for r in rows if not days_map[r["weekday"]].school_id]
            # (days_map قديم قبل الحفظ، نعيد القراءة)
            plan.refresh_from_db()
            days_map2 = {d.weekday: d for d in plan.days.all()}
            missing = [r["weekday_name"] for r in rows if not days_map2[r["weekday"]].school_id]

            if missing:
                messages.error(request, "لا يمكن اعتماد الخطة. الأيام الناقصة: " + "، ".join(missing))
                return redirect(f"/plan/?sup={supervisor.national_id}&week={week_no}")

            plan.status = Plan.STATUS_APPROVED
            plan.approved_at = timezone.now()
            plan.save(update_fields=["status", "approved_at"])
            messages.success(request, "تم اعتماد الخطة بنجاح.")
        else:
            messages.success(request, "تم حفظ الخطة.")

        return redirect(f"/plan/?sup={supervisor.national_id}&week={week_no}")

    # عرض
    context = {
        "cfg": cfg,
        "week_no": week_no,
        "supervisor": supervisor,
        "plan": plan,
        "rows": rows,
        "days_map": days_map,
        "schools": schools,
    }
    return render(request, "visits/plan.html", context)


def export_plan_excel(request: HttpRequest) -> HttpResponse:
    cfg = _get_config()
    week_no = int(request.GET.get("week", "1") or "1")
    week_no = max(1, min(cfg.weeks_count, week_no))

    try:
        supervisor = _get_supervisor(request)
    except RuntimeError as e:
        return HttpResponse(f"<h3 dir='rtl'>{e}</h3>")

    plan = get_object_or_404(Plan, supervisor=supervisor, week_no=week_no)
    rows = week_rows(cfg.start_week1_sunday, week_no)
    days_map = {d.weekday: d for d in plan.days.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {week_no}"

    headers = ["اليوم", "التاريخ (هجري)", "التاريخ (ميلادي)", "اسم المدرسة", "الرقم الإحصائي"]
    ws.append(headers)

    for r in rows:
        day = days_map.get(r["weekday"])
        school_name = day.school.name if day and day.school else ""
        stat_code = day.school.stat_code if day and day.school else ""
        ws.append([r["weekday_name"], r["hijri_date"], r["greg_date"].isoformat(), school_name, stat_code])

    # تحسين عرض الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    filename = f"خطة_الأسبوع_{week_no}_{supervisor.full_name}.xlsx".replace("/", "-").replace("\\", "-")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
