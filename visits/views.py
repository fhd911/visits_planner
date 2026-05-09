from __future__ import annotations

from dataclasses import dataclass
import re
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from typing import Any, Optional

from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.core.mail import EmailMultiAlternatives
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from random import randint

from .forms import WeeklyLetterLinkForm
from .models import (
    Assignment,
    EmailOTP,
    Plan,
    PlanDay,
    PlanWeek,
    Principal,
    School,
    Sector,
    SiteSetting,
    Supervisor,
    SupervisorNotification,
    UnlockRequest,
    WeeklyLetterLink,
)
from .services.drive_service import list_files_in_folder


# =============================================================================
# Constants
# =============================================================================
WEEKDAYS = [
    (0, "الأحد"),
    (1, "الإثنين"),
    (2, "الثلاثاء"),
    (3, "الأربعاء"),
    (4, "الخميس"),
]
WEEKDAY_MAP = dict(WEEKDAYS)
SESSION_SUP_ID = "visits_sup_id"
EMAIL_OTP_SESSION_KEY = "visits_email_otp"
EMAIL_OTP_EXPIRE_MINUTES = 10


# =============================================================================
# Generic helpers
# =============================================================================
def _digits(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())


def _safe_int(v: object, default: int = 1) -> int:
    try:
        return int(str(v).strip())
    except Exception:
        return default


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_ajax(request: HttpRequest) -> bool:
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _bool_from_post(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in ("1", "true", "yes", "on")


def _get_mobile_digits(sup: Supervisor) -> str:
    return _digits(getattr(sup, "mobile", "") or "")


def _supervisor_last4(sup: Supervisor) -> Optional[str]:
    d = _get_mobile_digits(sup)
    return d[-4:] if len(d) >= 4 else None


def _get_logged_in_supervisor(request: HttpRequest) -> Optional[Supervisor]:
    sid = request.session.get(SESSION_SUP_ID)
    if not sid:
        return None
    try:
        sid_int = int(sid)
    except Exception:
        return None
    return Supervisor.objects.filter(id=sid_int, is_active=True).first()


def _require_supervisor(request: HttpRequest) -> Supervisor:
    sup = _get_logged_in_supervisor(request)
    if not sup:
        raise Supervisor.DoesNotExist
    return sup


def _find_supervisor_by_nid(nid: str) -> Optional[Supervisor]:
    nid = _digits(nid)
    if not nid:
        return None
    return Supervisor.objects.filter(national_id=nid, is_active=True).first()


def _sup_nid_value(sup: Supervisor) -> str:
    return getattr(sup, "national_id", "") or ""


def _inject_week_no(plan: Plan) -> None:
    try:
        plan.week_no = plan.week.week_no  # type: ignore[attr-defined]
    except Exception:
        plan.week_no = None  # type: ignore[attr-defined]


def _supervisor_school_ids(supervisor: Supervisor) -> set[int]:
    return set(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
        ).values_list("school_id", flat=True)
    )


def _supervisor_schools_qs(supervisor: Supervisor):
    return (
        School.objects.filter(
            assignments__supervisor=supervisor,
            assignments__is_active=True,
            is_active=True,
        )
        .distinct()
        .order_by("name")
    )


def _supervisor_email_value(supervisor: Supervisor) -> str:
    return (getattr(supervisor, "email", "") or "").strip()


def _supervisor_needs_email_prompt(supervisor: Supervisor) -> bool:
    return not bool(_supervisor_email_value(supervisor))


def _supervisor_email_notifications_enabled(supervisor: Supervisor) -> bool:
    return bool(getattr(supervisor, "email_notifications_enabled", False))


def _generate_email_otp() -> str:
    return f"{randint(0, 999999):06d}"


def _email_otp_payload(request: HttpRequest) -> dict[str, Any]:
    return request.session.get(EMAIL_OTP_SESSION_KEY, {}) or {}


def _clear_email_otp_payload(request: HttpRequest) -> None:
    request.session.pop(EMAIL_OTP_SESSION_KEY, None)
    request.session.modified = True


def _set_email_otp_payload(
    request: HttpRequest,
    *,
    supervisor_id: int,
    email: str,
    email_notifications_enabled: bool,
    next_url: str,
    code: str,
) -> None:
    request.session[EMAIL_OTP_SESSION_KEY] = {
        "supervisor_id": supervisor_id,
        "email": email,
        "email_notifications_enabled": bool(email_notifications_enabled),
        "next_url": next_url,
        "code": code,
        "expires_at": (timezone.now() + timedelta(minutes=EMAIL_OTP_EXPIRE_MINUTES)).isoformat(),
    }
    request.session.modified = True


def _email_otp_is_valid(payload: dict[str, Any]) -> bool:
    expires_at = payload.get("expires_at")
    if not expires_at:
        return False
    try:
        dt = datetime.fromisoformat(expires_at)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
    except Exception:
        return False
    return timezone.now() <= dt


def _send_supervisor_email_otp(*, email: str, code: str, supervisor_name: str) -> None:
    subject = "رمز التحقق لتأكيد البريد الإلكتروني"

    plain_message = (
        f"السلام عليكم ورحمة الله وبركاته {supervisor_name}\n\n"
        f"نفيدكم بأنه تم طلب تأكيد البريد الإلكتروني المرتبط بحسابكم في بوابة الزيارات.\n\n"
        f"رمز التحقق الخاص بكم هو:\n"
        f"{code}\n\n"
        f"صلاحية الرمز: {EMAIL_OTP_EXPIRE_MINUTES} دقائق.\n\n"
        f"في حال لم يكن هذا الطلب صادرًا منكم، يرجى تجاهل هذه الرسالة.\n\n"
        f"مع خالص التحية والتقدير،\n"
        f"بوابة الزيارات"
    )

    html_message = f"""
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>{subject}</title>
    </head>
    <body style="margin:0;padding:0;background:#f2f5f7;font-family:'Cairo','Tajawal','Tahoma',Arial,sans-serif;color:#1f2937;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background:#f2f5f7;margin:0;padding:28px 0;">
        <tr>
          <td align="center">
            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="max-width:760px;background:#ffffff;border:1px solid #d8e1e6;border-radius:24px;overflow:hidden;box-shadow:0 16px 38px rgba(15,23,42,.08);">

              <tr>
                <td style="background:#175c49;padding:20px 34px 10px;text-align:right;">
                  <div style="font-size:13px;line-height:1.9;color:rgba(255,255,255,.92);font-weight:700;">
                    بوابة الزيارات
                  </div>
                </td>
              </tr>

              <tr>
                <td style="background:#175c49;padding:0 34px 18px;text-align:right;">
                  <div style="height:4px;background:linear-gradient(90deg,#caa85e,#f2dfb1,#caa85e);border-radius:999px;"></div>
                </td>
              </tr>

              <tr>
                <td style="background:#f8fbfa;padding:22px 34px 18px;border-bottom:1px solid #e5ece8;text-align:right;">
                  <div style="font-size:28px;line-height:1.7;color:#184c3c;font-weight:900;">
                    رمز التحقق لتأكيد البريد الإلكتروني
                  </div>
                  <div style="margin-top:8px;font-size:14px;line-height:2;color:#5f6b76;font-weight:600;">
                    إشعار آلي خاص بتأكيد البريد الإلكتروني وتفعيل التنبيهات المرتبطة بالخطة الأسبوعية والإشعارات الإدارية.
                  </div>
                </td>
              </tr>

              <tr>
                <td style="padding:34px 34px 10px;text-align:right;">
                  <p style="margin:0 0 14px;font-size:17px;line-height:2.1;color:#111827;font-weight:800;">
                    السلام عليكم ورحمة الله وبركاته
                  </p>

                  <p style="margin:0 0 10px;font-size:15px;line-height:2.1;color:#374151;">
                    الأستاذ/ <strong>{supervisor_name}</strong>
                  </p>

                  <p style="margin:0 0 20px;font-size:15px;line-height:2.2;color:#4b5563;">
                    نفيدكم بأنه تم طلب تأكيد البريد الإلكتروني المرتبط بحسابكم في
                    <strong style="color:#184c3c;">بوابة الزيارات</strong>،
                    وعليه نأمل استخدام رمز التحقق التالي لإتمام عملية التأكيد.
                  </p>

                  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="margin:26px 0 24px;">
                    <tr>
                      <td align="center">
                        <div style="display:inline-block;min-width:340px;background:linear-gradient(180deg,#fffdf8,#f8f4ea);border:1px solid #dcc79a;border-radius:24px;padding:18px 26px;box-shadow:0 8px 24px rgba(180,140,60,.08), inset 0 0 0 1px rgba(255,255,255,.7);">
                          <div style="font-size:13px;line-height:1.8;color:#8a6a2f;font-weight:800;margin-bottom:10px;">
                            رمز التحقق
                          </div>
                          <div style="font-size:38px;line-height:1;letter-spacing:12px;color:#184c3c;font-weight:900;direction:ltr;text-shadow:0 1px 0 rgba(255,255,255,.7);">
                            {code}
                          </div>
                        </div>
                      </td>
                    </tr>
                  </table>

                  <div style="background:#f8fafc;border:1px solid #e5e7eb;border-radius:15px;padding:14px 16px;margin:0 0 16px;">
                    <p style="margin:0;font-size:14px;line-height:2;color:#374151;">
                      <strong>مدة صلاحية الرمز:</strong> {EMAIL_OTP_EXPIRE_MINUTES} دقائق
                    </p>
                  </div>

                  <div style="background:#fffaf3;border:1px solid #ead7b0;border-radius:15px;padding:14px 16px;margin:0 0 18px;">
                    <p style="margin:0;font-size:14px;line-height:2.1;color:#8a5a16;">
                      في حال لم يكن هذا الطلب صادرًا منكم، يرجى تجاهل هذه الرسالة، ولن يتم اعتماد البريد الإلكتروني دون إدخال الرمز الصحيح.
                    </p>
                  </div>
                </td>
              </tr>

              <tr>
                <td style="padding:0 34px 28px;text-align:right;">
                  <div style="border-top:1px solid #e6ecef;padding-top:18px;">
                    <div style="font-size:14px;line-height:2;color:#374151;font-weight:800;">
                      مع خالص التحية والتقدير
                    </div>
                    <div style="font-size:13px;line-height:2;color:#6b7280;">
                      بوابة الزيارات
                    </div>
                  </div>
                </td>
              </tr>

              <tr>
                <td style="background:#f8fbfa;border-top:1px solid #e4ece8;padding:14px 24px;text-align:center;">
                  <div style="font-size:12px;line-height:1.9;color:#7b8794;">
                    هذه رسالة آلية صادرة من بوابة الزيارات، يرجى عدم الرد عليها.
                  </div>
                </td>
              </tr>

            </table>
          </td>
        </tr>
      </table>
    </body>
    </html>
    """

    msg = EmailMultiAlternatives(
        subject=subject,
        body=plain_message,
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=[email],
    )
    msg.attach_alternative(html_message, "text/html")
    msg.send(fail_silently=False)


# =============================================================================
# Admin/supervisor isolation
# =============================================================================
def _has_supervisor_session(request: HttpRequest) -> bool:
    return bool(request.session.get(SESSION_SUP_ID))


def _redirect_supervisor_from_admin(request: HttpRequest) -> HttpResponse | None:
    if _has_supervisor_session(request):
        messages.warning(request, "ليس لديك صلاحية لدخول صفحات الإدارة أثناء تسجيلك كمشرف.")
        week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
        return redirect(_plan_url(week_no))
    return None


def admin_only_view(view_func):
    @wraps(view_func)
    def _wrapped(request: HttpRequest, *args, **kwargs):
        supervisor_redirect = _redirect_supervisor_from_admin(request)
        if supervisor_redirect:
            return supervisor_redirect
        return staff_member_required(view_func)(request, *args, **kwargs)
    return _wrapped


# =============================================================================
# Site maintenance helpers
# =============================================================================
def _get_site_setting() -> SiteSetting:
    return SiteSetting.get_solo()


def _maintenance_message(setting: SiteSetting) -> str:
    return (
        setting.maintenance_message
        or "الموقع مغلق مؤقتًا للصيانة، وسيعود العمل خلال وقت قريب."
    )


def _parse_dt_local(value: str) -> Optional[datetime]:
    value = (value or "").strip()
    if not value:
        return None

    fmts = (
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    )
    for fmt in fmts:
        try:
            dt = datetime.strptime(value, fmt)
            return timezone.make_aware(dt, timezone.get_current_timezone())
        except Exception:
            continue
    return None


def _format_dt_local(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%dT%H:%M")


def _dt_iso_local(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return timezone.localtime(value).isoformat()


def _format_dt_ar_pretty(value: Optional[datetime]) -> str:
    if not value:
        return ""

    dt = timezone.localtime(value)
    weekdays = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
    months = [
        "يناير",
        "فبراير",
        "مارس",
        "أبريل",
        "مايو",
        "يونيو",
        "يوليو",
        "أغسطس",
        "سبتمبر",
        "أكتوبر",
        "نوفمبر",
        "ديسمبر",
    ]

    weekday_name = weekdays[dt.weekday()]
    month_name = months[dt.month - 1]
    hour12 = dt.hour % 12 or 12
    ampm = "ص" if dt.hour < 12 else "م"
    return f"{weekday_name} {dt.day} {month_name} {dt.year} — {hour12:02d}:{dt.minute:02d} {ampm}"


def _visit_type_export_label(value: str) -> str:
    if value == getattr(PlanDay, "VISIT_IN", "in"):
        return "حضوري"
    if value == getattr(PlanDay, "VISIT_REMOTE", "remote"):
        return "عن بعد"
    if value == getattr(PlanDay, "VISIT_NONE", "none"):
        return "بدون زيارة"
    return value or "—"


def _maintenance_is_active(setting: SiteSetting, *, persist: bool = False) -> bool:
    active = bool(setting.is_maintenance_mode)
    now = timezone.now()

    starts_at = getattr(setting, "maintenance_starts_at", None)
    ends_at = getattr(setting, "maintenance_ends_at", None)

    if not active:
        return False

    if starts_at and now < starts_at:
        return False

    if ends_at and now >= ends_at:
        if persist:
            setting.is_maintenance_mode = False
            setting.save(update_fields=["is_maintenance_mode", "updated_at"])
        return False

    return True


def _maintenance_context() -> dict[str, Any]:
    setting = _get_site_setting()
    active = _maintenance_is_active(setting, persist=True)
    starts_at = getattr(setting, "maintenance_starts_at", None)
    ends_at = getattr(setting, "maintenance_ends_at", None)
    now_dt = timezone.localtime(timezone.now())

    return {
        "site_setting": setting,
        "maintenance_message": _maintenance_message(setting),
        "expected_return_text": setting.expected_return_text or "",
        "allow_admin_only": setting.allow_admin_only,
        "maintenance_is_active": active,
        "maintenance_starts_at": starts_at,
        "maintenance_ends_at": ends_at,
        "maintenance_starts_at_value": _format_dt_local(starts_at),
        "maintenance_ends_at_value": _format_dt_local(ends_at),
        "maintenance_starts_at_iso": _dt_iso_local(starts_at),
        "maintenance_ends_at_iso": _dt_iso_local(ends_at),
        "maintenance_starts_at_gregorian": _format_dt_ar_pretty(starts_at),
        "maintenance_ends_at_gregorian": _format_dt_ar_pretty(ends_at),
        "maintenance_now": now_dt,
        "maintenance_now_iso": now_dt.isoformat(),
        "maintenance_now_gregorian": _format_dt_ar_pretty(now_dt),
        "maintenance_window_label": getattr(setting, "maintenance_window_label", "غير محدد"),
    }


def _is_admin_user(request: HttpRequest) -> bool:
    user = getattr(request, "user", None)
    return bool(user and user.is_authenticated and user.is_staff)


def _maintenance_allowed_for_request(request: HttpRequest, setting: SiteSetting) -> bool:
    if not _maintenance_is_active(setting, persist=True):
        return True

    if _is_admin_user(request):
        return True

    if request.path.startswith("/admin/"):
        return True

    return False


# =============================================================================
# Maintenance views
# =============================================================================
def maintenance_page_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if not _maintenance_is_active(setting, persist=True):
        if request.user.is_authenticated and request.user.is_staff:
            return redirect("visits:admin_dashboard")
        return redirect("visits:login")

    context = _maintenance_context()
    context.update(
        {
            "page_title": (getattr(setting, "site_name", "") or "بوابة الزيارات").strip() or "بوابة الزيارات",
            "page_subtitle": "الخدمة متوقفة مؤقتًا حسب الفترة المحددة من الإدارة.",
            "maintenance_has_window": bool(context.get("maintenance_starts_at") or context.get("maintenance_ends_at")),
            "maintenance_period_text": getattr(setting, "maintenance_window_label", "") or "",
        }
    )
    return render(request, "visits/maintenance.html", context)


@admin_only_view
def admin_maintenance_settings_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    _maintenance_is_active(setting, persist=True)
    return render(
        request,
        "visits/admin_maintenance_settings.html",
        {
            "site_setting": setting,
            "maintenance_starts_at_value": _format_dt_local(getattr(setting, "maintenance_starts_at", None)),
            "maintenance_ends_at_value": _format_dt_local(getattr(setting, "maintenance_ends_at", None)),
            "maintenance_window_label": getattr(setting, "maintenance_window_label", "غير محدد"),
            "maintenance_is_active": _maintenance_is_active(setting, persist=False),
        },
    )


@admin_only_view
@require_POST
def admin_toggle_maintenance_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()

    enable_raw = (
        request.POST.get("enable")
        or request.POST.get("is_maintenance_mode")
        or request.POST.get("maintenance")
        or ""
    ).strip().lower()

    if enable_raw in ("1", "true", "yes", "on", "enable"):
        setting.is_maintenance_mode = True
        msg = "تم تفعيل وضع الصيانة بنجاح."
    elif enable_raw in ("0", "false", "no", "off", "disable"):
        setting.is_maintenance_mode = False
        msg = "تم إيقاف وضع الصيانة بنجاح."
    else:
        setting.is_maintenance_mode = not setting.is_maintenance_mode
        msg = "تم تحديث حالة وضع الصيانة بنجاح."

    setting.allow_admin_only = _bool_from_post(
        request.POST.get("allow_admin_only"),
        default=setting.allow_admin_only,
    )
    setting.save()
    active = _maintenance_is_active(setting, persist=True)

    if _is_ajax(request):
        return JsonResponse(
            {
                "ok": True,
                "message": msg,
                "is_maintenance_mode": setting.is_maintenance_mode,
                "maintenance_is_active": active,
                "allow_admin_only": setting.allow_admin_only,
                "maintenance_window_label": getattr(setting, "maintenance_window_label", "غير محدد"),
            },
            status=200,
        )

    messages.success(request, msg)
    return redirect("visits:admin_maintenance_settings")


@admin_only_view
@require_POST
def admin_update_maintenance_message_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()

    starts_at_raw = (
        request.POST.get("maintenance_starts_at")
        or request.POST.get("starts_at")
        or ""
    )
    ends_at_raw = (
        request.POST.get("maintenance_ends_at")
        or request.POST.get("ends_at")
        or ""
    )

    starts_at = _parse_dt_local(starts_at_raw)
    ends_at = _parse_dt_local(ends_at_raw)

    if starts_at_raw.strip() and starts_at is None:
        messages.error(request, "صيغة تاريخ بداية الصيانة غير صحيحة.")
        return redirect("visits:admin_maintenance_settings")

    if ends_at_raw.strip() and ends_at is None:
        messages.error(request, "صيغة تاريخ نهاية الصيانة غير صحيحة.")
        return redirect("visits:admin_maintenance_settings")

    setting.site_name = (request.POST.get("site_name") or setting.site_name or "بوابة الزيارات").strip()
    setting.maintenance_message = (
        request.POST.get("maintenance_message")
        or request.POST.get("message")
        or ""
    ).strip() or None
    setting.expected_return_text = (
        request.POST.get("expected_return_text")
        or request.POST.get("expected_return")
        or ""
    ).strip() or None
    setting.maintenance_starts_at = starts_at
    setting.maintenance_ends_at = ends_at
    setting.allow_admin_only = _bool_from_post(
        request.POST.get("allow_admin_only"),
        default=setting.allow_admin_only,
    )
    setting.save()
    active = _maintenance_is_active(setting, persist=True)

    msg = "تم تحديث رسالة وإعدادات الصيانة بنجاح."

    if _is_ajax(request):
        return JsonResponse(
            {
                "ok": True,
                "message": msg,
                "site_name": setting.site_name,
                "maintenance_message": setting.maintenance_message or "",
                "expected_return_text": setting.expected_return_text or "",
                "allow_admin_only": setting.allow_admin_only,
                "is_maintenance_mode": setting.is_maintenance_mode,
                "maintenance_is_active": active,
                "maintenance_starts_at": _format_dt_local(setting.maintenance_starts_at),
                "maintenance_ends_at": _format_dt_local(setting.maintenance_ends_at),
                "maintenance_window_label": getattr(setting, "maintenance_window_label", "غير محدد"),
            },
            status=200,
        )

    messages.success(request, msg)
    return redirect("visits:admin_maintenance_settings")


# =============================================================================
# Google Drive helpers
# =============================================================================
def _extract_drive_folder_id(url: str) -> str | None:
    if not url:
        return None

    url = url.strip()
    patterns = [
        r"/folders/([a-zA-Z0-9_-]+)",
        r"[?&]id=([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    return None


def _find_supervisor_letter_in_folder(*, folder_id: str, national_id: str) -> str | None:
    files = list_files_in_folder(folder_id, page_size=200)

    national_id = _digits(national_id)
    if not national_id:
        return None

    for file_obj in files:
        name = (file_obj.get("name") or "").strip()
        if national_id in name:
            return file_obj.get("webViewLink")

    return None


# =============================================================================
# Week helpers
# =============================================================================
def _get_active_weeks_qs():
    return PlanWeek.objects.filter(is_break=False).order_by("week_no")


def _get_all_weeks_qs():
    return PlanWeek.objects.all().order_by("week_no")


def _get_default_week_no() -> int:
    w = _get_active_weeks_qs().first()
    return w.week_no if w else 1


def _week_exists(week_no: int) -> bool:
    return PlanWeek.objects.filter(week_no=week_no).exists()


def _resolve_week_or_404(week_no: int, *, allow_inactive: bool = False) -> PlanWeek:
    qs = PlanWeek.objects.all()
    if not allow_inactive:
        qs = qs.filter(is_break=False)
    return get_object_or_404(qs, week_no=week_no)


def _build_week_choices(active_only: bool = True) -> list[tuple[int, str]]:
    qs = _get_active_weeks_qs() if active_only else _get_all_weeks_qs()
    out: list[tuple[int, str]] = []
    for w in qs:
        label = f"الأسبوع {w.week_no}"
        if getattr(w, "title", ""):
            label += f" — {w.title}"
        if getattr(w, "is_break", False):
            label += " (إجازة)"
        out.append((w.week_no, label))
    return out


def _build_day_dates_from_week(week_obj: PlanWeek) -> dict[int, date]:
    start = week_obj.start_sunday
    return {
        0: start,
        1: start + timedelta(days=1),
        2: start + timedelta(days=2),
        3: start + timedelta(days=3),
        4: start + timedelta(days=4),
    }


def _current_week_obj() -> Optional[PlanWeek]:
    today = timezone.localdate()
    return (
        PlanWeek.objects.filter(start_sunday__lte=today, is_break=False)
        .order_by("-start_sunday", "-week_no")
        .first()
    )


def _plan_url(week_no: int) -> str:
    return f"{reverse('visits:plan')}?week={week_no}"


def _notifications_url(week_no: int | None = None) -> str:
    url = reverse("visits:notifications")
    return f"{url}?week={week_no}" if week_no else url


def _supervisor_visit_status_url(week_no: int) -> str:
    return f"{reverse('visits:supervisor_visit_status')}?week={week_no}"


def _admin_dashboard_url(
    week_no: int,
    *,
    show_all: bool = False,
    q: str = "",
    status: str = "all",
    ps: int | None = None,
    page: int | None = None,
) -> str:
    params = [f"week={week_no}"]
    if show_all:
        params.append("all=1")
    if q:
        params.append(f"q={q}")
    if status and status != "all":
        params.append(f"status={status}")
    if ps:
        params.append(f"ps={ps}")
    if page:
        params.append(f"page={page}")
    return f"{reverse('visits:admin_dashboard')}?{'&'.join(params)}"


def _admin_plan_detail_url(plan_id: int, *, week_no: int | None = None, next_url: str = "") -> str:
    url = reverse("visits:admin_plan_detail", args=[plan_id])
    params = []
    if week_no:
        params.append(f"week={week_no}")
    if next_url:
        params.append(f"next={next_url}")
    return f"{url}?{'&'.join(params)}" if params else url


def _resolve_admin_return_url(
    request: HttpRequest,
    *,
    plan: Plan,
    default_week_no: int,
    show_all: bool,
    q: str,
    status_filter: str,
    ps: int,
    page: int,
) -> str:
    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if next_url:
        return next_url

    back_to = (request.POST.get("back_to") or request.GET.get("back_to") or "").strip().lower()
    if back_to == "detail":
        return _admin_plan_detail_url(
            plan.id,
            week_no=plan.week.week_no,
            next_url=_admin_dashboard_url(
                default_week_no,
                show_all=show_all,
                q=q,
                status=status_filter,
                ps=ps,
                page=page,
            ),
        )

    return _admin_dashboard_url(
        default_week_no,
        show_all=show_all,
        q=q,
        status=status_filter,
        ps=ps,
        page=page,
    )


def _login_page_context() -> dict:
    return {
        "week_choices": _build_week_choices(active_only=False),
        "today": timezone.localdate(),
    }


# =============================================================================
# Status helpers
# =============================================================================
def _status_label(plan: Plan) -> str:
    if plan.status == Plan.STATUS_APPROVED:
        return "معتمدة"
    if plan.status == Plan.STATUS_UNLOCK_REQUESTED:
        return "طلب فك اعتماد"
    return "مسودة"


def _status_code(plan: Plan) -> str:
    if plan.status == Plan.STATUS_APPROVED:
        return "approved"
    if plan.status == Plan.STATUS_UNLOCK_REQUESTED:
        return "unlock"
    return "draft"


def _status_css(plan: Plan) -> str:
    return _status_code(plan)


def _day_is_filled(d: Optional[PlanDay]) -> bool:
    if not d:
        return False
    if getattr(d, "school_id", None):
        return True
    return getattr(d, "visit_type", "") == getattr(PlanDay, "VISIT_NONE", "none")


def _plan_filled_count(plan: Plan) -> int:
    day_map = {d.weekday: d for d in plan.days.all()}
    return sum(1 for wd, _ in WEEKDAYS if _day_is_filled(day_map.get(wd)))


def _plan_visit_counts(plan: Plan) -> dict[str, int]:
    assigned_school_ids = set(
        Assignment.objects.filter(
            supervisor=plan.supervisor,
            is_active=True,
            school__is_active=True,
        ).values_list("school_id", flat=True)
    )

    visited_school_ids = set(
        PlanDay.objects.filter(
            plan=plan,
            visit_type=getattr(PlanDay, "VISIT_IN", "in"),
            school_id__isnull=False,
        ).values_list("school_id", flat=True)
    )

    visited_school_ids &= assigned_school_ids
    total_assigned = len(assigned_school_ids)
    visited_count = len(visited_school_ids)
    remaining_count = max(total_assigned - visited_count, 0)

    return {
        "assigned_count": total_assigned,
        "visited_count": visited_count,
        "remaining_count": remaining_count,
    }


def _notify_supervisor(
    *,
    supervisor: Supervisor,
    plan: Plan | None,
    notif_type: str,
    title: str,
    message: str | None = None,
) -> None:
    SupervisorNotification.objects.create(
        supervisor=supervisor,
        plan=plan,
        notif_type=notif_type,
        title=title,
        message=message or "",
    )


def _get_control_followup_model():
    """Return ControlFollowUp model if it exists after migration."""
    try:
        return apps.get_model("visits", "ControlFollowUp")
    except Exception:
        return None


def _control_followup_status_label(status: str) -> str:
    labels = {
        "open": "مفتوحة",
        "notified": "تم التنبيه",
        "pending_admin": "بانتظار مراجعة الإدارة",
        "processed": "تمت المعالجة",
        "closed": "مغلقة إداريًا",
    }
    return labels.get(status or "", status or "—")


def _control_followup_type_label(issue_type: str) -> str:
    labels = {
        "incomplete_plan": "خطة غير مكتملة",
        "not_saved_plan": "لم يحفظ خطة الأسبوع",
        "unlock_request": "طلب فك اعتماد",
        "uncovered_schools": "مدارس غير مغطاة",
    }
    return labels.get(issue_type or "", issue_type or "—")


def _supervisor_open_control_followups_count(supervisor: Supervisor) -> int:
    ControlFollowUp = _get_control_followup_model()
    if not ControlFollowUp:
        return 0
    return ControlFollowUp.objects.filter(
        supervisor=supervisor,
        status__in=["open", "notified", "pending_admin"],
    ).count()


def _parse_admin_action_state(request: HttpRequest) -> dict[str, Any]:
    show_all = (request.POST.get("all") or request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")
    return {
        "show_all": show_all,
        "q": (request.POST.get("q") or request.GET.get("q") or "").strip(),
        "status_filter": (request.POST.get("status") or request.GET.get("status") or "all").strip(),
        "ps": _safe_int(request.POST.get("ps") or request.GET.get("ps") or 10, default=10),
        "page": _safe_int(request.POST.get("page") or request.GET.get("page") or 1, default=1),
    }


def _build_plan_action_meta(plan: Plan) -> dict[str, Any]:
    filled = _plan_filled_count(plan)
    status_code = _status_code(plan)
    return {
        "plan_id": plan.id,
        "week_no": plan.week.week_no,
        "status_code": status_code,
        "status_css": _status_css(plan),
        "status_label": _status_label(plan),
        "admin_note": plan.admin_note or "",
        "filled": filled,
        "is_full": filled == 5,
        "can_approve": status_code != "approved" and filled == 5,
        "can_back_to_draft": status_code != "draft",
        "can_unlock_approve": status_code == "unlock",
        "can_unlock_reject": status_code == "unlock",
    }


def _plan_ajax_payload(plan: Plan, message: str, ok: bool = True, *, errors: list[str] | None = None) -> dict:
    payload = {
        "ok": ok,
        "message": message,
        "errors": errors or [],
    }
    payload.update(_build_plan_action_meta(plan))
    return payload


def _admin_json_response(
    request: HttpRequest,
    *,
    plan: Plan,
    message: str,
    ok: bool,
    http_status: int,
    errors: list[str] | None = None,
) -> JsonResponse | None:
    if not _is_ajax(request):
        return None
    return JsonResponse(_plan_ajax_payload(plan, message, ok=ok, errors=errors), status=http_status)




def _build_dashboard_global_visit_counts(*, include_breaks: bool = False) -> dict[str, int]:
    """
    يحسب مؤشرات الزيارات على مستوى جميع الأسابيع للوحة الإدارة.
    هذه القيم تغذي بطاقات:
    - الزيارات الحضورية لجميع الأسابيع
    - الزيارات عن بُعد لجميع الأسابيع
    """
    qs = PlanDay.objects.all()
    if not include_breaks:
        qs = qs.filter(plan__week__is_break=False)

    visit_in = getattr(PlanDay, "VISIT_IN", "in")
    visit_remote = getattr(PlanDay, "VISIT_REMOTE", "remote")
    visit_none = getattr(PlanDay, "VISIT_NONE", "none")

    global_visit_in_total = qs.filter(visit_type=visit_in).count()
    global_visit_remote_total = qs.filter(visit_type=visit_remote).count()
    global_visit_none_total = qs.filter(visit_type=visit_none).count()

    global_visit_total = global_visit_in_total + global_visit_remote_total
    global_days_total = global_visit_total + global_visit_none_total

    global_visit_in_percent = round((global_visit_in_total / global_visit_total) * 100) if global_visit_total else 0
    global_visit_remote_percent = round((global_visit_remote_total / global_visit_total) * 100) if global_visit_total else 0
    global_visit_none_percent = round((global_visit_none_total / global_days_total) * 100) if global_days_total else 0

    return {
        "global_visit_in_total": global_visit_in_total,
        "global_visit_remote_total": global_visit_remote_total,
        "global_visit_none_total": global_visit_none_total,
        "global_visit_total": global_visit_total,
        "global_days_total": global_days_total,
        "global_visit_in_percent": global_visit_in_percent,
        "global_visit_remote_percent": global_visit_remote_percent,
        "global_visit_none_percent": global_visit_none_percent,
    }

def _build_chart_counts(week_obj: PlanWeek) -> dict:
    base = (
        PlanDay.objects.filter(plan__week=week_obj)
        .values("weekday", "visit_type")
        .annotate(total=Count("id"))
        .order_by("weekday", "visit_type")
    )

    in_map = {wd: 0 for wd, _ in WEEKDAYS}
    remote_map = {wd: 0 for wd, _ in WEEKDAYS}
    none_map = {wd: 0 for wd, _ in WEEKDAYS}

    for item in base:
        wd = item["weekday"]
        vt = item["visit_type"]
        total = item["total"]

        if vt == getattr(PlanDay, "VISIT_IN", "in"):
            in_map[wd] = total
        elif vt == getattr(PlanDay, "VISIT_REMOTE", "remote"):
            remote_map[wd] = total
        elif vt == getattr(PlanDay, "VISIT_NONE", "none"):
            none_map[wd] = total

    return {
        "chart_labels": [name for _, name in WEEKDAYS],
        "chart_in_values": [in_map[wd] for wd, _ in WEEKDAYS],
        "chart_remote_values": [remote_map[wd] for wd, _ in WEEKDAYS],
        "chart_none_values": [none_map[wd] for wd, _ in WEEKDAYS],
        "chart_in_total": sum(in_map.values()),
        "chart_remote_total": sum(remote_map.values()),
        "chart_none_total": sum(none_map.values()),
    }


# =============================================================================
# Excel helpers
# =============================================================================
def _gender_label(value: str) -> str:
    v = (value or "").strip().lower()
    if v in ("boys", "male", "m", "بنين"):
        return "بنين"
    if v in ("girls", "female", "f", "بنات"):
        return "بنات"
    return value or "—"


def _excel_response(wb: Workbook, filename: str) -> HttpResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_plan_excel_workbook(plan: Plan) -> Workbook:
    _inject_week_no(plan)

    wb = Workbook()
    ws = wb.active
    ws.title = f"الأسبوع {plan.week.week_no}"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=12)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:C1")
    ws["A1"] = f"خطة الأسبوع رقم {plan.week.week_no}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:C2")
    ws["A2"] = f"المشرف: {plan.supervisor.full_name} — الهوية: {_sup_nid_value(plan.supervisor)}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    ws.append(["", "", ""])

    headers = ["اليوم", "المدرسة/السبب", "نوع اليوم"]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    days = {d.weekday: d for d in plan.days.all().select_related("school")}
    row = header_row + 1
    none_val = getattr(PlanDay, "VISIT_NONE", "none")

    for wd, wd_name in WEEKDAYS:
        d = days.get(wd)
        if d and d.visit_type == none_val:
            reason = d.get_no_visit_reason_display() if getattr(d, "no_visit_reason", None) else "بدون زيارة"
            if getattr(d, "note", None):
                reason = f"{reason} — {d.note}"
            school_or_reason = reason
        else:
            school_or_reason = d.school.name if d and d.school else "—"

        visit_label = d.get_visit_type_display() if d else "—"

        ws.cell(row=row, column=1, value=wd_name).font = normal_font
        ws.cell(row=row, column=2, value=school_or_reason).font = normal_font
        ws.cell(row=row, column=3, value=visit_label).font = normal_font

        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2).alignment = right
        ws.cell(row=row, column=3).alignment = center

        for c in range(1, 4):
            ws.cell(row=row, column=c).border = border
        row += 1

    for col_i, width in {1: 16, 2: 55, 3: 18}.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb


def _build_supervisor_assignments_excel_workbook(supervisor: Supervisor) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "المدارس المسندة"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    assignments = (
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
            school__is_active=True,
        )
        .select_related("school")
        .order_by("school__name")
    )

    school_ids = [a.school_id for a in assignments]
    principals_map = {p.school_id: p for p in Principal.objects.filter(school_id__in=school_ids)}

    ws.merge_cells("A1:G1")
    ws["A1"] = "المدارس المسندة للمشرف التربوي"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:G2")
    ws["A2"] = f"المشرف: {supervisor.full_name} — الهوية: {_sup_nid_value(supervisor)}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:G3")
    ws["A3"] = f"عدد المدارس: {assignments.count()}"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center

    headers = ["م", "الرقم الإحصائي", "اسم المدرسة", "الجنس", "مدير المدرسة", "جوال المدير", "حالة الإسناد"]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for i, assignment in enumerate(assignments, start=1):
        school = assignment.school
        principal = principals_map.get(school.id)
        values = [
            i,
            school.stat_code or "—",
            school.name or "—",
            _gender_label(getattr(school, "gender", "")),
            getattr(principal, "full_name", "") or "—",
            getattr(principal, "mobile", "") or "—",
            "نشط" if assignment.is_active else "غير نشط",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 6, 7) else right
        row_idx += 1

    for col_i, width in {1: 8, 2: 18, 3: 42, 4: 12, 5: 28, 6: 18, 7: 14}.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb


def _plan_school_tracking(target: Plan | Supervisor) -> dict[str, Any]:
    supervisor = target.supervisor if isinstance(target, Plan) else target

    assigned_schools = list(_supervisor_schools_qs(supervisor))
    assigned_school_ids = {school.id for school in assigned_schools}

    planned_school_ids = (
        set(
            PlanDay.objects.filter(
                plan__supervisor=supervisor,
                plan__week__is_break=False,
                school_id__isnull=False,
            ).values_list("school_id", flat=True)
        )
        & assigned_school_ids
    )

    planned_schools = [school for school in assigned_schools if school.id in planned_school_ids]
    unplanned_schools = [school for school in assigned_schools if school.id not in planned_school_ids]

    return {
        "assigned_schools": assigned_schools,
        "planned_schools": planned_schools,
        "unplanned_schools": unplanned_schools,
        "assigned_count": len(assigned_schools),
        "planned_count": len(planned_schools),
        "unplanned_count": len(unplanned_schools),
    }


def _build_supervisor_school_list_workbook(
    *,
    supervisor: Supervisor,
    schools: list[School],
    report_title: str,
    report_scope_label: str = "من الأسبوع الأول إلى آخر أسبوع في الخطة",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "المدارس"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{report_title} — {report_scope_label}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:E2")
    ws["A2"] = f"المشرف: {supervisor.full_name} — الهوية: {_sup_nid_value(supervisor)}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:E3")
    ws["A3"] = f"عدد المدارس: {len(schools)}"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center

    headers = ["م", "الرقم الإحصائي", "اسم المدرسة", "الجنس", "القطاع"]
    header_row = 5

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for i, school in enumerate(schools, start=1):
        values = [
            i,
            school.stat_code or "—",
            school.name or "—",
            _gender_label(getattr(school, "gender", "")),
            getattr(school.sector, "name", "") or "—",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 4) else right

        row_idx += 1

    for col_i, width in {
        1: 8,
        2: 18,
        3: 42,
        4: 12,
        5: 24,
    }.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb


def _build_unassigned_schools_excel_workbook(
    *,
    q: str = "",
    gender: str = "",
    sector_id: int = 0,
    only_active: bool = True,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "المدارس غير المسندة"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    assigned_school_ids = _globally_assigned_school_ids()

    schools = (
        School.objects.select_related("sector")
        .exclude(id__in=assigned_school_ids)
        .order_by("name")
    )

    if only_active:
        schools = schools.filter(is_active=True)

    if q:
        schools = schools.filter(
            Q(name__icontains=q) | Q(stat_code__icontains=q)
        )

    if gender in ("boys", "girls"):
        schools = schools.filter(gender=gender)

    if sector_id:
        schools = schools.filter(sector_id=sector_id)

    school_ids = list(schools.values_list("id", flat=True))
    principals_map = {
        p.school_id: p
        for p in Principal.objects.filter(school_id__in=school_ids)
    }

    sector_name = ""
    if sector_id:
        sector_name = (
            Sector.objects.filter(id=sector_id)
            .values_list("name", flat=True)
            .first()
            or ""
        )

    filters_text = []
    if q:
        filters_text.append(f"البحث: {q}")
    if gender in ("boys", "girls"):
        filters_text.append(f"الجنس: {_gender_label(gender)}")
    if sector_name:
        filters_text.append(f"القطاع: {sector_name}")
    filters_text.append("المدارس النشطة فقط" if only_active else "جميع المدارس")

    ws.merge_cells("A1:I1")
    ws["A1"] = "المدارس غير المسندة"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:I2")
    ws["A2"] = " — ".join(filters_text) if filters_text else "بدون فلاتر"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:I3")
    ws["A3"] = f"عدد المدارس: {schools.count()}"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center

    headers = [
        "م",
        "الرقم الإحصائي",
        "اسم المدرسة",
        "الجنس",
        "القطاع",
        "مدير المدرسة",
        "جوال المدير",
        "حالة المدرسة",
        "حالة الإسناد",
    ]

    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for i, school in enumerate(schools, start=1):
        principal = principals_map.get(school.id)

        values = [
            i,
            school.stat_code or "—",
            school.name or "—",
            _gender_label(getattr(school, "gender", "")),
            getattr(school.sector, "name", "") or "—",
            getattr(principal, "full_name", "") or "—",
            getattr(principal, "mobile", "") or "—",
            "نشطة" if school.is_active else "غير نشطة",
            "غير مسندة",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 7, 8, 9) else right

        row_idx += 1

    for col_i, width in {
        1: 8,
        2: 18,
        3: 42,
        4: 12,
        5: 24,
        6: 28,
        7: 18,
        8: 14,
        9: 16,
    }.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb

def _build_admin_week_excel_workbook(week_obj: PlanWeek, plans) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"الأسبوع {week_obj.week_no}"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"تصدير خطط المشرفين — الأسبوع {week_obj.week_no}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    headers = ["م", "اسم المشرف", "السجل المدني", "الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الحالة", "مكتملة"]
    header_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    none_val = getattr(PlanDay, "VISIT_NONE", "none")
    row_idx = header_row + 1

    for i, plan in enumerate(plans, start=1):
        day_map = {d.weekday: d for d in plan.days.all()}
        day_values: list[str] = []
        filled = 0

        for wd, _wd_name in WEEKDAYS:
            d = day_map.get(wd)
            text = "—"

            if d:
                visit_type_label = _visit_type_export_label(getattr(d, "visit_type", ""))

                if d.visit_type == none_val:
                    reason = d.get_no_visit_reason_display() if getattr(d, "no_visit_reason", None) else "بدون زيارة"
                    if getattr(d, "note", None):
                        reason = f"{reason} — {d.note}"
                    text = f"{reason} ({visit_type_label})"
                    filled += 1
                elif d.school:
                    school_name = d.school.name
                    if getattr(d, "note", None):
                        school_name = f"{school_name} — {d.note}"
                    text = f"{school_name} ({visit_type_label})"
                    filled += 1
                else:
                    text = visit_type_label

            day_values.append(text)

        row_values = [
            i,
            plan.supervisor.full_name,
            _sup_nid_value(plan.supervisor),
            *day_values,
            _status_label(plan),
            "نعم" if filled == 5 else f"لا ({filled}/5)",
        ]

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 9, 10) else right

        row_idx += 1

    for col_i, width in {1: 8, 2: 24, 3: 18, 4: 32, 5: 32, 6: 32, 7: 32, 8: 32, 9: 16, 10: 14}.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb



def _plan_day_export_text(day: Optional[PlanDay]) -> str:
    """صياغة نص مختصر لليوم داخل تصدير جميع الخطط."""
    if not day:
        return "—"

    visit_type = getattr(day, "visit_type", "") or ""
    visit_type_label = _visit_type_export_label(visit_type)
    none_val = getattr(PlanDay, "VISIT_NONE", "none")

    if visit_type == none_val:
        reason = (
            day.get_no_visit_reason_display()
            if getattr(day, "no_visit_reason", None)
            else "بدون زيارة"
        )
        if getattr(day, "note", None):
            reason = f"{reason} — {day.note}"
        return f"{reason} ({visit_type_label})"

    if getattr(day, "school_id", None) and getattr(day, "school", None):
        school_name = day.school.name or "—"
        if getattr(day, "note", None):
            school_name = f"{school_name} — {day.note}"
        return f"{school_name} ({visit_type_label})"

    if getattr(day, "note", None):
        return f"{visit_type_label} — {day.note}"

    return visit_type_label or "—"


def _build_all_supervisor_plans_excel_workbook(plans: list[Plan]) -> Workbook:
    """
    يبني ملف Excel شامل لجميع خطط المشرفين عبر كل الأسابيع.

    يحتوي الملف على ثلاث أوراق:
    1) ملخص المشرفين.
    2) كل الخطط: صف لكل مشرف في كل أسبوع.
    3) تفصيل الأيام: صف لكل يوم داخل كل خطة.
    """
    wb = Workbook()

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")
    approved_fill = PatternFill("solid", fgColor="ECFDF3")
    draft_fill = PatternFill("solid", fgColor="F8FAFC")
    incomplete_fill = PatternFill("solid", fgColor="FFF7ED")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =========================================================================
    # Sheet 1: Summary by supervisor
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "ملخص المشرفين"
    ws_summary.sheet_view.rightToLeft = True

    ws_summary.merge_cells("A1:I1")
    ws_summary["A1"] = "ملخص خطط المشرفين على مستوى جميع الأسابيع"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = center
    ws_summary["A1"].fill = title_fill

    summary_headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "القطاع",
        "إجمالي الخطط",
        "معتمدة",
        "مسودة",
        "طلب فك اعتماد",
        "غير مكتملة",
    ]

    for col, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    summary: dict[int, dict[str, Any]] = {}

    for plan in plans:
        supervisor = plan.supervisor
        sid = supervisor.id

        if sid not in summary:
            summary[sid] = {
                "supervisor": supervisor,
                "total": 0,
                "approved": 0,
                "draft": 0,
                "unlock": 0,
                "incomplete": 0,
            }

        filled = _plan_filled_count(plan)
        summary[sid]["total"] += 1

        if plan.status == Plan.STATUS_APPROVED:
            summary[sid]["approved"] += 1
        elif plan.status == Plan.STATUS_UNLOCK_REQUESTED:
            summary[sid]["unlock"] += 1
        else:
            summary[sid]["draft"] += 1

        if filled < 5:
            summary[sid]["incomplete"] += 1

    row_idx = 4
    for i, item in enumerate(summary.values(), start=1):
        supervisor = item["supervisor"]
        values = [
            i,
            supervisor.full_name,
            _sup_nid_value(supervisor),
            getattr(getattr(supervisor, "sector", None), "name", "") or "—",
            item["total"],
            item["approved"],
            item["draft"],
            item["unlock"],
            item["incomplete"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws_summary.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 6, 7, 8, 9) else right

            if col == 6:
                cell.fill = approved_fill
            elif col == 9:
                cell.fill = incomplete_fill

        row_idx += 1

    for col_i, width in {
        1: 7,
        2: 30,
        3: 18,
        4: 22,
        5: 14,
        6: 12,
        7: 12,
        8: 16,
        9: 14,
    }.items():
        ws_summary.column_dimensions[get_column_letter(col_i)].width = width

    # =========================================================================
    # Sheet 2: Matrix plans
    # =========================================================================
    ws = wb.create_sheet("كل الخطط")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:O1")
    ws["A1"] = "جميع خطط المشرفين لكل الأسابيع"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "القطاع",
        "الأسبوع",
        "عنوان الأسبوع",
        "تاريخ بداية الأسبوع",
        "الأحد",
        "الإثنين",
        "الثلاثاء",
        "الأربعاء",
        "الخميس",
        "الحالة",
        "مكتملة",
        "ملاحظة الإدارة",
    ]

    header_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1

    for i, plan in enumerate(plans, start=1):
        day_map = {d.weekday: d for d in plan.days.all()}
        filled = _plan_filled_count(plan)
        week = plan.week
        supervisor = plan.supervisor

        row_values = [
            i,
            supervisor.full_name,
            _sup_nid_value(supervisor),
            getattr(getattr(supervisor, "sector", None), "name", "") or "—",
            week.week_no,
            getattr(week, "title", "") or "—",
            week.start_sunday.strftime("%Y-%m-%d") if getattr(week, "start_sunday", None) else "—",
            _plan_day_export_text(day_map.get(0)),
            _plan_day_export_text(day_map.get(1)),
            _plan_day_export_text(day_map.get(2)),
            _plan_day_export_text(day_map.get(3)),
            _plan_day_export_text(day_map.get(4)),
            _status_label(plan),
            "نعم" if filled == 5 else f"لا ({filled}/5)",
            getattr(plan, "admin_note", "") or "—",
        ]

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 7, 13, 14) else right

            if col == 13 and plan.status == Plan.STATUS_APPROVED:
                cell.fill = approved_fill
            elif col == 13 and plan.status == Plan.STATUS_DRAFT:
                cell.fill = draft_fill
            elif col == 14 and filled < 5:
                cell.fill = incomplete_fill

        row_idx += 1

    for col_i, width in {
        1: 7,
        2: 30,
        3: 18,
        4: 22,
        5: 10,
        6: 24,
        7: 18,
        8: 36,
        9: 36,
        10: 36,
        11: 36,
        12: 36,
        13: 15,
        14: 14,
        15: 38,
    }.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # =========================================================================
    # Sheet 3: Daily details
    # =========================================================================
    ws_days = wb.create_sheet("تفصيل الأيام")
    ws_days.sheet_view.rightToLeft = True

    ws_days.merge_cells("A1:K1")
    ws_days["A1"] = "تفصيل أيام الخطط لجميع المشرفين"
    ws_days["A1"].font = title_font
    ws_days["A1"].alignment = center
    ws_days["A1"].fill = title_fill

    day_headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "الأسبوع",
        "اليوم",
        "نوع الزيارة",
        "المدرسة",
        "الرقم الإحصائي",
        "سبب عدم الزيارة",
        "ملاحظة اليوم",
        "حالة الخطة",
    ]

    for col, header in enumerate(day_headers, start=1):
        cell = ws_days.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = 4
    counter = 1

    for plan in plans:
        day_map = {d.weekday: d for d in plan.days.all()}

        for weekday, weekday_name in WEEKDAYS:
            d = day_map.get(weekday)
            school = getattr(d, "school", None) if d else None
            visit_type_label = _visit_type_export_label(getattr(d, "visit_type", "")) if d else "—"

            no_visit_reason = "—"
            if d and getattr(d, "no_visit_reason", None):
                no_visit_reason = d.get_no_visit_reason_display()

            row_values = [
                counter,
                plan.supervisor.full_name,
                _sup_nid_value(plan.supervisor),
                plan.week.week_no,
                weekday_name,
                visit_type_label,
                getattr(school, "name", "") or "—",
                getattr(school, "stat_code", "") or "—",
                no_visit_reason,
                getattr(d, "note", "") if d else "—",
                _status_label(plan),
            ]

            for col, value in enumerate(row_values, start=1):
                cell = ws_days.cell(row=row_idx, column=col, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center if col in (1, 3, 4, 5, 6, 8, 11) else right

            row_idx += 1
            counter += 1

    for col_i, width in {
        1: 7,
        2: 30,
        3: 18,
        4: 10,
        5: 12,
        6: 15,
        7: 36,
        8: 18,
        9: 28,
        10: 40,
        11: 15,
    }.items():
        ws_days.column_dimensions[get_column_letter(col_i)].width = width

    return wb


def _get_assignment_dashboard_context() -> dict[str, Any]:
    """
    مؤشرات مختصرة للوحة الإدارة الرئيسية عن حالة الإسناد.

    لا يعتمد هذا المساعد على ملف خارجي، حتى يكون views.py مكتملًا بذاته.
    ويجلب آخر عملية معالجة فقط إذا كان Model:
    AssignmentReviewLog
    موجودًا ومهاجرًا في قاعدة البيانات.
    """
    active_assignments = Assignment.objects.filter(is_active=True)

    assigned_active_school_ids = (
        active_assignments
        .filter(school__is_active=True)
        .values_list("school_id", flat=True)
        .distinct()
    )

    unassigned_schools_count = (
        School.objects
        .filter(is_active=True)
        .exclude(id__in=assigned_active_school_ids)
        .count()
    )

    duplicate_assignments_count = (
        active_assignments
        .values("school_id")
        .annotate(total=Count("id"))
        .filter(total__gt=1, school_id__isnull=False)
        .count()
    )

    inactive_supervisor_assignments_count = (
        active_assignments
        .filter(supervisor__is_active=False)
        .count()
    )

    inactive_school_assignments_count = (
        active_assignments
        .filter(school__is_active=False)
        .count()
    )

    latest_assignment_log = None
    try:
        AssignmentReviewLog = apps.get_model("visits", "AssignmentReviewLog")
        latest_assignment_log = (
            AssignmentReviewLog.objects
            .select_related("school", "supervisor", "user")
            .order_by("-created_at", "-id")
            .first()
        )
    except Exception:
        latest_assignment_log = None

    return {
        "dashboard_unassigned_schools_count": unassigned_schools_count,
        "dashboard_duplicate_assignments_count": duplicate_assignments_count,
        "dashboard_inactive_supervisor_assignments_count": inactive_supervisor_assignments_count,
        "dashboard_inactive_school_assignments_count": inactive_school_assignments_count,
        "dashboard_latest_assignment_log": latest_assignment_log,
    }


# =============================================================================
# Auth views
# =============================================================================
def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated and request.user.is_staff:
        return redirect("visits:admin_dashboard")

    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    context = _login_page_context()

    if request.method == "POST":
        nid = _digits(
            (
                request.POST.get("nid")
                or request.POST.get("national_id")
                or request.POST.get("civil_registry")
                or ""
            ).strip()
        )
        last4 = _digits(
            (
                request.POST.get("last4")
                or request.POST.get("phone_last4")
                or request.POST.get("mobile_last4")
                or ""
            ).strip()
        )

        if len(nid) != 10:
            messages.error(request, "فضلاً أدخل رقم الهوية بشكل صحيح.")
            return render(request, "visits/login.html", context)

        if len(last4) != 4:
            messages.error(request, "فضلاً أدخل آخر 4 أرقام من الجوال بشكل صحيح.")
            return render(request, "visits/login.html", context)

        sup = _find_supervisor_by_nid(nid)
        if not sup:
            messages.error(request, "المشرف غير موجود أو غير مفعل.")
            return render(request, "visits/login.html", context)

        sup_last4 = _supervisor_last4(sup)
        if not sup_last4:
            messages.error(request, "لا يمكن التحقق لأن رقم جوال المشرف غير محفوظ. راجع الإدارة لإضافته.")
            return render(request, "visits/login.html", context)

        if sup_last4 != last4:
            messages.error(request, "بيانات التحقق غير صحيحة.")
            return render(request, "visits/login.html", context)

        logout(request)
        request.session.flush()
        request.session[SESSION_SUP_ID] = sup.id
        request.session.modified = True
        messages.success(request, f"مرحبًا بك {sup.full_name}.")
        return redirect("visits:supervisor_dashboard")

    return render(request, "visits/login.html", context)


def admin_login_view(request: HttpRequest) -> HttpResponse:
    supervisor_redirect = _redirect_supervisor_from_admin(request)
    if supervisor_redirect:
        return supervisor_redirect

    if request.user.is_authenticated and request.user.is_staff:
        return redirect("visits:admin_dashboard")

    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""

        if not username or not password:
            messages.error(request, "فضلاً أدخل اسم المستخدم وكلمة المرور.")
            return render(request, "visits/admin_login.html", {})

        user = authenticate(request, username=username, password=password)
        if not user:
            messages.error(request, "بيانات الدخول غير صحيحة.")
            return render(request, "visits/admin_login.html", {})

        if not user.is_staff:
            messages.error(request, "هذا الحساب لا يملك صلاحية الدخول للإدارة.")
            return render(request, "visits/admin_login.html", {})

        request.session.pop(SESSION_SUP_ID, None)
        login(request, user)
        messages.success(request, "تم تسجيل دخول الإدارة بنجاح.")
        return redirect("visits:admin_dashboard")

    return render(request, "visits/admin_login.html", {})


def logout_view(request: HttpRequest) -> HttpResponse:
    was_supervisor = bool(request.session.get(SESSION_SUP_ID))
    was_admin = bool(request.user.is_authenticated)

    request.session.pop(SESSION_SUP_ID, None)
    logout(request)

    if was_supervisor and not was_admin:
        messages.success(request, "تم تسجيل خروج المشرف بنجاح.")
        return redirect("visits:login")

    if was_admin:
        messages.success(request, "تم تسجيل خروج الإدارة بنجاح.")
        return redirect("visits:admin_login")

    messages.success(request, "تم تسجيل الخروج.")
    return redirect("visits:login")


def supervisor_dashboard_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_obj = _current_week_obj()
    if not week_obj:
        default_week_no = _get_default_week_no()
        week_obj = PlanWeek.objects.filter(week_no=default_week_no).first()

    current_week_no = week_obj.week_no if week_obj else _get_default_week_no()

    assignment_count = len(_supervisor_school_ids(supervisor))
    unread_count = supervisor.notifications.filter(is_read=False).count()
    control_followup_count = _supervisor_open_control_followups_count(supervisor)

    return render(
        request,
        "visits/supervisor_dashboard.html",
        {
            "sup": supervisor,
            "week": current_week_no,
            "current_week": current_week_no,
            "week_obj": week_obj,
            "assignment_count": assignment_count,
            "unread_count": unread_count,
            "control_followup_count": control_followup_count,
            "supervisor_email": _supervisor_email_value(supervisor),
            "email_notifications_enabled": _supervisor_email_notifications_enabled(supervisor),
        },
    )


# =============================================================================
# Letter views
# =============================================================================
def print_assignment_letter_view(request: HttpRequest) -> HttpResponse:
    nid = _digits((request.GET.get("nid") or "").strip())
    week = _safe_int(request.GET.get("week") or 0, default=0)

    if len(nid) != 10:
        messages.error(request, "يرجى إدخال السجل المدني بشكل صحيح.")
        return redirect("visits:login")

    if not PlanWeek.objects.filter(week_no=week).exists():
        messages.error(request, "يرجى اختيار أسبوع صحيح.")
        return redirect("visits:login")

    sup = _find_supervisor_by_nid(nid)
    if not sup:
        messages.error(request, "لم يتم العثور على مشرف بهذا السجل المدني.")
        return redirect("visits:login")

    link_obj = WeeklyLetterLink.objects.filter(week__week_no=week, is_active=True).select_related("week").first()
    if not link_obj or not link_obj.drive_url:
        messages.warning(request, "لا يوجد رابط خطاب منشور لهذا الأسبوع.")
        return redirect("visits:login")

    folder_id = _extract_drive_folder_id(link_obj.drive_url)
    if not folder_id:
        messages.error(request, "رابط الأسبوع غير صالح. يجب أن يكون رابط مجلد Google Drive صحيح.")
        return redirect("visits:login")

    try:
        file_url = _find_supervisor_letter_in_folder(folder_id=folder_id, national_id=nid)
    except Exception as e:
        messages.error(request, f"تعذر البحث عن خطاب المشرف داخل Google Drive: {e}")
        return redirect("visits:login")

    if not file_url:
        messages.warning(
            request,
            "لم يتم العثور على خطاب مطابق للسجل المدني داخل مجلد هذا الأسبوع. "
            "تأكد أن اسم الملف يحتوي السجل المدني للمشرف."
        )
        return redirect("visits:login")

    return redirect(file_url)


def current_week_letter_redirect_view(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_obj = _current_week_obj()
    if not week_obj:
        messages.error(request, "لم يتم العثور على أسبوع حالي.")
        return redirect("visits:supervisor_dashboard")

    link_obj = WeeklyLetterLink.objects.filter(week=week_obj, is_active=True).first()
    if not link_obj or not link_obj.drive_url:
        messages.warning(request, "لا يوجد رابط خطاب منشور لهذا الأسبوع.")
        return redirect(_plan_url(week_obj.week_no))

    folder_id = _extract_drive_folder_id(link_obj.drive_url)
    if not folder_id:
        messages.error(request, "رابط الأسبوع غير صالح. يجب أن يكون رابط مجلد Google Drive صحيح.")
        return redirect(_plan_url(week_obj.week_no))

    nid = _digits(_sup_nid_value(supervisor))
    if len(nid) != 10:
        messages.error(request, "السجل المدني للمشرف غير محفوظ بشكل صحيح.")
        return redirect(_plan_url(week_obj.week_no))

    try:
        file_url = _find_supervisor_letter_in_folder(folder_id=folder_id, national_id=nid)
    except Exception as e:
        messages.error(request, f"تعذر البحث عن الخطاب داخل Google Drive: {e}")
        return redirect(_plan_url(week_obj.week_no))

    if not file_url:
        messages.warning(
            request,
            "لم يتم العثور على خطابك داخل مجلد هذا الأسبوع. "
            "تأكد أن اسم الملف في Google Drive يحتوي السجل المدني."
        )
        return redirect(_plan_url(week_obj.week_no))

    return redirect(file_url)


@admin_only_view
def weekly_letters_drive_view(request: HttpRequest, week_number: int) -> HttpResponse:
    link_obj = WeeklyLetterLink.objects.filter(
        week__week_no=week_number,
        is_active=True,
    ).select_related("week").first()

    week_folder = None
    rows = []

    if link_obj and link_obj.drive_url:
        folder_id = _extract_drive_folder_id(link_obj.drive_url)
        if folder_id:
            try:
                files = list_files_in_folder(folder_id, page_size=300)
                week_folder = {
                    "id": folder_id,
                    "name": f"الأسبوع {week_number}",
                    "webViewLink": link_obj.drive_url,
                }
                for f in files:
                    filename = (f.get("name") or "").strip()
                    rows.append(
                        {
                            "school_code": filename.removesuffix(".pdf").strip(),
                            "name": filename,
                            "url": f.get("webViewLink"),
                            "file_id": f.get("id"),
                            "mime_type": f.get("mimeType"),
                        }
                    )
            except Exception as e:
                messages.error(request, f"تعذر قراءة ملفات Google Drive: {e}")
        else:
            messages.error(request, "رابط Google Drive المحفوظ غير صالح.")
    else:
        messages.warning(request, "لا يوجد رابط نشط محفوظ لهذا الأسبوع.")

    return render(
        request,
        "visits/weekly_letters_drive.html",
        {
            "week_number": week_number,
            "week_folder": week_folder,
            "rows": rows,
            "link_obj": link_obj,
        },
    )


# =============================================================================
# Supervisor plan views
# =============================================================================
def _plan_school_tracking(target: Plan | Supervisor) -> dict[str, Any]:
    supervisor = target.supervisor if isinstance(target, Plan) else target

    assigned_schools = list(_supervisor_schools_qs(supervisor))
    assigned_school_ids = {school.id for school in assigned_schools}

    planned_school_ids = (
        set(
            PlanDay.objects.filter(
                plan__supervisor=supervisor,
                plan__week__is_break=False,
                school_id__isnull=False,
            ).values_list("school_id", flat=True)
        )
        & assigned_school_ids
    )

    planned_schools = [school for school in assigned_schools if school.id in planned_school_ids]
    unplanned_schools = [school for school in assigned_schools if school.id not in planned_school_ids]

    return {
        "assigned_schools": assigned_schools,
        "planned_schools": planned_schools,
        "unplanned_schools": unplanned_schools,
        "assigned_count": len(assigned_schools),
        "planned_count": len(planned_schools),
        "unplanned_count": len(unplanned_schools),
    }


def plan_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_no = _safe_int(request.GET.get("week") or request.POST.get("week") or _get_default_week_no(), default=_get_default_week_no())

    if _supervisor_needs_email_prompt(supervisor):
        messages.info(request, "يرجى تسجيل بريدك الإلكتروني أولًا وتأكيده برمز تحقق لتصلك التنبيهات والإشعارات.")
        settings_url = reverse("visits:supervisor_email_settings")
        return redirect(f"{settings_url}?next={_plan_url(week_no)}")

    week_obj = _resolve_week_or_404(week_no, allow_inactive=False)

    plan, _ = Plan.objects.get_or_create(supervisor=supervisor, week=week_obj)
    _inject_week_no(plan)

    schools = _supervisor_schools_qs(supervisor)
    allowed_school_ids = _supervisor_school_ids(supervisor)
    days_map = {d.weekday: d for d in plan.days.all().select_related("school")}
    week_choices = _build_week_choices(active_only=True)
    day_dates = _build_day_dates_from_week(week_obj)
    notifications = supervisor.notifications.select_related("plan").order_by("-created_at")[:8]
    unread_count = supervisor.notifications.filter(is_read=False).count()
    week_letter = WeeklyLetterLink.objects.filter(week=week_obj, is_active=True).first()

    if request.method == "POST":
        if plan.status in (Plan.STATUS_APPROVED, Plan.STATUS_UNLOCK_REQUESTED):
            messages.info(request, "لا يمكن تعديل الخطة الآن. إذا كانت معتمدة فاطلب فك الاعتماد أولًا.")
            return redirect(_plan_url(week_obj.week_no))

        action = (request.POST.get("action") or "save").strip()

        for wd, _ in WEEKDAYS:
            sid = _safe_int((request.POST.get(f"school_{wd}") or "").strip(), default=0)
            vtype = (request.POST.get(f"visit_{wd}") or getattr(PlanDay, "VISIT_IN", "in")).strip()

            allowed_visit_types = {
                getattr(PlanDay, "VISIT_IN", "in"),
                getattr(PlanDay, "VISIT_REMOTE", "remote"),
                getattr(PlanDay, "VISIT_NONE", "none"),
            }
            if vtype not in allowed_visit_types:
                vtype = getattr(PlanDay, "VISIT_IN", "in")

            reason = (request.POST.get(f"reason_{wd}") or "").strip() or None
            note = (request.POST.get(f"note_{wd}") or "").strip() or None

            if vtype == getattr(PlanDay, "VISIT_NONE", "none"):
                sid = 0

            if sid and sid not in allowed_school_ids:
                messages.warning(request, f"تم تجاهل مدرسة غير مسندة للمشرف في يوم {WEEKDAY_MAP.get(wd, wd)}.")
                PlanDay.objects.filter(plan=plan, weekday=wd).delete()
                continue

            if (not sid) and (vtype != getattr(PlanDay, "VISIT_NONE", "none")):
                PlanDay.objects.filter(plan=plan, weekday=wd).delete()
                continue

            old_day = PlanDay.objects.filter(plan=plan, weekday=wd).first()
            visited = getattr(old_day, "visited", False) if old_day else False
            visited_at = getattr(old_day, "visited_at", None) if old_day else None
            visit_note = getattr(old_day, "visit_note", None) if old_day else None

            if vtype != getattr(PlanDay, "VISIT_IN", "in") or not sid:
                visited = False
                visited_at = None
                visit_note = None

            defaults = {
                "visit_type": vtype,
                "note": note,
                "visited": visited,
                "visited_at": visited_at,
                "visit_note": visit_note,
            }

            if sid:
                defaults["school_id"] = sid
                defaults["no_visit_reason"] = None
            else:
                defaults["school"] = None
                defaults["no_visit_reason"] = reason

            PlanDay.objects.update_or_create(plan=plan, weekday=wd, defaults=defaults)

        plan.saved_at = timezone.now()

        if action == "approve":
            if plan.is_fully_filled():
                plan.status = Plan.STATUS_APPROVED
                plan.approved_at = timezone.now()
                plan.admin_note = None
                plan.save(update_fields=["saved_at", "status", "approved_at", "admin_note"])
                messages.success(request, "تم اعتماد الخطة بنجاح.")
            else:
                plan.save(update_fields=["saved_at"])
                messages.warning(request, "تم الحفظ، لكن لا يمكن الاعتماد قبل اكتمال جميع الأيام.")
        else:
            plan.save(update_fields=["saved_at"])
            messages.success(request, "تم حفظ الخطة بنجاح.")

        return redirect(_plan_url(week_obj.week_no))

    school_tracking = _plan_school_tracking(supervisor)

    return render(
        request,
        "visits/plan.html",
        {
            "plan": plan,
            "schools": schools,
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "sup": supervisor,
            "weekdays": WEEKDAYS,
            "days_map": days_map,
            "week_choices": week_choices,
            "day_dates": day_dates,
            "today": timezone.localdate(),
            "notifications": notifications,
            "unread_count": unread_count,
            "visit_none_value": getattr(PlanDay, "VISIT_NONE", "none"),
            "visit_in_value": getattr(PlanDay, "VISIT_IN", "in"),
            "visit_remote_value": getattr(PlanDay, "VISIT_REMOTE", "remote"),
            "week_letter": week_letter,

            "assigned_count": school_tracking["assigned_count"],
            "planned_count": school_tracking["planned_count"],
            "unplanned_count": school_tracking["unplanned_count"],
            "assigned_schools": school_tracking["assigned_schools"],
            "planned_schools": school_tracking["planned_schools"],
            "unplanned_schools": school_tracking["unplanned_schools"],

            "visited_count": school_tracking["planned_count"],
            "remaining_count": school_tracking["unplanned_count"],

            "supervisor_email": _supervisor_email_value(supervisor),
            "email_notifications_enabled": _supervisor_email_notifications_enabled(supervisor),
        },
    )


def supervisor_email_settings_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    next_url = request.POST.get("next") or request.GET.get("next") or _plan_url(_get_default_week_no())

    if request.method == "POST":
        email = _cell_str(request.POST.get("email")).lower()
        enabled = _bool_from_post(request.POST.get("email_notifications_enabled"), default=True)

        if not email:
            messages.error(request, "يرجى إدخال البريد الإلكتروني.")
            return render(
                request,
                "visits/supervisor_email_settings.html",
                {
                    "sup": supervisor,
                    "next_url": next_url,
                    "supervisor_email": email,
                    "email_notifications_enabled": enabled,
                },
            )

        code = _generate_email_otp()
        expires_at = timezone.now() + timedelta(minutes=EMAIL_OTP_EXPIRE_MINUTES)

        old_qs = EmailOTP.objects.filter(supervisor=supervisor)
        if hasattr(EmailOTP, "is_used"):
            old_qs.filter(is_used=False).update(is_used=True)
        else:
            old_qs.delete()

        otp = EmailOTP.objects.create(
            supervisor=supervisor,
            email=email,
            code=code,
            expires_at=expires_at,
        )

        request.session[EMAIL_OTP_SESSION_KEY] = {
            "otp_id": otp.id,
            "supervisor_id": supervisor.id,
            "email_notifications_enabled": bool(enabled),
            "next_url": next_url,
        }
        request.session.modified = True

        try:
            _send_supervisor_email_otp(
                email=email,
                code=otp.code,
                supervisor_name=supervisor.full_name,
            )
            messages.success(request, "تم إرسال رمز التحقق إلى بريدك الإلكتروني.")
        except Exception:
            messages.warning(
                request,
                "تم حفظ طلب التحقق، لكن تعذر إرسال الرمز عبر البريد حاليًا. يمكنك إعادة الإرسال من صفحة التحقق.",
            )

        return redirect("visits:supervisor_email_verify")

    return render(
        request,
        "visits/supervisor_email_settings.html",
        {
            "sup": supervisor,
            "next_url": next_url,
            "supervisor_email": _supervisor_email_value(supervisor),
            "email_notifications_enabled": _supervisor_email_notifications_enabled(supervisor),
        },
    )


def supervisor_email_verify_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    payload = request.session.get(EMAIL_OTP_SESSION_KEY, {}) or {}
    otp_id = _safe_int(payload.get("otp_id") or 0, default=0)

    if not otp_id or payload.get("supervisor_id") != supervisor.id:
        messages.warning(request, "لا يوجد طلب تحقق نشط. يرجى إدخال البريد الإلكتروني أولًا.")
        return redirect("visits:supervisor_email_settings")

    otp = EmailOTP.objects.filter(id=otp_id, supervisor=supervisor).first()
    if not otp:
        request.session.pop(EMAIL_OTP_SESSION_KEY, None)
        request.session.modified = True
        messages.warning(request, "تعذر العثور على طلب التحقق. يرجى طلب رمز جديد.")
        return redirect("visits:supervisor_email_settings")

    is_used = bool(getattr(otp, "is_used", False))
    expires_at = getattr(otp, "expires_at", None)
    is_expired = bool(expires_at and timezone.now() > expires_at)

    if is_used or is_expired:
        request.session.pop(EMAIL_OTP_SESSION_KEY, None)
        request.session.modified = True
        messages.warning(request, "انتهت صلاحية رمز التحقق. يرجى طلب رمز جديد.")
        return redirect("visits:supervisor_email_settings")

    next_url = payload.get("next_url") or _plan_url(_get_default_week_no())
    masked_email = getattr(otp, "email", "")

    if request.method == "POST":
        code = _cell_str(request.POST.get("otp_code") or request.POST.get("code"))

        if code != getattr(otp, "code", ""):
            messages.error(request, "رمز التحقق غير صحيح.")
            return render(
                request,
                "visits/supervisor_email_verify.html",
                {
                    "sup": supervisor,
                    "masked_email": masked_email,
                    "next_url": next_url,
                },
            )

        supervisor.email = masked_email
        supervisor.email_notifications_enabled = bool(payload.get("email_notifications_enabled", True))
        if hasattr(supervisor, "email_verified"):
            supervisor.email_verified = True
            supervisor.save(update_fields=["email", "email_notifications_enabled", "email_verified"])
        else:
            supervisor.save(update_fields=["email", "email_notifications_enabled"])

        if hasattr(otp, "is_used"):
            otp.is_used = True
            if hasattr(otp, "used_at"):
                otp.used_at = timezone.now()
                otp.save(update_fields=["is_used", "used_at"])
            else:
                otp.save(update_fields=["is_used"])
        else:
            otp.delete()

        request.session.pop(EMAIL_OTP_SESSION_KEY, None)
        request.session.modified = True
        messages.success(request, "تم تأكيد البريد الإلكتروني وحفظه بنجاح.")
        return redirect(next_url)

    return render(
        request,
        "visits/supervisor_email_verify.html",
        {
            "sup": supervisor,
            "masked_email": masked_email,
            "next_url": next_url,
        },
    )


@require_POST
def supervisor_email_resend_otp_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    payload = request.session.get(EMAIL_OTP_SESSION_KEY, {}) or {}
    otp_id = _safe_int(payload.get("otp_id") or 0, default=0)

    if not otp_id or payload.get("supervisor_id") != supervisor.id:
        messages.warning(request, "لا يوجد طلب تحقق نشط لإعادة الإرسال.")
        return redirect("visits:supervisor_email_settings")

    otp = EmailOTP.objects.filter(id=otp_id, supervisor=supervisor).first()
    if not otp:
        request.session.pop(EMAIL_OTP_SESSION_KEY, None)
        request.session.modified = True
        messages.warning(request, "تعذر العثور على طلب التحقق. يرجى إدخال البريد الإلكتروني من جديد.")
        return redirect("visits:supervisor_email_settings")

    if getattr(otp, "is_used", False):
        request.session.pop(EMAIL_OTP_SESSION_KEY, None)
        request.session.modified = True
        messages.warning(request, "تم استخدام رمز التحقق السابق بالفعل. يرجى إدخال البريد الإلكتروني من جديد.")
        return redirect("visits:supervisor_email_settings")

    otp.code = _generate_email_otp()
    otp.expires_at = timezone.now() + timedelta(minutes=EMAIL_OTP_EXPIRE_MINUTES)

    update_fields = ["code", "expires_at"]
    if hasattr(otp, "is_used"):
        otp.is_used = False
        update_fields.append("is_used")
    if hasattr(otp, "used_at"):
        otp.used_at = None
        update_fields.append("used_at")

    otp.save(update_fields=update_fields)

    try:
        _send_supervisor_email_otp(
            email=otp.email,
            code=otp.code,
            supervisor_name=supervisor.full_name,
        )
        messages.success(request, "تمت إعادة إرسال رمز التحقق إلى بريدك الإلكتروني.")
    except Exception as e:
        messages.error(request, f"فشل إعادة إرسال رمز التحقق: {e}")

    return redirect("visits:supervisor_email_verify")


@require_POST
def toggle_email_notifications_view(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    next_url = request.POST.get("next") or request.GET.get("next") or _plan_url(_get_default_week_no())

    if not _supervisor_email_value(supervisor):
        payload = request.session.get(EMAIL_OTP_SESSION_KEY, {}) or {}
        otp_id = _safe_int(payload.get("otp_id") or 0, default=0)
        pending_otp = None
        if otp_id and payload.get("supervisor_id") == supervisor.id:
            pending_otp = EmailOTP.objects.filter(id=otp_id, supervisor=supervisor).first()

        if pending_otp and getattr(pending_otp, "email", ""):
            payload["email_notifications_enabled"] = not bool(payload.get("email_notifications_enabled", True))
            payload["next_url"] = next_url
            request.session[EMAIL_OTP_SESSION_KEY] = payload
            request.session.modified = True

            if payload["email_notifications_enabled"]:
                messages.success(request, "سيتم تفعيل التنبيهات البريدية بعد تأكيد البريد الإلكتروني.")
            else:
                messages.success(request, "سيتم إيقاف التنبيهات البريدية بعد تأكيد البريد الإلكتروني.")
            return redirect("visits:supervisor_email_verify")

        messages.warning(request, "سجل البريد الإلكتروني أولًا قبل تفعيل التنبيهات البريدية.")
        return redirect(f"{reverse('visits:supervisor_email_settings')}?next={next_url}")

    supervisor.email_notifications_enabled = not bool(getattr(supervisor, "email_notifications_enabled", False))
    supervisor.save(update_fields=["email_notifications_enabled"])

    if supervisor.email_notifications_enabled:
        messages.success(request, "تم تفعيل التنبيهات البريدية.")
    else:
        messages.success(request, "تم إيقاف التنبيهات البريدية.")

    return redirect(next_url)


def supervisor_visit_status_view(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
    week_obj = _resolve_week_or_404(week_no, allow_inactive=False)

    plan = get_object_or_404(
        Plan.objects.select_related("supervisor", "week").prefetch_related("days__school"),
        supervisor=supervisor,
        week=week_obj,
    )
    _inject_week_no(plan)

    school_days = [d for d in plan.days.all() if d.visit_type == getattr(PlanDay, "VISIT_IN", "in") and d.school_id]
    visited_days = [d for d in school_days if getattr(d, "visited", False)]
    unvisited_days = [d for d in school_days if not getattr(d, "visited", False)]

    notifications = supervisor.notifications.select_related("plan").order_by("-created_at")[:8]
    unread_count = supervisor.notifications.filter(is_read=False).count()
    visit_counts = _plan_visit_counts(plan)

    assigned_school_ids = set(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
            school__is_active=True,
        ).values_list("school_id", flat=True)
    )
    remaining_school_ids = assigned_school_ids - set(d.school_id for d in visited_days if d.school_id)
    remaining_schools = list(School.objects.filter(id__in=remaining_school_ids, is_active=True).order_by("name"))

    return render(
        request,
        "visits/supervisor_visit_status.html",
        {
            "sup": supervisor,
            "plan": plan,
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "school_days": school_days,
            "visited_days": visited_days,
            "unvisited_days": unvisited_days,
            "remaining_schools": remaining_schools,
            "assigned_count": visit_counts["assigned_count"],
            "visited_count": visit_counts["visited_count"],
            "remaining_count": visit_counts["remaining_count"],
            "notifications": notifications,
            "unread_count": unread_count,
        },
    )


@require_POST
def toggle_day_visited_view(request: HttpRequest, day_id: int) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    day = get_object_or_404(
        PlanDay.objects.select_related("plan", "plan__supervisor", "plan__week", "school"),
        id=day_id,
        plan__supervisor=supervisor,
    )

    if day.visit_type != getattr(PlanDay, "VISIT_IN", "in") or not day.school_id:
        msg = "لا يمكن تتبع هذا اليوم لأنه ليس زيارة حضورية مدرسية."
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": msg}, status=400)
        messages.warning(request, msg)
        return redirect(_supervisor_visit_status_url(day.plan.week.week_no))

    visited_raw = (request.POST.get("visited") or "").strip()
    note = (request.POST.get("visit_note") or request.POST.get("note") or "").strip() or None

    if visited_raw in ("1", "true", "yes", "on"):
        day.visited = True
        day.visited_at = timezone.now()
        day.visit_note = note
        msg = "تم تسجيل الزيارة بنجاح."
    else:
        day.visited = False
        day.visited_at = None
        day.visit_note = None
        msg = "تم إلغاء تسجيل الزيارة."

    day.save()
    visit_counts = _plan_visit_counts(day.plan)

    if _is_ajax(request):
        return JsonResponse(
            {
                "ok": True,
                "message": msg,
                "day_id": day.id,
                "plan_id": day.plan.id,
                "visited": day.visited,
                "visited_at": day.visited_at.strftime("%Y-%m-%d %H:%M") if day.visited_at else "",
                "visit_note": day.visit_note or "",
                "assigned_count": visit_counts["assigned_count"],
                "visited_count": visit_counts["visited_count"],
                "remaining_count": visit_counts["remaining_count"],
            },
            status=200,
        )

    messages.success(request, msg)
    return redirect(_supervisor_visit_status_url(day.plan.week.week_no))


# =============================================================================
# Notifications
# =============================================================================
def notifications_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
    notifications_qs = supervisor.notifications.select_related("plan", "plan__week").all()

    paginator = Paginator(notifications_qs, 20)
    page = _safe_int(request.GET.get("page") or 1, default=1)
    page_obj = paginator.get_page(page)

    unread_count = supervisor.notifications.filter(is_read=False).count()

    return render(
        request,
        "visits/notifications.html",
        {
            "notifications": page_obj.object_list,
            "page_obj": page_obj,
            "sup": supervisor,
            "unread_count": unread_count,
            "week": week_no,
            "supervisor_email": _supervisor_email_value(supervisor),
            "email_notifications_enabled": _supervisor_email_notifications_enabled(supervisor),
        },
    )


@require_POST
def mark_notification_read_view(request: HttpRequest, notif_id: int) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    notif = get_object_or_404(SupervisorNotification, id=notif_id, supervisor=supervisor)
    if not notif.is_read:
        notif.is_read = True
        notif.save(update_fields=["is_read"])

    next_url = request.POST.get("next") or request.GET.get("next") or _notifications_url()
    return redirect(next_url)


@require_POST
def mark_all_notifications_read_view(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    supervisor.notifications.filter(is_read=False).update(is_read=True)
    messages.success(request, "تم تعليم جميع الإشعارات كمقروءة.")
    next_url = request.POST.get("next") or request.GET.get("next") or _notifications_url()
    return redirect(next_url)


# =============================================================================
# Supervisor exports
# =============================================================================
def export_plan_excel(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
    week_obj = _resolve_week_or_404(week_no, allow_inactive=False)

    plan = get_object_or_404(
        Plan.objects.select_related("supervisor", "week").prefetch_related("days__school"),
        supervisor=supervisor,
        week=week_obj,
    )
    wb = _build_plan_excel_workbook(plan)
    filename = f"خطة_الأسبوع_{week_obj.week_no}_{_sup_nid_value(supervisor)}.xlsx"
    return _excel_response(wb, filename)


def export_plan_planned_schools_excel(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    tracking = _plan_school_tracking(supervisor)

    wb = _build_supervisor_school_list_workbook(
        supervisor=supervisor,
        schools=tracking["planned_schools"],
        report_title="المدارس المدرجة في الخطة",
        report_scope_label="من الأسبوع الأول إلى آخر أسبوع في الخطة",
    )
    filename = f"المدارس_المدرجة_جميع_الأسابيع_{_sup_nid_value(supervisor)}.xlsx"
    return _excel_response(wb, filename)


def export_plan_unplanned_schools_excel(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    tracking = _plan_school_tracking(supervisor)

    wb = _build_supervisor_school_list_workbook(
        supervisor=supervisor,
        schools=tracking["unplanned_schools"],
        report_title="المدارس غير المدرجة في الخطة",
        report_scope_label="من الأسبوع الأول إلى آخر أسبوع في الخطة",
    )
    filename = f"المدارس_غير_المدرجة_جميع_الأسابيع_{_sup_nid_value(supervisor)}.xlsx"
    return _excel_response(wb, filename)


def export_supervisor_assignments_excel(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    wb = _build_supervisor_assignments_excel_workbook(supervisor)
    filename = f"المدارس_المسندة_{_sup_nid_value(supervisor)}.xlsx"
    return _excel_response(wb, filename)


# =============================================================================
# Admin schools
# =============================================================================
@admin_only_view
def admin_school_list_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    gender = _cell_str(request.GET.get("gender"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_active = request.GET.get("active", "1") == "1"

    assignment_status = _cell_str(request.GET.get("assignment") or "all")
    if assignment_status not in ("all", "assigned", "unassigned"):
        assignment_status = "all"

    base_qs = School.objects.select_related("sector").order_by("name")

    if only_active:
        base_qs = base_qs.filter(is_active=True)

    if q:
        base_qs = base_qs.filter(
            Q(name__icontains=q) | Q(stat_code__icontains=q)
        )

    if gender in ("boys", "girls"):
        base_qs = base_qs.filter(gender=gender)

    if sector_id:
        base_qs = base_qs.filter(sector_id=sector_id)

    assigned_school_ids = set(
        Assignment.objects.filter(
            is_active=True,
            school_id__isnull=False,
        ).values_list("school_id", flat=True)
    )

    kpi_total = base_qs.count()
    kpi_active = base_qs.filter(is_active=True).count()
    kpi_inactive = base_qs.filter(is_active=False).count()
    kpi_assigned = base_qs.filter(id__in=assigned_school_ids).count()
    kpi_unassigned = base_qs.exclude(id__in=assigned_school_ids).count()

    qs = base_qs

    if assignment_status == "assigned":
        qs = qs.filter(id__in=assigned_school_ids)
    elif assignment_status == "unassigned":
        qs = qs.exclude(id__in=assigned_school_ids)

    sectors = Sector.objects.filter(is_active=True).order_by("name")

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    schools = list(page_obj.object_list)
    school_ids = [school.id for school in schools]

    active_assignments = (
        Assignment.objects.filter(
            is_active=True,
            school_id__in=school_ids,
        )
        .select_related("supervisor")
        .order_by("school_id", "id")
    )

    assignment_by_school: dict[int, Assignment] = {}
    assignment_count_by_school: dict[int, int] = {}

    for assignment in active_assignments:
        school_id = assignment.school_id
        assignment_count_by_school[school_id] = assignment_count_by_school.get(school_id, 0) + 1

        # في حال وجود أكثر من إسناد نشط لنفس المدرسة، نعرض أول إسناد
        # ونمرر العدد للقالب لإظهار تنبيه خفيف.
        if school_id not in assignment_by_school:
            assignment_by_school[school_id] = assignment

    for school in schools:
        assignment = assignment_by_school.get(school.id)
        school.active_assignment = assignment
        school.assigned_supervisor = assignment.supervisor if assignment else None
        school.active_assignment_count = assignment_count_by_school.get(school.id, 0)

    return render(
        request,
        "visits/admin_school_list.html",
        {
            "rows": schools,
            "page_obj": page_obj,
            "q": q,
            "gender": gender,
            "sector_id": sector_id,
            "only_active": only_active,
            "assignment_status": assignment_status,
            "sectors": sectors,
            "kpi_total": kpi_total,
            "kpi_active": kpi_active,
            "kpi_inactive": kpi_inactive,
            "kpi_assigned": kpi_assigned,
            "kpi_unassigned": kpi_unassigned,
        },
    )


@admin_only_view
@require_POST
def admin_school_save_view(request: HttpRequest) -> HttpResponse:
    school_id = _safe_int(request.POST.get("school_id") or 0, default=0)

    name = _cell_str(request.POST.get("name"))
    stat_code = _cell_str(request.POST.get("stat_code"))
    gender = _cell_str(request.POST.get("gender"))
    sector_id = _safe_int(request.POST.get("sector") or request.POST.get("sector_id") or 0, default=0)
    is_active = _bool_from_post(request.POST.get("is_active"), default=True)

    if not name:
        messages.error(request, "اسم المدرسة مطلوب.")
        return redirect("visits:admin_school_list")

    sector = Sector.objects.filter(id=sector_id).first() if sector_id else None

    if school_id:
        school = get_object_or_404(School, id=school_id)
        school.name = name
        if hasattr(school, "stat_code"):
            school.stat_code = stat_code
        if hasattr(school, "gender") and gender:
            school.gender = gender
        if hasattr(school, "sector"):
            school.sector = sector
        if hasattr(school, "is_active"):
            school.is_active = is_active
        school.save()
        messages.success(request, "تم تحديث المدرسة بنجاح.")
    else:
        data = {"name": name}
        if hasattr(School, "stat_code"):
            data["stat_code"] = stat_code
        if hasattr(School, "gender"):
            data["gender"] = gender or "boys"
        if hasattr(School, "sector"):
            data["sector"] = sector
        if hasattr(School, "is_active"):
            data["is_active"] = is_active
        School.objects.create(**data)
        messages.success(request, "تمت إضافة المدرسة بنجاح.")

    return redirect("visits:admin_school_list")


@admin_only_view
@require_POST
def admin_school_toggle_active_view(request: HttpRequest, school_id: int) -> HttpResponse:
    school = get_object_or_404(School, id=school_id)
    school.is_active = not bool(school.is_active)
    school.save(update_fields=["is_active"])
    messages.success(request, "تم تحديث حالة المدرسة.")
    return redirect("visits:admin_school_list")


# =============================================================================
# Admin supervisors
# =============================================================================
@admin_only_view
def admin_supervisor_list_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    gender = _cell_str(request.GET.get("gender"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_active = request.GET.get("active", "1") == "1"

    qs = (
        Supervisor.objects.select_related("sector")
        .annotate(
            active_assignment_count=Count("assignments", filter=Q(assignments__is_active=True), distinct=True),
            total_assignment_count=Count("assignments", distinct=True),
        )
        .order_by("full_name")
    )

    if only_active:
        qs = qs.filter(is_active=True)
    if q:
        qs = qs.filter(Q(full_name__icontains=q) | Q(national_id__icontains=q))
    if gender in ("boys", "girls"):
        qs = qs.filter(gender=gender)
    if sector_id:
        qs = qs.filter(sector_id=sector_id)

    sectors = Sector.objects.filter(is_active=True).order_by("name")
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/admin_supervisor_list.html",
        {
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "gender": gender,
            "sector_id": sector_id,
            "only_active": only_active,
            "sectors": sectors,
            "kpi_total": qs.count(),
            "kpi_active": qs.filter(is_active=True).count(),
            "kpi_inactive": qs.filter(is_active=False).count(),
        },
    )


@admin_only_view
@require_POST
def admin_supervisor_save_view(request: HttpRequest) -> HttpResponse:
    supervisor_id = _safe_int(request.POST.get("supervisor_id") or 0, default=0)

    full_name = _cell_str(request.POST.get("full_name"))
    national_id = _cell_str(request.POST.get("national_id"))
    mobile = _cell_str(request.POST.get("mobile"))
    email = _cell_str(request.POST.get("email")).lower()
    gender = _cell_str(request.POST.get("gender"))
    sector_id = _safe_int(request.POST.get("sector") or request.POST.get("sector_id") or 0, default=0)
    is_active = _bool_from_post(request.POST.get("is_active"), default=True)
    email_notifications_enabled = _bool_from_post(request.POST.get("email_notifications_enabled"), default=True)

    if not full_name:
        messages.error(request, "اسم المشرف مطلوب.")
        return redirect("visits:admin_supervisor_list")

    sector = Sector.objects.filter(id=sector_id).first() if sector_id else None

    if supervisor_id:
        supervisor = get_object_or_404(Supervisor, id=supervisor_id)
        supervisor.full_name = full_name
        if hasattr(supervisor, "national_id"):
            supervisor.national_id = national_id
        if hasattr(supervisor, "mobile"):
            supervisor.mobile = mobile
        if hasattr(supervisor, "email"):
            supervisor.email = email or None
        if hasattr(supervisor, "email_notifications_enabled"):
            supervisor.email_notifications_enabled = email_notifications_enabled
        if hasattr(supervisor, "email_verified"):
            supervisor.email_verified = False if email else False
        if hasattr(supervisor, "gender") and gender:
            supervisor.gender = gender
        if hasattr(supervisor, "sector"):
            supervisor.sector = sector
        if hasattr(supervisor, "is_active"):
            supervisor.is_active = is_active
        supervisor.save()
        messages.success(request, "تم تحديث المشرف بنجاح.")
    else:
        data = {"full_name": full_name}
        if hasattr(Supervisor, "national_id"):
            data["national_id"] = national_id
        if hasattr(Supervisor, "mobile"):
            data["mobile"] = mobile
        if hasattr(Supervisor, "gender"):
            data["gender"] = gender or "boys"
        if hasattr(Supervisor, "sector"):
            data["sector"] = sector
        if hasattr(Supervisor, "is_active"):
            data["is_active"] = is_active
        Supervisor.objects.create(**data)
        messages.success(request, "تمت إضافة المشرف بنجاح.")

    return redirect("visits:admin_supervisor_list")


@admin_only_view
@require_POST
def admin_supervisor_toggle_active_view(request: HttpRequest, supervisor_id: int) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)
    supervisor.is_active = not bool(supervisor.is_active)
    supervisor.save(update_fields=["is_active"])
    messages.success(request, "تم تحديث حالة المشرف.")
    return redirect("visits:admin_supervisor_list")


def _supervisor_assignment_scope(supervisor: Supervisor):
    qs = School.objects.filter(is_active=True)
    gender = getattr(supervisor, "gender", "")
    if gender in ("boys", "girls"):
        qs = qs.filter(gender=gender)
    if getattr(supervisor, "sector_id", None):
        qs = qs.filter(sector_id=supervisor.sector_id)
    return qs.order_by("name")


def _globally_assigned_school_ids(*, exclude_supervisor_id: int | None = None) -> set[int]:
    qs = Assignment.objects.filter(is_active=True).exclude(school_id__isnull=True)
    if exclude_supervisor_id:
        qs = qs.exclude(supervisor_id=exclude_supervisor_id)
    return set(qs.values_list("school_id", flat=True))


def _find_existing_school_for_assignment(*, stat_code: str, name: str, gender: str, sector_id: int | None):
    qs = School.objects.all()
    if stat_code:
        obj = qs.filter(stat_code=stat_code).first()
        if obj:
            return obj
    if name:
        qs = qs.filter(name=name)
        if gender in ("boys", "girls"):
            qs = qs.filter(gender=gender)
        if sector_id:
            qs = qs.filter(sector_id=sector_id)
        return qs.first()
    return None


# =============================================================================
# Admin assignments
# =============================================================================
@admin_only_view
def admin_assignments_overview_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    gender = _cell_str(request.GET.get("gender"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_active = request.GET.get("active", "1") == "1"
    workload = _cell_str(request.GET.get("workload") or "all") or "all"

    valid_workloads = {"all", "zero", "light", "balanced", "high", "very_high"}
    if workload not in valid_workloads:
        workload = "all"

    supervisors = (
        Supervisor.objects.select_related("sector")
        .annotate(
            active_assignment_count=Count(
                "assignments",
                filter=Q(assignments__is_active=True, assignments__school__is_active=True),
                distinct=True,
            ),
            total_assignment_count=Count("assignments", distinct=True),
        )
        .order_by("full_name")
    )

    if only_active:
        supervisors = supervisors.filter(is_active=True)

    if q:
        supervisors = supervisors.filter(
            Q(full_name__icontains=q) | Q(national_id__icontains=q)
        )

    if gender in ("boys", "girls"):
        supervisors = supervisors.filter(gender=gender)

    if sector_id:
        supervisors = supervisors.filter(sector_id=sector_id)

    if workload == "zero":
        supervisors = supervisors.filter(active_assignment_count=0)
    elif workload == "light":
        supervisors = supervisors.filter(
            active_assignment_count__gte=1,
            active_assignment_count__lte=10,
        )
    elif workload == "balanced":
        supervisors = supervisors.filter(
            active_assignment_count__gte=11,
            active_assignment_count__lte=25,
        )
    elif workload == "high":
        supervisors = supervisors.filter(
            active_assignment_count__gte=26,
            active_assignment_count__lte=35,
        )
    elif workload == "very_high":
        supervisors = supervisors.filter(active_assignment_count__gte=36)

    sectors = Sector.objects.filter(is_active=True).order_by("name")

    assigned_school_ids = _globally_assigned_school_ids()

    unassigned_schools_qs = School.objects.exclude(id__in=assigned_school_ids)
    if only_active:
        unassigned_schools_qs = unassigned_schools_qs.filter(is_active=True)
    if q:
        unassigned_schools_qs = unassigned_schools_qs.filter(
            Q(name__icontains=q) | Q(stat_code__icontains=q)
        )
    if gender in ("boys", "girls"):
        unassigned_schools_qs = unassigned_schools_qs.filter(gender=gender)
    if sector_id:
        unassigned_schools_qs = unassigned_schools_qs.filter(sector_id=sector_id)

    return render(
        request,
        "visits/admin_assignments_overview.html",
        {
            "rows": supervisors,
            "q": q,
            "gender": gender,
            "sector_id": sector_id,
            "only_active": only_active,
            "workload": workload,
            "sectors": sectors,
            "kpi_supervisors": supervisors.count(),
            "kpi_assignments": Assignment.objects.filter(
                is_active=True,
                school__is_active=True,
            ).count(),
            "kpi_schools": School.objects.filter(is_active=True).count(),
            "kpi_unassigned_schools": unassigned_schools_qs.count(),
        },
    )




@admin_only_view
def admin_export_unassigned_schools_excel(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    gender = _cell_str(request.GET.get("gender"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_active = request.GET.get("active", "1") == "1"

    wb = _build_unassigned_schools_excel_workbook(
        q=q,
        gender=gender,
        sector_id=sector_id,
        only_active=only_active,
    )
    return _excel_response(wb, "المدارس_غير_المسندة.xlsx")

@admin_only_view
def admin_export_unassigned_schools_excel(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    gender = _cell_str(request.GET.get("gender"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_active = request.GET.get("active", "1") == "1"

    wb = _build_unassigned_schools_excel_workbook(
        q=q,
        gender=gender,
        sector_id=sector_id,
        only_active=only_active,
    )
    return _excel_response(wb, "المدارس_غير_المسندة.xlsx")

@admin_only_view
def admin_supervisor_assignments_view(request: HttpRequest, supervisor_id: int) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor.objects.select_related("sector"), id=supervisor_id)
    q = _cell_str(request.GET.get("q"))

    assigned_qs = (
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
            school__isnull=False,
            school__is_active=True,
        )
        .select_related("school")
        .order_by("school__name")
    )

    if q:
        assigned_qs = assigned_qs.filter(
            Q(school__name__icontains=q) | Q(school__stat_code__icontains=q)
        )

    assigned_items = list(assigned_qs)
    school_ids = [a.school_id for a in assigned_items if a.school_id]
    principals_map = {p.school_id: p for p in Principal.objects.filter(school_id__in=school_ids)}

    rows = []
    for i, assignment in enumerate(assigned_items, start=1):
        school = assignment.school
        principal = principals_map.get(school.id) if school else None
        rows.append(
            {
                "index": i,
                "assignment": assignment,
                "school": school,
                "principal": principal,
                "is_active": assignment.is_active,
                "stat_code": getattr(school, "stat_code", "") or "—",
                "school_name": getattr(school, "name", "") or "—",
                "gender_label": _gender_label(getattr(school, "gender", "")),
                "principal_name": getattr(principal, "full_name", "") or "—",
                "principal_mobile": getattr(principal, "mobile", "") or "—",
            }
        )

    available_scope_qs = _supervisor_assignment_scope(supervisor)
    busy_school_ids = _globally_assigned_school_ids(exclude_supervisor_id=supervisor.id)
    available_schools_qs = available_scope_qs.exclude(id__in=busy_school_ids)

    if q:
        available_schools_qs = available_schools_qs.filter(
            Q(name__icontains=q) | Q(stat_code__icontains=q)
        )

    available_schools_qs = available_schools_qs.order_by("name")
    available_schools = list(available_schools_qs[:100])

    total_assigned_count = Assignment.objects.filter(
        supervisor=supervisor,
        is_active=True,
        school__isnull=False,
        school__is_active=True,
    ).count()
    total_available_count = available_scope_qs.exclude(id__in=busy_school_ids).count()

    return render(
        request,
        "visits/admin_supervisor_assignments.html",
        {
            "supervisor": supervisor,
            "supervisor_id": supervisor.id,
            "sup": supervisor,
            "assigned_items": assigned_items,
            "available_schools": available_schools,
            "q": q,
            "stats": {
                "assigned_count": total_assigned_count,
                "available_count": total_available_count,
                "search_count": len(available_schools),
            },
            # توافق مع أي قالب قديم
            "rows": rows,
            "kpi_total": total_assigned_count,
            "kpi_active": total_assigned_count,
            "kpi_inactive": 0,
            "kpi_available": total_available_count,
        },
    )


@admin_only_view
@require_POST
def admin_add_assignment_view(request: HttpRequest, supervisor_id: int) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)

    next_url = (
        request.POST.get("next")
        or reverse("visits:admin_supervisor_assignments", args=[supervisor.id])
    )

    action = _cell_str(request.POST.get("action"))
    school_id = _safe_int(request.POST.get("school_id") or request.POST.get("school") or 0, default=0)
    new_school_name = _cell_str(request.POST.get("new_school_name") or request.POST.get("school_name") or request.POST.get("name"))
    new_stat_code = _cell_str(request.POST.get("new_stat_code") or request.POST.get("stat_code"))
    new_gender = _cell_str(request.POST.get("new_gender") or request.POST.get("gender")) or getattr(supervisor, "gender", "")

    school: School | None = None

    if action == "add_existing":
        if not school_id:
            messages.error(request, "يرجى اختيار مدرسة متاحة.")
            return redirect(next_url)
        school = get_object_or_404(School, id=school_id, is_active=True)
    elif action == "create_school":
        if not new_school_name:
            messages.error(request, "اسم المدرسة مطلوب.")
            return redirect(next_url)
        if not new_stat_code:
            messages.error(request, "الرقم الإحصائي مطلوب عند إضافة مدرسة جديدة.")
            return redirect(next_url)
        if new_gender not in ("boys", "girls"):
            new_gender = getattr(supervisor, "gender", "") or "boys"
        school = _find_existing_school_for_assignment(
            stat_code=new_stat_code,
            name=new_school_name,
            gender=new_gender,
            sector_id=getattr(supervisor, "sector_id", None),
        )
        if school is None:
            create_data = {
                "name": new_school_name,
                "stat_code": new_stat_code,
                "gender": new_gender,
                "is_active": True,
            }
            if getattr(supervisor, "sector_id", None):
                create_data["sector_id"] = supervisor.sector_id
            school = School.objects.create(**create_data)
    else:
        if school_id:
            school = get_object_or_404(School, id=school_id, is_active=True)
        elif new_school_name:
            if not new_stat_code:
                messages.error(request, "الرقم الإحصائي مطلوب عند إضافة مدرسة جديدة.")
                return redirect(next_url)
            if new_gender not in ("boys", "girls"):
                new_gender = getattr(supervisor, "gender", "") or "boys"
            school = _find_existing_school_for_assignment(
                stat_code=new_stat_code,
                name=new_school_name,
                gender=new_gender,
                sector_id=getattr(supervisor, "sector_id", None),
            )
            if school is None:
                create_data = {
                    "name": new_school_name,
                    "stat_code": new_stat_code,
                    "gender": new_gender,
                    "is_active": True,
                }
                if getattr(supervisor, "sector_id", None):
                    create_data["sector_id"] = supervisor.sector_id
                school = School.objects.create(**create_data)
        else:
            messages.error(request, "يرجى اختيار مدرسة أو إدخال بيانات مدرسة جديدة.")
            return redirect(next_url)

    if school is None:
        messages.error(request, "تعذر تحديد المدرسة المطلوبة.")
        return redirect(next_url)

    if getattr(supervisor, "gender", "") in ("boys", "girls") and getattr(school, "gender", "") not in ("", supervisor.gender):
        messages.error(request, "لا يمكن إسناد مدرسة بجنس مختلف عن جنس المشرف.")
        return redirect(next_url)

    if getattr(supervisor, "sector_id", None) and getattr(school, "sector_id", None) and school.sector_id != supervisor.sector_id:
        messages.error(request, "لا يمكن إسناد مدرسة من قطاع مختلف عن قطاع المشرف.")
        return redirect(next_url)

    existing_elsewhere = (
        Assignment.objects.filter(school=school, is_active=True, school__is_active=True)
        .exclude(supervisor=supervisor)
        .select_related("supervisor")
        .first()
    )
    if existing_elsewhere:
        messages.error(
            request,
            f"هذه المدرسة مسندة بالفعل إلى المشرف {existing_elsewhere.supervisor.full_name}.",
        )
        return redirect(next_url)

    assignment, created = Assignment.objects.get_or_create(
        supervisor=supervisor,
        school=school,
        defaults={"is_active": True},
    )

    if created:
        messages.success(request, "تمت إضافة الإسناد بنجاح.")
    else:
        if not assignment.is_active:
            assignment.is_active = True
            assignment.save(update_fields=["is_active"])
            messages.success(request, "تمت إعادة تفعيل الإسناد بنجاح.")
        else:
            messages.info(request, "هذا الإسناد موجود مسبقًا.")

    return redirect(next_url)


@admin_only_view
@require_POST
def admin_delete_assignment_view(request: HttpRequest, assignment_id: int) -> HttpResponse:
    assignment = get_object_or_404(Assignment.objects.select_related("supervisor"), id=assignment_id)

    next_url = (
        request.POST.get("next")
        or reverse("visits:admin_supervisor_assignments", args=[assignment.supervisor_id])
    )

    if hasattr(assignment, "is_active"):
        assignment.is_active = False
        assignment.save(update_fields=["is_active"])
        messages.success(request, "تم إلغاء الإسناد بنجاح.")
    else:
        assignment.delete()
        messages.success(request, "تم حذف الإسناد بنجاح.")

    return redirect(next_url)


# =============================================================================
# Admin sectors
# =============================================================================
@admin_only_view
def admin_sector_list_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    show_inactive = request.GET.get("inactive") == "1"

    sectors = (
        Sector.objects.annotate(
            schools_count=Count("schools", distinct=True),
            supervisors_count=Count("supervisors", distinct=True),
        )
        .order_by("name")
    )

    if q:
        sectors = sectors.filter(name__icontains=q)
    if not show_inactive:
        sectors = sectors.filter(is_active=True)

    return render(
        request,
        "visits/admin_sector_list.html",
        {
            "rows": sectors,
            "q": q,
            "show_inactive": show_inactive,
        },
    )


@admin_only_view
@require_POST
def admin_sector_save_view(request: HttpRequest) -> HttpResponse:
    sector_id = _safe_int(request.POST.get("sector_id") or 0, default=0)
    name = _cell_str(request.POST.get("name"))

    if not name:
        messages.error(request, "اسم القطاع مطلوب.")
        return redirect("visits:admin_sector_list")

    if sector_id:
        sector = get_object_or_404(Sector, id=sector_id)
        sector.name = name
        sector.save()
        messages.success(request, "تم تحديث بيانات القطاع بنجاح.")
    else:
        if Sector.objects.filter(name=name).exists():
            messages.error(request, "يوجد قطاع بهذا الاسم بالفعل.")
            return redirect("visits:admin_sector_list")
        Sector.objects.create(name=name, is_active=True)
        messages.success(request, "تمت إضافة القطاع بنجاح.")

    return redirect("visits:admin_sector_list")


@admin_only_view
@require_POST
def admin_sector_toggle_active_view(request: HttpRequest, sector_id: int) -> HttpResponse:
    sector = get_object_or_404(Sector, id=sector_id)
    sector.is_active = not bool(sector.is_active)
    sector.save(update_fields=["is_active"])

    if sector.is_active:
        messages.success(request, f"تم تفعيل القطاع «{sector.name}».")
    else:
        messages.success(request, f"تم تعطيل القطاع «{sector.name}».")
    return redirect("visits:admin_sector_list")


# =============================================================================
# Unlock request
# =============================================================================
@require_POST
def request_unlock_view(request: HttpRequest) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    plan_id = _safe_int(request.POST.get("plan") or 0, default=0)
    plan = get_object_or_404(Plan, id=plan_id, supervisor=supervisor)
    _inject_week_no(plan)

    if plan.status != Plan.STATUS_APPROVED:
        messages.info(request, "لا يمكن طلب فك اعتماد إلا لخطة معتمدة.")
        return redirect(_plan_url(plan.week.week_no))

    reason = _cell_str(request.POST.get("reason"))
    if not reason:
        reason = "طلب المشرف فك اعتماد الخطة لإجراء تعديل على بيانات الخطة الأسبوعية."

    req, created = UnlockRequest.objects.get_or_create(
        plan=plan,
        defaults={
            "reason": reason,
            "status": UnlockRequest.STATUS_PENDING,
            "resolved_at": None,
        },
    )

    if not created and req.status == UnlockRequest.STATUS_PENDING:
        messages.info(request, "يوجد طلب فك اعتماد سابق لهذه الخطة بانتظار الإدارة.")
        return redirect(_plan_url(plan.week.week_no))

    req.reason = reason
    req.status = UnlockRequest.STATUS_PENDING
    req.resolved_at = None
    req.save(update_fields=["reason", "status", "resolved_at"])

    plan.status = Plan.STATUS_UNLOCK_REQUESTED
    plan.save(update_fields=["status"])

    messages.success(request, "تم إرسال طلب فك الاعتماد.")
    return redirect(_plan_url(plan.week.week_no))


# =============================================================================
# Admin visit follow-up
# =============================================================================
def _build_supervisor_visit_followup_row(supervisor: Supervisor) -> dict[str, Any]:
    assigned_school_ids = set(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
            school__is_active=True,
        ).values_list("school_id", flat=True)
    )

    visited_school_ids = set(
        PlanDay.objects.filter(
            plan__supervisor=supervisor,
            visit_type=PlanDay.VISIT_IN,
            school__isnull=False,
        ).values_list("school_id", flat=True)
    )

    assigned_count = len(assigned_school_ids)
    visited_count = len(assigned_school_ids & visited_school_ids)
    remaining_count = max(assigned_count - visited_count, 0)
    progress_percent = round((visited_count / assigned_count) * 100, 1) if assigned_count else 0

    return {
        "supervisor": supervisor,
        "assigned_count": assigned_count,
        "visited_count": visited_count,
        "remaining_count": remaining_count,
        "progress_percent": progress_percent,
    }


def _build_global_visit_followup_stats() -> dict[str, Any]:
    supervisors = (
        Supervisor.objects.filter(is_active=True)
        .select_related("sector")
        .order_by("full_name")
    )

    rows: list[dict[str, Any]] = []
    total_assigned = 0
    total_planned = 0
    total_remaining = 0
    supervisors_with_remaining = 0

    for supervisor in supervisors:
        assigned_schools = list(
            School.objects.filter(
                assignments__supervisor=supervisor,
                assignments__is_active=True,
                is_active=True,
            )
            .distinct()
            .order_by("name")
        )
        assigned_school_ids = {s.id for s in assigned_schools}

        planned_schools = list(
            School.objects.filter(
                planday__plan__supervisor=supervisor,
                planday__school__isnull=False,
                is_active=True,
            )
            .distinct()
            .order_by("name")
        )
        planned_school_ids = {s.id for s in planned_schools}

        planned_school_ids &= assigned_school_ids
        planned_schools = [s for s in planned_schools if s.id in planned_school_ids]

        remaining_schools = [s for s in assigned_schools if s.id not in planned_school_ids]
        remaining_school_ids = {s.id for s in remaining_schools}

        assigned_count = len(assigned_school_ids)
        planned_count = len(planned_school_ids)
        remaining_count = len(remaining_school_ids)

        total_assigned += assigned_count
        total_planned += planned_count
        total_remaining += remaining_count

        if remaining_count > 0:
            supervisors_with_remaining += 1

        rows.append(
            {
                "supervisor": supervisor,
                "assigned_count": assigned_count,
                "planned_count": planned_count,
                "visited_count": planned_count,   # للإبقاء على التوافق مع القالب الحالي مؤقتًا
                "remaining_count": remaining_count,
                "progress_percent": round((planned_count / assigned_count) * 100, 1) if assigned_count else 0,
                "planned_schools": planned_schools,
                "remaining_schools": remaining_schools,
                "planned_school_names": "، ".join(s.name for s in planned_schools) or "—",
                "remaining_school_names": "، ".join(s.name for s in remaining_schools) or "—",
            }
        )

    return {
        "visit_followup_rows": rows,
        "visit_followup_total_assigned": total_assigned,
        "visit_followup_total_visited": total_planned,   # للإبقاء على التوافق مع القالب الحالي مؤقتًا
        "visit_followup_total_planned": total_planned,
        "visit_followup_total_remaining": total_remaining,
        "visit_followup_supervisors_with_remaining": supervisors_with_remaining,
    }


def _filter_visit_followup_rows(
    *,
    rows: list[dict[str, Any]],
    q: str = "",
    sector_id: int = 0,
    only_remaining: bool = False,
) -> list[dict[str, Any]]:
    filtered = list(rows)

    if q:
        q_lower = q.lower()
        filtered = [
            row for row in filtered
            if q_lower in ((row["supervisor"].full_name or "").lower())
            or q_lower in ((row["supervisor"].national_id or "").lower())
        ]

    if sector_id:
        filtered = [
            row for row in filtered
            if row["supervisor"].sector_id == sector_id
        ]

    if only_remaining:
        filtered = [
            row for row in filtered
            if row["remaining_count"] > 0
        ]

    return filtered


def _build_visit_followup_excel_workbook(rows, report_title: str = "متابعة الزيارات على مستوى جميع الأسابيع") -> Workbook:
    wb = Workbook()

    # =========================================
    # الورقة الأولى: ملخص
    # =========================================
    ws = wb.active
    ws.title = "ملخص المتابعة"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = report_title
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:H2")
    ws["A2"] = f"تاريخ التصدير: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    headers = [
        "م",
        "اسم المشرف",
        "رقم الهوية",
        "القطاع",
        "المدارس المسندة",
        "المدارس المدرجة في الخطة",
        "المتبقي",
        "نسبة الإنجاز",
    ]
    header_row = 4

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for index, row in enumerate(rows, start=1):
        supervisor = row["supervisor"]
        sector_name = supervisor.sector.name if getattr(supervisor, "sector", None) else "—"

        values = [
            index,
            supervisor.full_name or "—",
            supervisor.national_id or "—",
            sector_name,
            row["assigned_count"],
            row["planned_count"],
            row["remaining_count"],
            f'{row["progress_percent"]}%',
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 6, 7, 8) else right

        row_idx += 1

    for col_i, width in {
        1: 8,
        2: 28,
        3: 18,
        4: 22,
        5: 18,
        6: 24,
        7: 16,
        8: 16,
    }.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # =========================================
    # الورقة الثانية: التفاصيل بالأسماء
    # =========================================
    ws2 = wb.create_sheet(title="تفاصيل المدارس")
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"{report_title} — تفاصيل المدارس"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center
    ws2["A1"].fill = title_fill

    ws2.merge_cells("A2:F2")
    ws2["A2"] = f"تاريخ التصدير: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"
    ws2["A2"].font = bold_font
    ws2["A2"].alignment = center

    detail_headers = [
        "م",
        "اسم المشرف",
        "رقم الهوية",
        "القطاع",
        "المدارس المدرجة في الخطة",
        "المدارس المتبقية",
    ]
    detail_header_row = 4

    for col, header in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=detail_header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = detail_header_row + 1
    for index, row in enumerate(rows, start=1):
        supervisor = row["supervisor"]
        sector_name = supervisor.sector.name if getattr(supervisor, "sector", None) else "—"

        values = [
            index,
            supervisor.full_name or "—",
            supervisor.national_id or "—",
            sector_name,
            row.get("planned_school_names", "—"),
            row.get("remaining_school_names", "—"),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3) else right

        row_idx += 1

    for col_i, width in {
        1: 8,
        2: 26,
        3: 18,
        4: 20,
        5: 60,
        6: 60,
    }.items():
        ws2.column_dimensions[get_column_letter(col_i)].width = width

    return wb


@admin_only_view
def admin_visit_followup_dashboard_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_remaining = str(request.GET.get("only_remaining") or "").strip().lower() in ("1", "true", "on", "yes")

    stats = _build_global_visit_followup_stats()
    rows = _filter_visit_followup_rows(
        rows=stats["visit_followup_rows"],
        q=q,
        sector_id=sector_id,
        only_remaining=only_remaining,
    )

    sectors = Sector.objects.filter(is_active=True).order_by("name")

    return render(
        request,
        "visits/admin_visit_followup_dashboard.html",
        {
            **stats,
            "rows": rows,
            "q": q,
            "sector_id": sector_id,
            "only_remaining": only_remaining,
            "sectors": sectors,
        },
    )


@admin_only_view
def admin_visit_followup_export_excel_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    only_remaining = str(request.GET.get("only_remaining") or "").strip().lower() in ("1", "true", "on", "yes")

    stats = _build_global_visit_followup_stats()
    rows = _filter_visit_followup_rows(
        rows=stats["visit_followup_rows"],
        q=q,
        sector_id=sector_id,
        only_remaining=only_remaining,
    )

    title_parts = ["متابعة الزيارات على مستوى جميع الأسابيع"]
    if q:
        title_parts.append(f"بحث: {q}")

    if sector_id:
        sector = Sector.objects.filter(id=sector_id).first()
        if sector:
            title_parts.append(f"القطاع: {sector.name}")

    if only_remaining:
        title_parts.append("المتأخرون فقط")

    workbook = _build_visit_followup_excel_workbook(
        rows,
        report_title=" — ".join(title_parts),
    )
    return _excel_response(workbook, "visit_followup_report.xlsx")


@admin_only_view
@require_POST
def admin_notify_supervisor_visit_followup_view(request: HttpRequest, supervisor_id: int) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)
    row = _build_supervisor_visit_followup_row(supervisor)

    if row["remaining_count"] <= 0:
        messages.info(request, "لا توجد مدارس متبقية على هذا المشرف.")
        return redirect("visits:admin_visit_followup_dashboard")

    _notify_supervisor(
        supervisor=supervisor,
        plan=None,
        notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
        title="تنبيه بمتابعة الزيارات",
        message=(
            "نأمل استكمال الزيارات المتبقية. "
            f"عدد المدارس التي لم يتم تسجيل زيارتها حتى الآن: {row['remaining_count']}."
        ),
    )

    messages.success(request, f"تم إرسال تنبيه إلى المشرف {supervisor.full_name}.")
    return redirect("visits:admin_visit_followup_dashboard")


@admin_only_view
@require_POST
def admin_notify_all_supervisors_visit_followup_view(request: HttpRequest) -> HttpResponse:
    supervisors = Supervisor.objects.filter(is_active=True).order_by("full_name")
    notified = 0

    for supervisor in supervisors:
        row = _build_supervisor_visit_followup_row(supervisor)
        if row["remaining_count"] <= 0:
            continue

        _notify_supervisor(
            supervisor=supervisor,
            plan=None,
            notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
            title="تنبيه بمتابعة الزيارات",
            message=(
                "نأمل استكمال الزيارات المتبقية. "
                f"عدد المدارس التي لم يتم تسجيل زيارتها حتى الآن: {row['remaining_count']}."
            ),
        )
        notified += 1

    if notified:
        messages.success(request, f"تم إرسال {notified} تنبيهًا للمشرفين المتأخرين.")
    else:
        messages.info(request, "لا يوجد مشرفون متأخرون حاليًا.")

    return redirect("visits:admin_visit_followup_dashboard")


# =============================================================================
# Admin dashboard / detail / exports
# =============================================================================
@admin_only_view
def admin_dashboard_view(request: HttpRequest) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")

    weeks_qs = _get_all_weeks_qs() if show_all else _get_active_weeks_qs()
    week_choices = _build_week_choices(active_only=(not show_all))
    default_week = weeks_qs.first()

    if not default_week:
        messages.warning(request, "لا يوجد أسابيع في جدول PlanWeek. أضف الأسابيع أولًا.")
        return render(request, "visits/admin_dashboard.html", {"rows": [], "page_obj": None})

    week_no = _safe_int(request.GET.get("week") or default_week.week_no, default=default_week.week_no)
    if not _week_exists(week_no):
        week_no = default_week.week_no

    week_obj = _resolve_week_or_404(week_no, allow_inactive=show_all)
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "all").strip().lower()
    if status not in ("all", "approved", "draft", "unlock", "not_full"):
        status = "all"

    page_sizes = [10, 20, 30, 50]
    page_size = _safe_int(request.GET.get("ps") or 10, default=10)
    if page_size not in page_sizes:
        page_size = 10

    page = _safe_int(request.GET.get("page") or 1, default=1)

    base_qs = (
        Plan.objects.filter(week=week_obj)
        .select_related("supervisor", "week")
        .prefetch_related("days__school")
        .order_by("supervisor__full_name")
    )

    if q:
        base_qs = base_qs.filter(Q(supervisor__full_name__icontains=q) | Q(supervisor__national_id__icontains=q))

    plans_all = list(base_qs)
    for p in plans_all:
        _inject_week_no(p)

    supervisor_ids = [p.supervisor_id for p in plans_all]
    assignment_counts = {}
    if supervisor_ids:
        assignment_counts = dict(
            Assignment.objects.filter(
                supervisor_id__in=supervisor_ids,
                is_active=True,
                school__is_active=True,
            )
            .values("supervisor_id")
            .annotate(c=Count("id"))
            .values_list("supervisor_id", "c")
        )

    kpi_total = len(plans_all)
    kpi_approved = sum(1 for p in plans_all if p.status == Plan.STATUS_APPROVED)
    kpi_drafts = sum(1 for p in plans_all if p.status == Plan.STATUS_DRAFT)
    kpi_unlock = sum(1 for p in plans_all if p.status == Plan.STATUS_UNLOCK_REQUESTED)
    kpi_filled = sum(1 for p in plans_all if p.is_fully_filled())
    kpi_not_filled = kpi_total - kpi_filled

    plans_filtered = plans_all
    if status == "approved":
        plans_filtered = [p for p in plans_all if p.status == Plan.STATUS_APPROVED]
    elif status == "draft":
        plans_filtered = [p for p in plans_all if p.status == Plan.STATUS_DRAFT]
    elif status == "unlock":
        plans_filtered = [p for p in plans_all if p.status == Plan.STATUS_UNLOCK_REQUESTED]
    elif status == "not_full":
        plans_filtered = [p for p in plans_all if not p.is_fully_filled()]

    rows = []
    for p in plans_filtered:
        day_map = {d.weekday: d for d in p.days.all()}
        filled = sum(1 for wd, _ in WEEKDAYS if _day_is_filled(day_map.get(wd)))
        rows.append(
            {
                "plan": p,
                "sup": p.supervisor,
                "filled": filled,
                "is_full": filled == 5,
                "day_map": day_map,
                "assignment_count": assignment_counts.get(p.supervisor_id, 0),
            }
        )

    paginator = Paginator(rows, page_size)
    page_obj = paginator.get_page(page)
    chart_data = _build_chart_counts(week_obj)
    global_visit_data = _build_dashboard_global_visit_counts(include_breaks=show_all)
    site_setting = _get_site_setting()
    _maintenance_is_active(site_setting, persist=True)

    return render(
        request,
        "visits/admin_dashboard.html",
        {
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "q": q,
            "status": status,
            "page_size": page_size,
            "page_sizes": page_sizes,
            "weekdays": WEEKDAYS,
            "week_choices": week_choices,
            "kpi_total": kpi_total,
            "kpi_approved": kpi_approved,
            "kpi_drafts": kpi_drafts,
            "kpi_unlock": kpi_unlock,
            "kpi_filled": kpi_filled,
            "kpi_not_filled": kpi_not_filled,
            "show_all": show_all,
            "week_letter": WeeklyLetterLink.objects.filter(week=week_obj).first(),
            "site_setting": site_setting,
            **_get_assignment_dashboard_context(),
            **chart_data,
            **global_visit_data,
        },
    )


@admin_only_view
def admin_plan_detail_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(
        Plan.objects.select_related("supervisor", "week").prefetch_related("days__school"),
        id=plan_id,
    )
    _inject_week_no(plan)

    day_map = {d.weekday: d for d in plan.days.all().select_related("school")}
    filled = sum(1 for wd, _ in WEEKDAYS if _day_is_filled(day_map.get(wd)))
    day_dates = _build_day_dates_from_week(plan.week)
    visit_counts = _plan_visit_counts(plan)

    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "all").strip()
    ps = _safe_int(request.GET.get("ps") or 10, default=10)
    page = _safe_int(request.GET.get("page") or 1, default=1)

    next_url = (request.GET.get("next") or "").strip()
    if not next_url:
        next_url = _admin_dashboard_url(
            plan.week.week_no,
            show_all=show_all,
            q=q,
            status=status_filter,
            ps=ps,
            page=page,
        )

    return render(
        request,
        "visits/admin_plan_detail.html",
        {
            "plan": plan,
            "sup": plan.supervisor,
            "weekdays": WEEKDAYS,
            "day_map": day_map,
            "filled": filled,
            "week": plan.week.week_no,
            "week_obj": plan.week,
            "day_dates": day_dates,
            "next_url": next_url,
            "back_to": "detail",
            "show_all": show_all,
            "q": q,
            "status": status_filter,
            "page_size": ps,
            "page_number": page,
            "assigned_count": visit_counts["assigned_count"],
            "visited_count": visit_counts["visited_count"],
            "remaining_count": visit_counts["remaining_count"],
        },
    )


@admin_only_view
def admin_plan_export_excel(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(
        Plan.objects.select_related("supervisor", "week").prefetch_related("days__school"),
        id=plan_id,
    )
    wb = _build_plan_excel_workbook(plan)
    filename = f"خطة_الأسبوع_{plan.week.week_no}_{_sup_nid_value(plan.supervisor)}.xlsx"
    return _excel_response(wb, filename)


@admin_only_view
def admin_export_supervisor_assignments_excel(request: HttpRequest, supervisor_id: int) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor, id=supervisor_id, is_active=True)
    wb = _build_supervisor_assignments_excel_workbook(supervisor)
    filename = f"المدارس_المسندة_{_sup_nid_value(supervisor)}.xlsx"
    return _excel_response(wb, filename)




def _build_week_visit_summary_rows(week_obj: PlanWeek) -> list[dict[str, Any]]:
    plans = (
        Plan.objects.filter(week=week_obj)
        .select_related("supervisor", "supervisor__sector", "week")
        .prefetch_related("days__school")
        .order_by("supervisor__full_name")
    )

    rows: list[dict[str, Any]] = []

    for plan in plans:
        assigned_school_ids = set(
            Assignment.objects.filter(
                supervisor=plan.supervisor,
                is_active=True,
                school__is_active=True,
            ).values_list("school_id", flat=True)
        )

        visited_school_ids = set(
            PlanDay.objects.filter(
                plan=plan,
                visit_type=getattr(PlanDay, "VISIT_IN", "in"),
                school_id__isnull=False,
                visited=True,
            ).values_list("school_id", flat=True)
        )

        visited_school_ids &= assigned_school_ids
        unvisited_school_ids = assigned_school_ids - visited_school_ids

        visited_schools = list(
            School.objects.filter(id__in=visited_school_ids, is_active=True).order_by("name")
        )
        unvisited_schools = list(
            School.objects.filter(id__in=unvisited_school_ids, is_active=True).order_by("name")
        )

        rows.append(
            {
                "supervisor": plan.supervisor,
                "plan": plan,
                "sector_name": getattr(getattr(plan.supervisor, "sector", None), "name", "") or "—",
                "assigned_count": len(assigned_school_ids),
                "visited_count": len(visited_school_ids),
                "unvisited_count": len(unvisited_school_ids),
                "visited_school_names": [s.name for s in visited_schools],
                "unvisited_school_names": [s.name for s in unvisited_schools],
            }
        )

    return rows


def _build_week_visit_summary_excel_workbook(week_obj: PlanWeek) -> Workbook:
    rows = _build_week_visit_summary_rows(week_obj)

    wb = Workbook()
    ws = wb.active
    ws.title = f"ملخص الأسبوع {week_obj.week_no}"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=12)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")
    visited_fill = PatternFill("solid", fgColor="ECFDF3")
    unvisited_fill = PatternFill("solid", fgColor="FFF7ED")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_assigned = sum(r["assigned_count"] for r in rows)
    total_visited = sum(r["visited_count"] for r in rows)
    total_unvisited = sum(r["unvisited_count"] for r in rows)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"تقرير المدارس المزارة وغير المزارة — الأسبوع {week_obj.week_no}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"إجمالي المدارس المسندة: {total_assigned}  |  "
        f"المزارة: {total_visited}  |  "
        f"غير المزارة: {total_unvisited}"
    )
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "القطاع",
        "عدد المدارس المسندة",
        "عدد المدارس المزارة",
        "عدد المدارس غير المزارة",
        "تفاصيل المدارس غير المزارة",
    ]

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for i, row in enumerate(rows, start=1):
        values = [
            i,
            row["supervisor"].full_name,
            _sup_nid_value(row["supervisor"]),
            row["sector_name"],
            row["assigned_count"],
            row["visited_count"],
            row["unvisited_count"],
            "، ".join(row["unvisited_school_names"]) if row["unvisited_school_names"] else "—",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 6, 7) else right

            if col == 6:
                cell.fill = visited_fill
            elif col == 7:
                cell.fill = unvisited_fill

        row_idx += 1

    for col_i, width in {
        1: 8,
        2: 26,
        3: 18,
        4: 20,
        5: 18,
        6: 18,
        7: 20,
        8: 60,
    }.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    return wb


@admin_only_view
def admin_export_week_excel(request: HttpRequest) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")
    weeks_qs = _get_all_weeks_qs() if show_all else _get_active_weeks_qs()
    default_week = weeks_qs.first()

    if not default_week:
        messages.warning(request, "لا يوجد أسابيع متاحة للتصدير.")
        return redirect("visits:admin_dashboard")

    week_no = _safe_int(request.GET.get("week") or default_week.week_no, default=default_week.week_no)
    if not _week_exists(week_no):
        week_no = default_week.week_no

    week_obj = _resolve_week_or_404(week_no, allow_inactive=show_all)
    plans = (
        Plan.objects.filter(week=week_obj)
        .select_related("supervisor", "week")
        .prefetch_related("days__school")
        .order_by("supervisor__full_name")
    )
    wb = _build_admin_week_excel_workbook(week_obj, plans)
    filename = f"خطط_الأسبوع_{week_obj.week_no}.xlsx"
    return _excel_response(wb, filename)


@admin_only_view
def admin_export_all_plans_excel(request: HttpRequest) -> HttpResponse:
    """تصدير جميع خطط المشرفين لجميع الأسابيع في ملف Excel واحد."""
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes", "on")
    q = _cell_str(request.GET.get("q"))
    status = _cell_str(request.GET.get("status") or "all").lower()
    supervisor_id = _safe_int(request.GET.get("supervisor") or 0, default=0)
    from_week = _safe_int(request.GET.get("from_week") or 0, default=0)
    to_week = _safe_int(request.GET.get("to_week") or 0, default=0)

    if status not in ("all", "approved", "draft", "unlock", "not_full"):
        status = "all"

    plans_qs = (
        Plan.objects
        .select_related("supervisor", "supervisor__sector", "week")
        .prefetch_related("days__school")
        .order_by("supervisor__full_name", "week__week_no")
    )

    if not show_all:
        plans_qs = plans_qs.filter(week__is_break=False)

    if supervisor_id:
        plans_qs = plans_qs.filter(supervisor_id=supervisor_id)

    if from_week:
        plans_qs = plans_qs.filter(week__week_no__gte=from_week)

    if to_week:
        plans_qs = plans_qs.filter(week__week_no__lte=to_week)

    if status == "approved":
        plans_qs = plans_qs.filter(status=Plan.STATUS_APPROVED)
    elif status == "draft":
        plans_qs = plans_qs.filter(status=Plan.STATUS_DRAFT)
    elif status == "unlock":
        plans_qs = plans_qs.filter(status=Plan.STATUS_UNLOCK_REQUESTED)

    if q:
        plans_qs = plans_qs.filter(
            Q(supervisor__full_name__icontains=q)
            | Q(supervisor__national_id__icontains=q)
            | Q(days__school__name__icontains=q)
            | Q(days__school__stat_code__icontains=q)
        ).distinct()

    plans = list(plans_qs)

    if status == "not_full":
        plans = [plan for plan in plans if not plan.is_fully_filled()]

    if not plans:
        messages.warning(request, "لا توجد خطط مطابقة لمعايير التصدير.")
        return redirect("visits:admin_dashboard")

    wb = _build_all_supervisor_plans_excel_workbook(plans)
    return _excel_response(wb, "جميع_خطط_المشرفين_لكل_الأسابيع.xlsx")


@admin_only_view
def admin_export_week_visit_summary_excel(request: HttpRequest) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")
    weeks_qs = _get_all_weeks_qs() if show_all else _get_active_weeks_qs()
    default_week = weeks_qs.first()

    if not default_week:
        messages.warning(request, "لا يوجد أسابيع متاحة للتصدير.")
        return redirect("visits:admin_dashboard")

    week_no = _safe_int(request.GET.get("week") or default_week.week_no, default=default_week.week_no)
    if not _week_exists(week_no):
        week_no = default_week.week_no

    week_obj = _resolve_week_or_404(week_no, allow_inactive=show_all)
    wb = _build_week_visit_summary_excel_workbook(week_obj)
    filename = f"ملخص_الزيارات_الأسبوع_{week_obj.week_no}.xlsx"
    return _excel_response(wb, filename)



# =============================================================================
# Admin reports center and control follow-ups
# =============================================================================
def _reports_default_week() -> PlanWeek | None:
    return _current_week_obj() or _get_active_weeks_qs().first() or _get_all_weeks_qs().first()


def _resolve_report_week(request: HttpRequest, *, allow_inactive: bool = False) -> PlanWeek | None:
    default_week = _reports_default_week()
    if not default_week:
        return None
    week_no = _safe_int(request.GET.get("week") or request.POST.get("week") or default_week.week_no, default=default_week.week_no)
    week = PlanWeek.objects.filter(week_no=week_no).first()
    if not week:
        return default_week
    if not allow_inactive and getattr(week, "is_break", False):
        return default_week
    return week


def _control_report_config(report_type: str) -> dict[str, str]:
    configs = {
        "incomplete_plans": {
            "issue_type": "incomplete_plan",
            "title": "الخطط غير المكتملة",
            "subtitle": "خطط محفوظة أو منشأة لكنها لم تستكمل الأيام الخمسة.",
            "empty": "لا توجد خطط غير مكتملة حسب الفلاتر الحالية.",
        },
        "not_saved_plans": {
            "issue_type": "not_saved_plan",
            "title": "المشرفون الذين لم يحفظوا خطة الأسبوع",
            "subtitle": "مشرفون نشطون لديهم إسناد ولم يظهر لهم حفظ فعلي لخطة الأسبوع.",
            "empty": "لا يوجد مشرفون بدون حفظ للخطة في الأسبوع المحدد.",
        },
        "unlock_requests": {
            "issue_type": "unlock_request",
            "title": "طلبات فك الاعتماد",
            "subtitle": "الخطط التي لديها طلب فك اعتماد قائم وتحتاج قرارًا إداريًا.",
            "empty": "لا توجد طلبات فك اعتماد حسب الفلاتر الحالية.",
        },
        "uncovered_schools": {
            "issue_type": "uncovered_schools",
            "title": "المدارس غير المغطاة",
            "subtitle": "مدارس مسندة للمشرف ولم تظهر في أي خطة حتى الأسبوع المحدد.",
            "empty": "لا توجد مدارس غير مغطاة حسب الفلاتر الحالية.",
        },
    }
    return configs.get(report_type, configs["incomplete_plans"])


def _control_report_filter_q(rows: list[dict[str, Any]], q: str) -> list[dict[str, Any]]:
    q = (q or "").strip().lower()
    if not q:
        return rows

    filtered = []
    for row in rows:
        supervisor = row.get("supervisor")
        haystack = " ".join(
            [
                getattr(supervisor, "full_name", "") or "",
                _sup_nid_value(supervisor) if supervisor else "",
                row.get("title", "") or "",
                row.get("description", "") or "",
                " ".join(row.get("school_names", []) or []),
            ]
        ).lower()
        if q in haystack:
            filtered.append(row)
    return filtered


def _build_control_report_rows(report_type: str, *, week_obj: PlanWeek, q: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    active_supervisors = Supervisor.objects.filter(is_active=True).order_by("full_name")

    if report_type == "incomplete_plans":
        plans = (
            Plan.objects.filter(week=week_obj)
            .select_related("supervisor", "week")
            .prefetch_related("days__school")
            .order_by("supervisor__full_name")
        )
        for plan in plans:
            filled = _plan_filled_count(plan)
            if filled >= 5:
                continue
            rows.append(
                {
                    "supervisor": plan.supervisor,
                    "plan": plan,
                    "week": week_obj,
                    "title": f"خطة غير مكتملة للأسبوع {week_obj.week_no}",
                    "description": f"اكتمل من الخطة {filled}/5 أيام فقط.",
                    "detail": f"{filled}/5",
                    "school_names": [],
                }
            )

    elif report_type == "not_saved_plans":
        supervisors = active_supervisors.filter(assignments__is_active=True, assignments__school__is_active=True).distinct()
        plan_map = {
            p.supervisor_id: p
            for p in Plan.objects.filter(week=week_obj).select_related("supervisor", "week").prefetch_related("days__school")
        }
        for supervisor in supervisors:
            plan = plan_map.get(supervisor.id)
            saved_at = getattr(plan, "saved_at", None) if plan else None
            if saved_at:
                continue
            rows.append(
                {
                    "supervisor": supervisor,
                    "plan": plan,
                    "week": week_obj,
                    "title": f"لم يحفظ خطة الأسبوع {week_obj.week_no}",
                    "description": "لا يوجد حفظ فعلي للخطة حتى الآن، أو لم يتم إنشاء الخطة لهذا الأسبوع.",
                    "detail": "لم تحفظ",
                    "school_names": [],
                }
            )

    elif report_type == "unlock_requests":
        plans = (
            Plan.objects.filter(week=week_obj, status=Plan.STATUS_UNLOCK_REQUESTED)
            .select_related("supervisor", "week")
            .prefetch_related("days__school")
            .order_by("supervisor__full_name")
        )
        for plan in plans:
            rows.append(
                {
                    "supervisor": plan.supervisor,
                    "plan": plan,
                    "week": week_obj,
                    "title": f"طلب فك اعتماد للأسبوع {week_obj.week_no}",
                    "description": getattr(plan, "admin_note", "") or "يوجد طلب فك اعتماد يحتاج مراجعة الإدارة.",
                    "detail": "طلب قائم",
                    "school_names": [],
                }
            )

    elif report_type == "uncovered_schools":
        supervisors = active_supervisors.filter(assignments__is_active=True, assignments__school__is_active=True).distinct()
        for supervisor in supervisors:
            assigned_schools = list(_supervisor_schools_qs(supervisor))
            assigned_ids = {s.id for s in assigned_schools}
            if not assigned_ids:
                continue
            covered_ids = set(
                PlanDay.objects.filter(
                    plan__supervisor=supervisor,
                    plan__week__is_break=False,
                    plan__week__week_no__lte=week_obj.week_no,
                    school_id__isnull=False,
                ).values_list("school_id", flat=True)
            ) & assigned_ids
            uncovered = [school for school in assigned_schools if school.id not in covered_ids]
            if not uncovered:
                continue
            school_names = [school.name for school in uncovered]
            rows.append(
                {
                    "supervisor": supervisor,
                    "plan": None,
                    "week": week_obj,
                    "title": f"مدارس غير مغطاة حتى الأسبوع {week_obj.week_no}",
                    "description": f"عدد المدارس غير المغطاة: {len(uncovered)} — " + "، ".join(school_names[:8]) + (" ..." if len(school_names) > 8 else ""),
                    "detail": str(len(uncovered)),
                    "school_names": school_names,
                }
            )

    return _control_report_filter_q(rows, q)


def _control_followup_unique_key(*, issue_type: str, supervisor_id: int, week_id: int | None, plan_id: int | None) -> str:
    return f"{issue_type}:sup={supervisor_id}:week={week_id or 0}:plan={plan_id or 0}"


def _create_or_update_control_followup_from_row(
    *,
    report_type: str,
    row: dict[str, Any],
    notified: bool = False,
    admin_note: str = "",
):
    ControlFollowUp = _get_control_followup_model()
    if not ControlFollowUp:
        return None

    config = _control_report_config(report_type)
    issue_type = config["issue_type"]
    supervisor = row["supervisor"]
    plan = row.get("plan")
    week = row.get("week")
    unique_key = _control_followup_unique_key(
        issue_type=issue_type,
        supervisor_id=supervisor.id,
        week_id=getattr(week, "id", None),
        plan_id=getattr(plan, "id", None),
    )

    defaults = {
        "issue_type": issue_type,
        "supervisor": supervisor,
        "plan": plan,
        "week": week,
        "title": row.get("title") or config["title"],
        "description": row.get("description") or config["subtitle"],
        "admin_note": admin_note or "",
    }

    obj, created = ControlFollowUp.objects.get_or_create(unique_key=unique_key, defaults=defaults)

    changed_fields: list[str] = []
    for field in ("title", "description", "admin_note"):
        value = defaults[field]
        if value and getattr(obj, field, "") != value:
            setattr(obj, field, value)
            changed_fields.append(field)

    if getattr(obj, "status", "open") == "closed":
        obj.status = "open"
        obj.closed_at = None
        changed_fields.extend(["status", "closed_at"])

    if notified:
        obj.status = "notified"
        obj.notification_count = (getattr(obj, "notification_count", 0) or 0) + 1
        obj.last_notification_at = timezone.now()
        changed_fields.extend(["status", "notification_count", "last_notification_at"])

    if changed_fields:
        changed_fields.append("updated_at")
        obj.save(update_fields=list(dict.fromkeys(changed_fields)))

    return obj


def _build_control_report_excel_workbook(*, report_type: str, week_obj: PlanWeek, rows: list[dict[str, Any]]) -> Workbook:
    config = _control_report_config(report_type)
    wb = Workbook()
    ws = wb.active
    ws.title = config["title"][:31]
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{config['title']} — الأسبوع {week_obj.week_no}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    headers = ["م", "اسم المشرف", "السجل المدني", "الأسبوع", "نوع الحالة", "الوصف", "التفصيل", "المدارس/الملاحظات"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = 4
    for i, row in enumerate(rows, start=1):
        supervisor = row["supervisor"]
        school_names = row.get("school_names") or []
        values = [
            i,
            supervisor.full_name,
            _sup_nid_value(supervisor),
            row.get("week").week_no if row.get("week") else "—",
            config["title"],
            row.get("description") or "—",
            row.get("detail") or "—",
            "، ".join(school_names) if school_names else "—",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 4, 5, 7) else right
        row_idx += 1

    for col_i, width in {1: 7, 2: 28, 3: 18, 4: 10, 5: 24, 6: 48, 7: 16, 8: 60}.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width
    return wb


@admin_only_view
def admin_reports_view(request: HttpRequest) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes", "on")
    week_obj = _resolve_report_week(request, allow_inactive=show_all)
    week_no = week_obj.week_no if week_obj else _get_default_week_no()

    ControlFollowUp = _get_control_followup_model()
    control_followup_open_count = 0
    control_followup_pending_count = 0
    if ControlFollowUp:
        control_followup_open_count = ControlFollowUp.objects.filter(status__in=["open", "notified"]).count()
        control_followup_pending_count = ControlFollowUp.objects.filter(status="pending_admin").count()

    return render(
        request,
        "visits/admin_reports.html",
        {
            "week": week_no,
            "week_obj": week_obj,
            "week_choices": _build_week_choices(active_only=not show_all),
            "show_all": show_all,
            "control_followup_open_count": control_followup_open_count,
            "control_followup_pending_count": control_followup_pending_count,
            "control_followup_model_ready": bool(ControlFollowUp),
        },
    )


@admin_only_view
def admin_control_report_view(request: HttpRequest, report_type: str) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes", "on")
    week_obj = _resolve_report_week(request, allow_inactive=show_all)
    if not week_obj:
        messages.warning(request, "لا يوجد أسبوع متاح لعرض التقرير.")
        return redirect("visits:admin_reports")

    q = _cell_str(request.GET.get("q"))
    config = _control_report_config(report_type)
    rows = _build_control_report_rows(report_type, week_obj=week_obj, q=q)

    paginator = Paginator(rows, _safe_int(request.GET.get("ps") or 25, default=25))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/admin_control_report.html",
        {
            "report_type": report_type,
            "report_title": config["title"],
            "report_subtitle": config["subtitle"],
            "empty_message": config["empty"],
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "week_choices": _build_week_choices(active_only=not show_all),
            "show_all": show_all,
            "q": q,
            "total_count": len(rows),
            "control_followup_model_ready": bool(_get_control_followup_model()),
        },
    )


@admin_only_view
def admin_control_report_export_excel_view(request: HttpRequest, report_type: str) -> HttpResponse:
    show_all = (request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes", "on")
    week_obj = _resolve_report_week(request, allow_inactive=show_all)
    if not week_obj:
        messages.warning(request, "لا يوجد أسبوع متاح للتصدير.")
        return redirect("visits:admin_reports")

    q = _cell_str(request.GET.get("q"))
    rows = _build_control_report_rows(report_type, week_obj=week_obj, q=q)
    wb = _build_control_report_excel_workbook(report_type=report_type, week_obj=week_obj, rows=rows)
    filename = f"تقرير_رقابي_{report_type}_الأسبوع_{week_obj.week_no}.xlsx"
    return _excel_response(wb, filename)


@admin_only_view
@require_POST
def admin_control_report_notify_view(request: HttpRequest, report_type: str) -> HttpResponse:
    ControlFollowUp = _get_control_followup_model()
    if not ControlFollowUp:
        messages.error(request, "لم يتم تفعيل نموذج سجل المتابعة الرقابية بعد. أضف الموديل ثم نفذ makemigrations و migrate.")
        return redirect("visits:admin_reports")

    show_all = (request.POST.get("all") or request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes", "on")
    week_obj = _resolve_report_week(request, allow_inactive=show_all)
    if not week_obj:
        messages.warning(request, "لا يوجد أسبوع متاح للتنبيه.")
        return redirect("visits:admin_reports")

    q = _cell_str(request.POST.get("q") or request.GET.get("q"))
    note = _cell_str(request.POST.get("note"))
    config = _control_report_config(report_type)
    rows = _build_control_report_rows(report_type, week_obj=week_obj, q=q)

    sent_count = 0
    for row in rows:
        supervisor = row["supervisor"]
        plan = row.get("plan")
        message = note or row.get("description") or config["subtitle"]
        _notify_supervisor(
            supervisor=supervisor,
            plan=plan,
            notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
            title=f"تنبيه رقابي: {config['title']}",
            message=message,
        )
        _create_or_update_control_followup_from_row(
            report_type=report_type,
            row=row,
            notified=True,
            admin_note=message,
        )
        sent_count += 1

    messages.success(request, f"تم إرسال التنبيه وتسجيل المتابعة لعدد {sent_count} حالة.")
    params = f"week={week_obj.week_no}"
    if show_all:
        params += "&all=1"
    if q:
        params += f"&q={q}"
    return redirect(f"{reverse('visits:admin_control_report', args=[report_type])}?{params}")


def _control_followups_base_qs():
    ControlFollowUp = _get_control_followup_model()
    if not ControlFollowUp:
        return None
    return (
        ControlFollowUp.objects
        .select_related("supervisor", "plan", "week")
        .order_by("-updated_at", "-created_at", "-id")
    )


@admin_only_view
def admin_control_followups_view(request: HttpRequest) -> HttpResponse:
    qs = _control_followups_base_qs()
    if qs is None:
        messages.error(request, "لم يتم تفعيل نموذج سجل المتابعة الرقابية بعد. أضف الموديل ثم نفذ makemigrations و migrate.")
        return redirect("visits:admin_reports")

    status_filter = _cell_str(request.GET.get("status") or "active")
    issue_type = _cell_str(request.GET.get("issue_type") or "all")
    q = _cell_str(request.GET.get("q"))

    if status_filter == "active":
        qs = qs.filter(status__in=["open", "notified", "pending_admin"])
    elif status_filter != "all":
        qs = qs.filter(status=status_filter)

    if issue_type != "all":
        qs = qs.filter(issue_type=issue_type)

    if q:
        qs = qs.filter(
            Q(supervisor__full_name__icontains=q)
            | Q(supervisor__national_id__icontains=q)
            | Q(title__icontains=q)
            | Q(description__icontains=q)
            | Q(supervisor_response__icontains=q)
        )

    paginator = Paginator(qs, _safe_int(request.GET.get("ps") or 25, default=25))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    stats_qs = _control_followups_base_qs()
    stats = {
        "open_count": stats_qs.filter(status="open").count(),
        "notified_count": stats_qs.filter(status="notified").count(),
        "pending_count": stats_qs.filter(status="pending_admin").count(),
        "processed_count": stats_qs.filter(status="processed").count(),
        "closed_count": stats_qs.filter(status="closed").count(),
    }

    return render(
        request,
        "visits/admin_control_followups.html",
        {
            "page_obj": page_obj,
            "rows": page_obj.object_list,
            "status_filter": status_filter,
            "issue_type": issue_type,
            "q": q,
            "stats": stats,
            "status_choices": [
                ("active", "النشطة"),
                ("all", "الكل"),
                ("open", "مفتوحة"),
                ("notified", "تم التنبيه"),
                ("pending_admin", "بانتظار مراجعة الإدارة"),
                ("processed", "تمت المعالجة"),
                ("closed", "مغلقة إداريًا"),
            ],
            "issue_type_choices": [
                ("all", "كل الأنواع"),
                ("incomplete_plan", "خطة غير مكتملة"),
                ("not_saved_plan", "لم يحفظ خطة الأسبوع"),
                ("unlock_request", "طلب فك اعتماد"),
                ("uncovered_schools", "مدارس غير مغطاة"),
            ],
        },
    )


@admin_only_view
def admin_control_followups_export_excel_view(request: HttpRequest) -> HttpResponse:
    qs = _control_followups_base_qs()
    if qs is None:
        messages.error(request, "لم يتم تفعيل نموذج سجل المتابعة الرقابية بعد.")
        return redirect("visits:admin_reports")

    wb = Workbook()
    ws = wb.active
    ws.title = "سجل المتابعة الرقابية"
    ws.sheet_view.rightToLeft = True
    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = "سجل المتابعة الرقابية"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    headers = ["م", "نوع الحالة", "المشرف", "السجل المدني", "الأسبوع", "العنوان", "الحالة", "عدد التنبيهات", "آخر تنبيه", "إفادة المشرف", "ملاحظة الإدارة"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = 4
    for i, obj in enumerate(qs, start=1):
        values = [
            i,
            _control_followup_type_label(obj.issue_type),
            obj.supervisor.full_name,
            _sup_nid_value(obj.supervisor),
            obj.week.week_no if obj.week else "—",
            obj.title,
            _control_followup_status_label(obj.status),
            obj.notification_count,
            timezone.localtime(obj.last_notification_at).strftime("%Y-%m-%d %H:%M") if obj.last_notification_at else "—",
            obj.supervisor_response or "—",
            obj.admin_review_note or obj.admin_note or "—",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 4, 5, 7, 8, 9) else right
        row_idx += 1

    for col_i, width in {1: 7, 2: 22, 3: 28, 4: 18, 5: 10, 6: 36, 7: 20, 8: 14, 9: 20, 10: 48, 11: 48}.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width
    return _excel_response(wb, "سجل_المتابعة_الرقابية.xlsx")


@admin_only_view
@require_POST
def admin_control_followup_notify_view(request: HttpRequest, pk: int) -> HttpResponse:
    qs = _control_followups_base_qs()
    if qs is None:
        messages.error(request, "لم يتم تفعيل نموذج سجل المتابعة الرقابية بعد.")
        return redirect("visits:admin_reports")
    obj = get_object_or_404(qs, pk=pk)
    note = _cell_str(request.POST.get("note")) or obj.admin_note or obj.description

    _notify_supervisor(
        supervisor=obj.supervisor,
        plan=obj.plan,
        notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
        title=f"تنبيه رقابي: {_control_followup_type_label(obj.issue_type)}",
        message=note,
    )
    obj.status = "notified"
    obj.notification_count = (obj.notification_count or 0) + 1
    obj.last_notification_at = timezone.now()
    obj.admin_note = note
    obj.save(update_fields=["status", "notification_count", "last_notification_at", "admin_note", "updated_at"])
    messages.success(request, "تم إرسال التنبيه وتحديث سجل المتابعة.")
    return redirect("visits:admin_control_followups")


@admin_only_view
@require_POST
def admin_control_followup_update_view(request: HttpRequest, pk: int) -> HttpResponse:
    qs = _control_followups_base_qs()
    if qs is None:
        messages.error(request, "لم يتم تفعيل نموذج سجل المتابعة الرقابية بعد.")
        return redirect("visits:admin_reports")
    obj = get_object_or_404(qs, pk=pk)
    action = _cell_str(request.POST.get("action"))
    note = _cell_str(request.POST.get("admin_review_note"))

    if action == "accept":
        obj.status = "processed"
        obj.resolved_at = timezone.now()
        obj.admin_review_note = note or "تم قبول معالجة المشرف وإغلاقها كمنجزة."
        obj.save(update_fields=["status", "resolved_at", "admin_review_note", "updated_at"])
        _notify_supervisor(
            supervisor=obj.supervisor,
            plan=obj.plan,
            notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
            title="تم قبول معالجة الملاحظة الرقابية",
            message=obj.admin_review_note,
        )
        messages.success(request, "تم قبول المعالجة وتحديث الحالة.")
    elif action == "return":
        obj.status = "notified"
        obj.admin_review_note = note or "تمت إعادة الملاحظة للمشرف لاستكمال المعالجة."
        obj.notification_count = (obj.notification_count or 0) + 1
        obj.last_notification_at = timezone.now()
        obj.save(update_fields=["status", "admin_review_note", "notification_count", "last_notification_at", "updated_at"])
        _notify_supervisor(
            supervisor=obj.supervisor,
            plan=obj.plan,
            notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
            title="إعادة ملاحظة رقابية لاستكمال المعالجة",
            message=obj.admin_review_note,
        )
        messages.success(request, "تمت إعادة الملاحظة للمشرف.")
    elif action == "close":
        obj.status = "closed"
        obj.closed_at = timezone.now()
        obj.admin_review_note = note or "تم إغلاق الحالة إداريًا."
        obj.save(update_fields=["status", "closed_at", "admin_review_note", "updated_at"])
        messages.success(request, "تم إغلاق الحالة إداريًا.")
    else:
        messages.warning(request, "الإجراء غير معروف.")

    return redirect("visits:admin_control_followups")


def _supervisor_control_followups_qs(supervisor: Supervisor):
    ControlFollowUp = _get_control_followup_model()
    if not ControlFollowUp:
        return None
    return (
        ControlFollowUp.objects
        .filter(supervisor=supervisor)
        .select_related("plan", "week")
        .order_by("-updated_at", "-created_at", "-id")
    )


def supervisor_control_followups_view(request: HttpRequest) -> HttpResponse:
    setting = _get_site_setting()
    if _maintenance_is_active(setting, persist=True) and not _maintenance_allowed_for_request(request, setting):
        return redirect("visits:maintenance_page")

    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    qs = _supervisor_control_followups_qs(supervisor)
    if qs is None:
        messages.warning(request, "لم يتم تفعيل الملاحظات الرقابية بعد.")
        return redirect("visits:supervisor_dashboard")

    status_filter = _cell_str(request.GET.get("status") or "active")
    if status_filter == "active":
        qs = qs.filter(status__in=["open", "notified", "pending_admin"])
    elif status_filter != "all":
        qs = qs.filter(status=status_filter)

    paginator = Paginator(qs, _safe_int(request.GET.get("ps") or 10, default=10))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/supervisor_control_followups.html",
        {
            "sup": supervisor,
            "page_obj": page_obj,
            "rows": page_obj.object_list,
            "status_filter": status_filter,
            "active_count": _supervisor_open_control_followups_count(supervisor),
        },
    )


@require_POST
def supervisor_control_followup_respond_view(request: HttpRequest, pk: int) -> HttpResponse:
    try:
        supervisor = _require_supervisor(request)
    except Supervisor.DoesNotExist:
        return redirect("visits:login")

    qs = _supervisor_control_followups_qs(supervisor)
    if qs is None:
        messages.warning(request, "لم يتم تفعيل الملاحظات الرقابية بعد.")
        return redirect("visits:supervisor_dashboard")

    obj = get_object_or_404(qs, pk=pk)
    response = _cell_str(request.POST.get("supervisor_response"))
    if not response:
        messages.warning(request, "يرجى كتابة إفادتكم قبل الإرسال.")
        return redirect("visits:supervisor_control_followups")

    obj.supervisor_response = response
    obj.supervisor_response_at = timezone.now()
    obj.status = "pending_admin"
    obj.save(update_fields=["supervisor_response", "supervisor_response_at", "status", "updated_at"])
    messages.success(request, "تم إرسال الإفادة للإدارة، وبانتظار مراجعتها.")
    return redirect("visits:supervisor_control_followups")


# =============================================================================
# Admin actions
# =============================================================================
@admin_only_view
@require_POST
def admin_plan_approve_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(Plan.objects.select_related("supervisor", "week").prefetch_related("days"), id=plan_id)
    _inject_week_no(plan)

    state = _parse_admin_action_state(request)
    return_url = _resolve_admin_return_url(
        request,
        plan=plan,
        default_week_no=plan.week.week_no,
        show_all=state["show_all"],
        q=state["q"],
        status_filter=state["status_filter"],
        ps=state["ps"],
        page=state["page"],
    )

    if plan.status == Plan.STATUS_APPROVED:
        msg = "الخطة معتمدة مسبقًا."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
        if ajax_response:
            return ajax_response
        messages.info(request, msg)
        return redirect(return_url)

    if not plan.is_fully_filled():
        msg = "لا يمكن اعتماد الخطة قبل اكتمال جميع الأيام (5/5)."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=False, http_status=400, errors=[msg])
        if ajax_response:
            return ajax_response
        messages.warning(request, msg)
        return redirect(return_url)

    plan.status = Plan.STATUS_APPROVED
    plan.approved_at = timezone.now()
    plan.admin_note = None
    plan.save(update_fields=["status", "approved_at", "admin_note"])
    UnlockRequest.objects.filter(plan=plan).delete()

    _notify_supervisor(
        supervisor=plan.supervisor,
        plan=plan,
        notif_type=SupervisorNotification.TYPE_APPROVED,
        title="تم اعتماد خطتك",
        message=f"تم اعتماد خطة الأسبوع {plan.week.week_no} بنجاح.",
    )

    msg = "تم اعتماد الخطة."
    ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
    if ajax_response:
        return ajax_response

    messages.success(request, msg)
    return redirect(return_url)


@admin_only_view
@require_POST
def admin_plan_back_to_draft_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(Plan.objects.select_related("supervisor", "week").prefetch_related("days"), id=plan_id)
    _inject_week_no(plan)

    note = (request.POST.get("note") or "").strip()
    state = _parse_admin_action_state(request)
    return_url = _resolve_admin_return_url(
        request,
        plan=plan,
        default_week_no=plan.week.week_no,
        show_all=state["show_all"],
        q=state["q"],
        status_filter=state["status_filter"],
        ps=state["ps"],
        page=state["page"],
    )

    if plan.status == Plan.STATUS_DRAFT:
        msg = "الخطة بالفعل مسودة."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
        if ajax_response:
            return ajax_response
        messages.info(request, msg)
        return redirect(return_url)

    plan.status = Plan.STATUS_DRAFT
    plan.approved_at = None
    plan.admin_note = note or None
    plan.save(update_fields=["status", "approved_at", "admin_note"])
    UnlockRequest.objects.filter(plan=plan).delete()

    _notify_supervisor(
        supervisor=plan.supervisor,
        plan=plan,
        notif_type=SupervisorNotification.TYPE_RETURNED,
        title="تمت إعادة الخطة للمراجعة",
        message=f"تمت إعادة خطة الأسبوع {plan.week.week_no} للمراجعة." + (f" الملاحظة: {note}" if note else ""),
    )

    msg = "تم إرجاع الخطة إلى مسودة."
    ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
    if ajax_response:
        return ajax_response

    messages.success(request, msg)
    return redirect(return_url)


@admin_only_view
@require_POST
def admin_send_notification_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(Plan.objects.select_related("supervisor", "week").prefetch_related("days"), id=plan_id)
    _inject_week_no(plan)

    note = (request.POST.get("note") or "").strip()
    title = (request.POST.get("title") or "تنبيه إداري").strip()

    state = _parse_admin_action_state(request)
    return_url = _resolve_admin_return_url(
        request,
        plan=plan,
        default_week_no=plan.week.week_no,
        show_all=state["show_all"],
        q=state["q"],
        status_filter=state["status_filter"],
        ps=state["ps"],
        page=state["page"],
    )

    if not note:
        msg = "نص التنبيه مطلوب."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=False, http_status=400, errors=[msg])
        if ajax_response:
            return ajax_response
        messages.warning(request, msg)
        return redirect(return_url)

    _notify_supervisor(
        supervisor=plan.supervisor,
        plan=plan,
        notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
        title=title,
        message=note,
    )

    msg = "تم إرسال التنبيه إلى المشرف."
    ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
    if ajax_response:
        return ajax_response

    messages.success(request, msg)
    return redirect(return_url)


@admin_only_view
@require_POST
def admin_send_notification_all_view(request: HttpRequest) -> HttpResponse:
    show_all = (request.POST.get("all") or request.GET.get("all") or "0").strip().lower() in ("1", "true", "yes")

    weeks_qs = _get_all_weeks_qs() if show_all else _get_active_weeks_qs()
    default_week = weeks_qs.first()

    if not default_week:
        messages.warning(request, "لا يوجد أسابيع متاحة.")
        return redirect("visits:admin_dashboard")

    week_no = _safe_int(request.POST.get("week") or request.GET.get("week") or default_week.week_no, default=default_week.week_no)
    if not _week_exists(week_no):
        week_no = default_week.week_no

    week_obj = _resolve_week_or_404(week_no, allow_inactive=show_all)
    q = (request.POST.get("q") or request.GET.get("q") or "").strip()
    status_filter = (request.POST.get("status") or request.GET.get("status") or "all").strip().lower()
    ps = _safe_int(request.POST.get("ps") or request.GET.get("ps") or 10, default=10)
    page = _safe_int(request.POST.get("page") or request.GET.get("page") or 1, default=1)
    title = (request.POST.get("title") or "تنبيه إداري عام").strip()
    note = (request.POST.get("note") or "").strip()

    return_url = _admin_dashboard_url(
        week_obj.week_no,
        show_all=show_all,
        q=q,
        status=status_filter,
        ps=ps,
        page=page,
    )

    if not note:
        msg = "نص التنبيه العام مطلوب."
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": msg, "errors": [msg]}, status=400)
        messages.warning(request, msg)
        return redirect(return_url)

    plans_qs = Plan.objects.filter(week=week_obj).select_related("supervisor").order_by("supervisor__full_name")
    if q:
        plans_qs = plans_qs.filter(Q(supervisor__full_name__icontains=q) | Q(supervisor__national_id__icontains=q))

    plans = list(plans_qs)

    if status_filter == "approved":
        plans = [p for p in plans if p.status == Plan.STATUS_APPROVED]
    elif status_filter == "draft":
        plans = [p for p in plans if p.status == Plan.STATUS_DRAFT]
    elif status_filter == "unlock":
        plans = [p for p in plans if p.status == Plan.STATUS_UNLOCK_REQUESTED]
    elif status_filter == "not_full":
        plans = [p for p in plans if not p.is_fully_filled()]

    sent_count = 0
    touched_supervisor_ids: set[int] = set()

    for plan in plans:
        sup = plan.supervisor
        if not sup or not getattr(sup, "is_active", False):
            continue
        if sup.id in touched_supervisor_ids:
            continue
        _notify_supervisor(
            supervisor=sup,
            plan=plan,
            notif_type=SupervisorNotification.TYPE_ADMIN_ALERT,
            title=title,
            message=note,
        )
        touched_supervisor_ids.add(sup.id)
        sent_count += 1

    msg = f"تم إرسال التنبيه العام إلى {sent_count} مشرف/ة."
    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": msg, "sent_count": sent_count, "week_no": week_obj.week_no}, status=200)

    messages.success(request, msg)
    return redirect(return_url)


@admin_only_view
@require_POST
def admin_unlock_approve_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(Plan.objects.select_related("supervisor", "week").prefetch_related("days"), id=plan_id)
    _inject_week_no(plan)

    state = _parse_admin_action_state(request)
    return_url = _resolve_admin_return_url(
        request,
        plan=plan,
        default_week_no=plan.week.week_no,
        show_all=state["show_all"],
        q=state["q"],
        status_filter=state["status_filter"],
        ps=state["ps"],
        page=state["page"],
    )

    unlock = get_object_or_404(UnlockRequest, plan=plan)

    if unlock.status != UnlockRequest.STATUS_PENDING:
        msg = "تمت معالجة طلب فك الاعتماد سابقًا."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=False, http_status=400, errors=[msg])
        if ajax_response:
            return ajax_response
        messages.warning(request, msg)
        return redirect(return_url)

    unlock.status = UnlockRequest.STATUS_APPROVED
    unlock.resolved_at = timezone.now()
    unlock.save(update_fields=["status", "resolved_at"])

    plan.status = Plan.STATUS_DRAFT
    plan.approved_at = None
    plan.save(update_fields=["status", "approved_at"])

    _notify_supervisor(
        supervisor=plan.supervisor,
        plan=plan,
        notif_type=SupervisorNotification.TYPE_UNLOCK_APPROVED,
        title="تمت الموافقة على فك اعتماد الخطة",
        message=f"تمت الموافقة على فك اعتماد خطة الأسبوع {plan.week.week_no}. يمكنك التعديل الآن.",
    )

    msg = "تمت الموافقة على فك الاعتماد وإرجاع الخطة إلى مسودة."
    ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
    if ajax_response:
        return ajax_response

    messages.success(request, msg)
    return redirect(return_url)


@admin_only_view
@require_POST
def admin_unlock_reject_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(Plan.objects.select_related("supervisor", "week").prefetch_related("days"), id=plan_id)
    _inject_week_no(plan)

    state = _parse_admin_action_state(request)
    return_url = _resolve_admin_return_url(
        request,
        plan=plan,
        default_week_no=plan.week.week_no,
        show_all=state["show_all"],
        q=state["q"],
        status_filter=state["status_filter"],
        ps=state["ps"],
        page=state["page"],
    )

    unlock = get_object_or_404(UnlockRequest, plan=plan)
    note = (request.POST.get("note") or "").strip()

    if unlock.status != UnlockRequest.STATUS_PENDING:
        msg = "تمت معالجة طلب فك الاعتماد سابقًا."
        ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=False, http_status=400, errors=[msg])
        if ajax_response:
            return ajax_response
        messages.warning(request, msg)
        return redirect(return_url)

    unlock.status = UnlockRequest.STATUS_REJECTED
    unlock.resolved_at = timezone.now()
    unlock.save(update_fields=["status", "resolved_at"])

    plan.status = Plan.STATUS_APPROVED
    plan.save(update_fields=["status"])

    _notify_supervisor(
        supervisor=plan.supervisor,
        plan=plan,
        notif_type=SupervisorNotification.TYPE_UNLOCK_REJECTED,
        title="تم رفض طلب فك اعتماد الخطة",
        message=f"تم رفض طلب فك اعتماد خطة الأسبوع {plan.week.week_no}." + (f" الملاحظة: {note}" if note else ""),
    )

    msg = "تم رفض طلب فك الاعتماد."
    ajax_response = _admin_json_response(request, plan=plan, message=msg, ok=True, http_status=200)
    if ajax_response:
        return ajax_response

    messages.success(request, msg)
    return redirect(return_url)


# =============================================================================
# Weekly letter links admin
# =============================================================================
def _weekly_letter_links_stats() -> dict[str, int]:
    return {
        "total_links": WeeklyLetterLink.objects.count(),
        "active_links": WeeklyLetterLink.objects.filter(is_active=True).count(),
        "inactive_links": WeeklyLetterLink.objects.filter(is_active=False).count(),
        "total_weeks": WeeklyLetterLink.objects.values("week_id").distinct().count(),
    }


def _weekly_letter_links_list_context(*, form=None, edit_obj=None) -> dict[str, Any]:
    rows = (
        WeeklyLetterLink.objects
        .select_related("week")
        .order_by("week__week_no", "-is_active", "-id")
    )

    context = {
        "rows": rows,
        "form": form or WeeklyLetterLinkForm(),
        "edit_obj": edit_obj,
    }
    context.update(_weekly_letter_links_stats())
    return context


@admin_only_view
def weekly_letter_links_list_view(request: HttpRequest) -> HttpResponse:
    return render(
        request,
        "visits/weekly_letter_links_list.html",
        _weekly_letter_links_list_context(),
    )


@admin_only_view
def weekly_letter_link_create_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = WeeklyLetterLinkForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)

            if obj.is_active:
                WeeklyLetterLink.objects.filter(
                    week=obj.week,
                    is_active=True,
                ).update(is_active=False)

            obj.save()
            messages.success(request, "تم حفظ رابط الأسبوع بنجاح.")
            return redirect("visits:weekly_letter_links_list")

        messages.error(request, "تعذر حفظ رابط الأسبوع. تحقق من الحقول.")
        return render(
            request,
            "visits/weekly_letter_link_form.html",
            {
                "form": form,
                "edit_obj": None,
                "page_title": "إضافة رابط خطاب أسبوعي",
                "submit_label": "حفظ الرابط",
            },
        )

    return render(
        request,
        "visits/weekly_letter_link_form.html",
        {
            "form": WeeklyLetterLinkForm(),
            "edit_obj": None,
            "page_title": "إضافة رابط خطاب أسبوعي",
            "submit_label": "حفظ الرابط",
        },
    )


@admin_only_view
def weekly_letter_link_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(WeeklyLetterLink, pk=pk)

    if request.method == "POST":
        form = WeeklyLetterLinkForm(request.POST, instance=obj)
        if form.is_valid():
            updated = form.save(commit=False)

            if updated.is_active:
                WeeklyLetterLink.objects.filter(
                    week=updated.week,
                    is_active=True,
                ).exclude(id=updated.id).update(is_active=False)

            updated.save()
            messages.success(request, "تم تحديث رابط الأسبوع بنجاح.")
            return redirect("visits:weekly_letter_links_list")

        messages.error(request, "تعذر تحديث الرابط. تحقق من الحقول.")
        return render(
            request,
            "visits/weekly_letter_link_form.html",
            {
                "form": form,
                "edit_obj": obj,
                "page_title": "تعديل رابط الخطاب",
                "submit_label": "حفظ التعديلات",
            },
        )

    return render(
        request,
        "visits/weekly_letter_link_form.html",
        {
            "form": WeeklyLetterLinkForm(instance=obj),
            "edit_obj": obj,
            "page_title": "تعديل رابط الخطاب",
            "submit_label": "حفظ التعديلات",
        },
    )


@admin_only_view
@require_POST
def weekly_letter_link_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(WeeklyLetterLink, pk=pk)
    obj.delete()
    messages.success(request, "تم حذف الرابط.")
    return redirect("visits:weekly_letter_links_list")


# =============================================================================
# Weekly assignment letters status report
# =============================================================================
def _weekly_letter_drive_file_key(file_obj: dict) -> str:
    """مفتاح آمن للملف حتى نميّز الملفات المطابقة وغير المطابقة."""
    return str(
        file_obj.get("id")
        or file_obj.get("webViewLink")
        or file_obj.get("webContentLink")
        or file_obj.get("name")
        or id(file_obj)
    )


def _weekly_letter_drive_file_name(file_obj: dict) -> str:
    return (file_obj.get("name") or "").strip()


def _weekly_letter_drive_file_url(file_obj: dict) -> str:
    return (
        file_obj.get("webViewLink")
        or file_obj.get("webContentLink")
        or file_obj.get("alternateLink")
        or ""
    )


def _weekly_letter_drive_file_is_pdf(file_obj: dict) -> bool:
    """
    يحدد هل الملف ملف PDF.

    بعض ملفات Google Drive قد يظهر اسمها مثل: pdf.1021880941
    دون امتداد .pdf، لذلك نعتمد على mimeType متى توفر، ثم اسم الملف.
    """
    name = _weekly_letter_drive_file_name(file_obj).lower()
    mime_type = (file_obj.get("mimeType") or file_obj.get("mime_type") or "").lower()

    if mime_type == "application/pdf":
        return True

    if name.endswith(".pdf"):
        return True

    if name.startswith("pdf."):
        return True

    return False


def _weekly_letter_nids_from_file_name(name: str) -> list[str]:
    """
    يستخرج أرقام هوية محتملة من اسم الملف.
    نعتمد على 10 أرقام متتالية؛ لأنها صيغة السجل المدني المتوقعة.
    """
    name = name or ""
    found = re.findall(r"\d{10}", name)

    # احتياط للأسماء التي تحتوي رموزًا بين الأرقام أو اسمها كله عبارة عن رقم.
    digits = _digits(name)
    if not found and len(digits) == 10:
        found.append(digits)

    # إزالة التكرار مع الحفاظ على الترتيب.
    unique: list[str] = []
    for nid in found:
        if nid not in unique:
            unique.append(nid)
    return unique


def _weekly_letter_file_for_nid(files: list[dict], national_id: str) -> dict | None:
    """
    يبحث عن ملف خطاب التكليف داخل ملفات Google Drive
    بناءً على وجود رقم الهوية الوطنية داخل اسم ملف PDF.
    """
    nid = _digits(national_id)
    if not nid:
        return None

    for file_obj in files:
        if not _weekly_letter_drive_file_is_pdf(file_obj):
            continue

        name = _weekly_letter_drive_file_name(file_obj)
        name_digits = _digits(name)

        if nid in name or nid in name_digits:
            return file_obj

    return None


def _weekly_letter_supervisors_qs():
    """
    مصدر تقرير خطابات التكليف.

    لا نعتمد على وجود خطة أسبوعية فقط؛ لأن المطلوب معرفة من تم تصدير
    خطاب تكليف له من المشرفين المعنيين. لذلك يكون المصدر الأساسي:
    المشرفون النشطون الذين لديهم إسناد نشط على مدرسة نشطة.
    """
    return (
        Supervisor.objects
        .filter(
            is_active=True,
            assignments__is_active=True,
            assignments__school__is_active=True,
        )
        .distinct()
        .order_by("full_name")
    )


def _weekly_letter_status_context(
    *,
    week_obj: PlanWeek,
    filter_status: str = "all",
) -> dict[str, Any]:
    """
    يبني بيانات تقرير حالة خطابات التكليف للأسبوع المحدد.

    المنطق الإداري النهائي:
    - حالة الخطاب تُحسب فقط من وجود ملف PDF مطابق برقم الهوية داخل مجلد الأسبوع.
    - جاهزية الخطة معلومة مساعدة ولا تؤثر على حكم: تم التصدير / لم يتم التصدير.
    - إذا تعذر فحص مجلد Drive تكون الحالة: غير مفحوص.
    - تُعرض ملفات PDF غير المطابقة لأي مشرف معني في قسم مستقل لمراجعة جودة التسمية.
    """
    link_obj = (
        WeeklyLetterLink.objects
        .filter(week=week_obj, is_active=True)
        .select_related("week")
        .first()
    )

    folder_id = None
    drive_files: list[dict] = []
    drive_error = ""
    can_inspect_drive = False

    if link_obj and getattr(link_obj, "drive_url", ""):
        folder_id = _extract_drive_folder_id(link_obj.drive_url)

        if folder_id:
            try:
                drive_files = list_files_in_folder(folder_id, page_size=1000)
                can_inspect_drive = True
            except Exception as exc:
                drive_error = f"تعذر فحص مجلد الخطابات: {exc}"
                drive_files = []
                can_inspect_drive = False
        else:
            drive_error = "رابط مجلد الخطابات غير صالح."
    else:
        drive_error = "لا يوجد رابط مجلد خطابات مفعّل لهذا الأسبوع."

    supervisors = list(_weekly_letter_supervisors_qs())
    supervisor_ids = [sup.id for sup in supervisors]
    supervisor_nid_set = {
        _digits(_sup_nid_value(sup))
        for sup in supervisors
        if _digits(_sup_nid_value(sup))
    }

    plans = (
        Plan.objects
        .filter(week=week_obj, supervisor_id__in=supervisor_ids)
        .select_related("supervisor", "week")
        .prefetch_related("days")
    )
    plan_by_supervisor_id = {plan.supervisor_id: plan for plan in plans}

    drive_pdf_files = [file_obj for file_obj in drive_files if _weekly_letter_drive_file_is_pdf(file_obj)] if can_inspect_drive else []
    matched_pdf_file_keys: set[str] = set()
    unmatched_pdf_files: list[dict[str, Any]] = []

    if can_inspect_drive:
        for file_obj in drive_pdf_files:
            file_name = _weekly_letter_drive_file_name(file_obj)
            file_url = _weekly_letter_drive_file_url(file_obj)
            extracted_nids = _weekly_letter_nids_from_file_name(file_name)
            key = _weekly_letter_drive_file_key(file_obj)

            if extracted_nids and any(nid in supervisor_nid_set for nid in extracted_nids):
                matched_pdf_file_keys.add(key)
                continue

            if extracted_nids:
                reason = "رقم الهوية في اسم الملف لا يطابق مشرفًا معنيًا في هذا التقرير."
                extracted_label = "، ".join(extracted_nids)
            else:
                reason = "لم يتم العثور على رقم هوية من 10 أرقام في اسم الملف."
                extracted_label = "—"

            unmatched_pdf_files.append(
                {
                    "file_name": file_name,
                    "file_url": file_url,
                    "extracted_nids": extracted_nids,
                    "extracted_label": extracted_label,
                    "reason": reason,
                }
            )

    all_rows: list[dict[str, Any]] = []

    for supervisor in supervisors:
        plan = plan_by_supervisor_id.get(supervisor.id)
        national_id = _sup_nid_value(supervisor)
        filled_count = _plan_filled_count(plan) if plan else 0
        is_full = bool(plan and filled_count == len(WEEKDAYS))
        found_file = _weekly_letter_file_for_nid(drive_files, national_id) if can_inspect_drive else None

        if plan:
            plan_status_label = _status_label(plan)
            plan_status_code = _status_code(plan)
        else:
            plan_status_label = "لا توجد خطة"
            plan_status_code = "no-plan"

        readiness_code = "ready" if is_full else "not_ready"
        readiness_label = "خطة مكتملة" if is_full else "خطة غير مكتملة/غير موجودة"

        if not can_inspect_drive:
            letter_status_code = "unchecked"
            letter_status_label = "غير مفحوص"
            letter_status_note = drive_error or "تعذر فحص مجلد الخطابات."
            letter_file_name = ""
            letter_file_url = ""
        elif found_file:
            letter_status_code = "exported"
            letter_status_label = "تم تصدير الخطاب"
            letter_status_note = "تم العثور على ملف PDF برقم الهوية."
            letter_file_name = _weekly_letter_drive_file_name(found_file)
            letter_file_url = _weekly_letter_drive_file_url(found_file)
        else:
            letter_status_code = "missing"
            letter_status_label = "لم يتم تصدير الخطاب"
            letter_status_note = "لم يتم العثور على ملف PDF مطابق برقم الهوية."
            letter_file_name = ""
            letter_file_url = ""

        all_rows.append(
            {
                "plan": plan,
                "supervisor": supervisor,
                "supervisor_name": supervisor.full_name,
                "national_id": national_id,
                "plan_status_label": plan_status_label,
                "plan_status_code": plan_status_code,
                "filled_count": filled_count,
                "is_full": is_full,
                "readiness_code": readiness_code,
                "readiness_label": readiness_label,
                "letter_status_code": letter_status_code,
                "letter_status_label": letter_status_label,
                "letter_status_note": letter_status_note,
                "letter_file_name": letter_file_name,
                "letter_file_url": letter_file_url,
            }
        )

    total_count = len(all_rows)
    ready_count = sum(1 for row in all_rows if row["readiness_code"] == "ready")
    not_ready_count = sum(1 for row in all_rows if row["readiness_code"] == "not_ready")
    exported_count = sum(1 for row in all_rows if row["letter_status_code"] == "exported")
    unchecked_count = sum(1 for row in all_rows if row["letter_status_code"] == "unchecked")
    inspected_count = total_count - unchecked_count
    missing_count = sum(1 for row in all_rows if row["letter_status_code"] == "missing")

    exported_percent = round((exported_count / inspected_count) * 100) if inspected_count else 0
    exported_percent_available = bool(inspected_count)

    drive_files_count = len(drive_files) if can_inspect_drive else 0
    drive_pdf_count = len(drive_pdf_files) if can_inspect_drive else 0
    drive_matched_pdf_count = len(matched_pdf_file_keys) if can_inspect_drive else 0
    drive_unmatched_pdf_count = len(unmatched_pdf_files) if can_inspect_drive else 0
    drive_duplicate_or_extra_pdf_count = max(drive_matched_pdf_count - exported_count, 0) if can_inspect_drive else 0

    rows = all_rows
    if filter_status == "exported":
        rows = [row for row in all_rows if row["letter_status_code"] == "exported"]
    elif filter_status == "missing":
        rows = [row for row in all_rows if row["letter_status_code"] == "missing"]
    elif filter_status == "unchecked":
        rows = [row for row in all_rows if row["letter_status_code"] == "unchecked"]
    elif filter_status == "not_ready":
        rows = [row for row in all_rows if row["readiness_code"] == "not_ready"]
    elif filter_status == "ready":
        rows = [row for row in all_rows if row["readiness_code"] == "ready"]

    return {
        "letter_link": link_obj,
        "letter_folder_id": folder_id or "",
        "letter_drive_error": drive_error,
        "letter_can_inspect_drive": can_inspect_drive,
        "letter_rows": rows,
        "letter_all_rows_count": total_count,
        "letter_ready_count": ready_count,
        "letter_not_ready_count": not_ready_count,
        "letter_exported_count": exported_count,
        "letter_missing_count": missing_count,
        "letter_unchecked_count": unchecked_count,
        "letter_inspected_count": inspected_count,
        "letter_exported_percent": exported_percent,
        "letter_exported_percent_available": exported_percent_available,
        "letter_filter_status": filter_status,
        "letter_source_label": "المشرفون النشطون الذين لديهم إسناد نشط",
        "drive_files_count": drive_files_count,
        "drive_pdf_count": drive_pdf_count,
        "drive_matched_pdf_count": drive_matched_pdf_count,
        "drive_unmatched_pdf_count": drive_unmatched_pdf_count,
        "drive_duplicate_or_extra_pdf_count": drive_duplicate_or_extra_pdf_count,
        "drive_unmatched_files": unmatched_pdf_files,
    }


def _build_weekly_letter_status_workbook(
    *,
    week_obj: PlanWeek,
    rows: list[dict[str, Any]],
    filter_status: str = "all",
    unmatched_files: list[dict[str, Any]] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"الأسبوع {week_obj.week_no}"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    bold_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    title_fill = PatternFill("solid", fgColor="E8F5E9")
    exported_fill = PatternFill("solid", fgColor="ECFDF3")
    missing_fill = PatternFill("solid", fgColor="FFF7ED")
    unchecked_fill = PatternFill("solid", fgColor="F8FAFC")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"تقرير حالة خطابات التكليف — الأسبوع {week_obj.week_no}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    filter_label = {
        "all": "الكل",
        "ready": "لديهم خطة مكتملة",
        "exported": "تم تصدير الخطاب",
        "missing": "لم يتم تصدير الخطاب",
        "not_ready": "خطة غير مكتملة/غير موجودة",
        "unchecked": "غير مفحوص",
    }.get(filter_status, "الكل")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"الفلتر: {filter_label}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center

    headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "حالة الخطة",
        "اكتمال الخطة",
        "مؤشر الخطة",
        "حالة الخطاب",
        "سبب الحالة",
        "اسم ملف الخطاب",
        "رابط الخطاب",
    ]

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1

    for i, row in enumerate(rows, start=1):
        values = [
            i,
            row["supervisor_name"],
            row["national_id"],
            row["plan_status_label"],
            "مكتملة" if row["is_full"] else f"{row['filled_count']}/{len(WEEKDAYS)}",
            row["readiness_label"],
            row["letter_status_label"],
            row["letter_status_note"] or "—",
            row["letter_file_name"] or "—",
            row["letter_file_url"] or "—",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 4, 5, 6, 7) else right

            if col == 7:
                if row["letter_status_code"] == "exported":
                    cell.fill = exported_fill
                elif row["letter_status_code"] == "missing":
                    cell.fill = missing_fill
                elif row["letter_status_code"] == "unchecked":
                    cell.fill = unchecked_fill

        row_idx += 1

    widths = {
        1: 7,
        2: 32,
        3: 18,
        4: 16,
        5: 14,
        6: 22,
        7: 20,
        8: 40,
        9: 38,
        10: 55,
    }

    for col_i, width in widths.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    unmatched_files = unmatched_files or []
    if unmatched_files:
        ws_unmatched = wb.create_sheet("ملفات غير مطابقة")
        ws_unmatched.sheet_view.rightToLeft = True

        ws_unmatched.merge_cells("A1:E1")
        ws_unmatched["A1"] = f"ملفات PDF غير مطابقة لأي مشرف معني — الأسبوع {week_obj.week_no}"
        ws_unmatched["A1"].font = title_font
        ws_unmatched["A1"].alignment = center
        ws_unmatched["A1"].fill = title_fill

        unmatched_headers = ["م", "اسم الملف", "الأرقام المستخرجة", "سبب عدم المطابقة", "رابط الملف"]
        for col, header in enumerate(unmatched_headers, start=1):
            cell = ws_unmatched.cell(row=3, column=col, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row_idx = 4
        for i, file_row in enumerate(unmatched_files, start=1):
            values = [
                i,
                file_row.get("file_name") or "—",
                file_row.get("extracted_label") or "—",
                file_row.get("reason") or "—",
                file_row.get("file_url") or "—",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws_unmatched.cell(row=row_idx, column=col, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center if col in (1, 3) else right
            row_idx += 1

        for col_i, width in {1: 7, 2: 42, 3: 24, 4: 44, 5: 58}.items():
            ws_unmatched.column_dimensions[get_column_letter(col_i)].width = width

    return wb


@admin_only_view
def admin_weekly_letter_status_view(request: HttpRequest) -> HttpResponse:
    week_no = _safe_int(
        request.GET.get("week") or _get_default_week_no(),
        default=_get_default_week_no(),
    )

    week_obj = _resolve_week_or_404(week_no, allow_inactive=True)
    filter_status = _cell_str(request.GET.get("letter_status") or "all")

    if filter_status not in ("all", "ready", "exported", "missing", "not_ready", "unchecked"):
        filter_status = "all"

    context = _weekly_letter_status_context(
        week_obj=week_obj,
        filter_status=filter_status,
    )

    context.update(
        {
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "week_choices": _build_week_choices(active_only=False),
            "filter_status": filter_status,
        }
    )

    return render(
        request,
        "visits/admin_weekly_letter_status.html",
        context,
    )


@admin_only_view
def admin_weekly_letter_status_export_excel_view(request: HttpRequest) -> HttpResponse:
    week_no = _safe_int(
        request.GET.get("week") or _get_default_week_no(),
        default=_get_default_week_no(),
    )

    week_obj = _resolve_week_or_404(week_no, allow_inactive=True)
    filter_status = _cell_str(request.GET.get("letter_status") or "all")

    if filter_status not in ("all", "ready", "exported", "missing", "not_ready", "unchecked"):
        filter_status = "all"

    context = _weekly_letter_status_context(
        week_obj=week_obj,
        filter_status=filter_status,
    )

    wb = _build_weekly_letter_status_workbook(
        week_obj=week_obj,
        rows=context["letter_rows"],
        filter_status=filter_status,
        unmatched_files=context.get("drive_unmatched_files") or [],
    )

    filename = f"weekly-letter-status-week-{week_obj.week_no}.xlsx"
    return _excel_response(wb, filename)



# =============================================================================
# Read-only management portal
# =============================================================================
DEPARTMENT_GROUPS = {
    "مدير القسم",
    "مدير قسم",
    "department_manager",
    "readonly_department_manager",
}

UNIT_GROUP_PREFIXES = (
    "مدير وحدة:",
    "مدير وحدة -",
    "مدير وحدة ",
    "unit_manager:",
    "readonly_unit:",
)


@dataclass(frozen=True)
class ReadOnlyScope:
    allowed: bool
    role: str
    role_label: str
    all_scope: bool
    sector_ids: tuple[int, ...]


def _user_group_names(user) -> set[str]:
    if not getattr(user, "is_authenticated", False):
        return set()
    return set(user.groups.values_list("name", flat=True))


def _parse_sector_from_group_name(group_name: str, sectors_by_name: dict[str, int]) -> int | None:
    raw = (group_name or "").strip()
    if not raw:
        return None

    for prefix in UNIT_GROUP_PREFIXES:
        if raw.startswith(prefix):
            value = raw[len(prefix):].strip()
            if not value:
                return None
            if value.isdigit():
                return int(value)
            return sectors_by_name.get(value)

    return None


def get_readonly_scope_for_user(user) -> ReadOnlyScope:
    """
    صلاحيات الاطلاع دون أي صلاحيات تنفيذية.

    طرق الإسناد عبر مجموعات Django:
    - مدير القسم / department_manager: اطلاع على جميع الخطط والإسنادات.
    - مدير وحدة:<sector_id> أو unit_manager:<sector_id>: اطلاع على قطاعات محددة.
    - يمكن أيضًا استخدام اسم القطاع بدل رقمه: مدير وحدة - أبها.
    """
    if not getattr(user, "is_authenticated", False):
        return ReadOnlyScope(False, "none", "غير مصرح", False, ())

    # مدير النظام الحالي يستطيع دخول بوابة الاطلاع أيضًا.
    if getattr(user, "is_superuser", False):
        return ReadOnlyScope(True, "department", "مدير النظام", True, ())

    groups = _user_group_names(user)

    if groups & DEPARTMENT_GROUPS:
        return ReadOnlyScope(True, "department", "مدير القسم", True, ())

    sectors_by_name = {
        (s.name or "").strip(): s.id
        for s in Sector.objects.filter(is_active=True)
        if (s.name or "").strip()
    }

    sector_ids: set[int] = set()
    for group_name in groups:
        sid = _parse_sector_from_group_name(group_name, sectors_by_name)
        if sid:
            sector_ids.add(sid)

    if sector_ids:
        return ReadOnlyScope(True, "unit", "مدير وحدة", False, tuple(sorted(sector_ids)))

    return ReadOnlyScope(False, "none", "غير مصرح", False, ())


def _readonly_access_required(view_func):
    @wraps(view_func)
    def _wrapped(request: HttpRequest, *args, **kwargs):
        if not request.user.is_authenticated:
            login_url = f"{reverse('visits:readonly_login')}?next={request.get_full_path()}"
            return redirect(login_url)

        scope = get_readonly_scope_for_user(request.user)
        if not scope.allowed:
            messages.error(request, "لا تملك صلاحية الدخول إلى بوابة الاطلاع.")
            return redirect("visits:readonly_login")

        # منع تداخل جلسة المشرف مع بوابة الاطلاع.
        # بوابة الاطلاع تعتمد على مستخدم Django، بينما بوابة المشرف تعتمد على visits_sup_id.
        if request.session.get(SESSION_SUP_ID):
            request.session.pop(SESSION_SUP_ID, None)
            request.session.modified = True

        request.readonly_scope = scope  # type: ignore[attr-defined]
        return view_func(request, *args, **kwargs)

    return _wrapped


def _allowed_supervisor_ids(scope: ReadOnlyScope) -> set[int]:
    if scope.all_scope:
        return set(Supervisor.objects.filter(is_active=True).values_list("id", flat=True))

    if not scope.sector_ids:
        return set()

    by_assignment = set(
        Assignment.objects.filter(
            is_active=True,
            school__is_active=True,
            school__sector_id__in=scope.sector_ids,
            supervisor__is_active=True,
        ).values_list("supervisor_id", flat=True).distinct()
    )

    by_supervisor_sector = set(
        Supervisor.objects.filter(
            is_active=True,
            sector_id__in=scope.sector_ids,
        ).values_list("id", flat=True)
    )

    return by_assignment | by_supervisor_sector


def _scope_sectors_qs(scope: ReadOnlyScope):
    qs = Sector.objects.filter(is_active=True).order_by("name")
    if scope.all_scope:
        return qs
    return qs.filter(id__in=scope.sector_ids)


def _readonly_scope_sector_names(scope: ReadOnlyScope) -> list[str]:
    if scope.all_scope:
        return []
    return list(
        Sector.objects.filter(id__in=scope.sector_ids, is_active=True)
        .order_by("name")
        .values_list("name", flat=True)
    )


def _plan_sector_names_for_scope(plan: Plan, scope: ReadOnlyScope) -> str:
    """
    يعرض القطاعات المرتبطة بالمشرف داخل نطاق الاطلاع.
    لا يستخدم للصلاحية الأمنية، بل كقيمة عرض فقط داخل الجداول.
    """
    qs = (
        Assignment.objects.filter(
            supervisor=plan.supervisor,
            is_active=True,
            school__is_active=True,
            school__sector__isnull=False,
        )
        .select_related("school__sector")
        .values_list("school__sector__name", flat=True)
        .distinct()
        .order_by("school__sector__name")
    )

    if not scope.all_scope:
        qs = qs.filter(school__sector_id__in=scope.sector_ids)

    names = [name for name in qs if name]

    if not names:
        sector = getattr(plan.supervisor, "sector", None)
        sector_name = getattr(sector, "name", "") if sector else ""
        if sector_name:
            names = [sector_name]

    return "، ".join(names) if names else "—"


def _plans_qs_for_scope(scope: ReadOnlyScope, *, week_obj: PlanWeek | None = None):
    qs = (
        Plan.objects.select_related("supervisor", "week")
        .prefetch_related("days", "days__school")
        .order_by("supervisor__full_name", "week__week_no")
    )

    if week_obj is not None:
        qs = qs.filter(week=week_obj)

    if scope.all_scope:
        return qs

    allowed = _allowed_supervisor_ids(scope)
    if not allowed:
        return qs.none()

    return qs.filter(supervisor_id__in=allowed)


def _assignments_qs_for_scope(scope: ReadOnlyScope):
    qs = (
        Assignment.objects.select_related("supervisor", "school", "school__sector")
        .order_by("school__sector__name", "school__name", "supervisor__full_name")
    )

    if scope.all_scope:
        return qs

    if not scope.sector_ids:
        return qs.none()

    return qs.filter(school__sector_id__in=scope.sector_ids)


def _plan_row(plan: Plan) -> dict[str, Any]:
    filled = _plan_filled_count(plan)
    return {
        "plan": plan,
        "filled": filled,
        "status_label": _status_label(plan),
        "status_code": _status_code(plan),
        "is_full": filled == len(WEEKDAYS),
    }


def readonly_login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        scope = get_readonly_scope_for_user(request.user)
        if scope.allowed:
            if request.session.get(SESSION_SUP_ID):
                request.session.pop(SESSION_SUP_ID, None)
                request.session.modified = True
            return redirect("visits:readonly_dashboard")

    next_url = _cell_str(request.GET.get("next") or request.POST.get("next") or "")

    if request.method == "POST":
        username = _cell_str(request.POST.get("username"))
        password = request.POST.get("password") or ""

        if not username or not password:
            messages.error(request, "فضلاً أدخل اسم المستخدم وكلمة المرور.")
            return render(request, "visits/readonly_login.html", {"next": next_url})

        user = authenticate(request, username=username, password=password)
        if not user:
            messages.error(request, "بيانات الدخول غير صحيحة.")
            return render(request, "visits/readonly_login.html", {"next": next_url})

        scope = get_readonly_scope_for_user(user)
        if not scope.allowed:
            messages.error(request, "هذا الحساب لا يملك صلاحية الاطلاع.")
            return render(request, "visits/readonly_login.html", {"next": next_url})

        login(request, user)

        # عند دخول بوابة الاطلاع نتأكد من إزالة أي جلسة مشرف سابقة
        # حتى لا يظهر هيدر أو روابط بوابة المشرف بالخطأ.
        request.session.pop(SESSION_SUP_ID, None)
        request.session.modified = True

        messages.success(request, "تم تسجيل الدخول إلى بوابة الاطلاع بنجاح.")

        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            return redirect(next_url)

        return redirect("visits:readonly_dashboard")

    return render(request, "visits/readonly_login.html", {"next": next_url})


def readonly_logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    messages.info(request, "تم تسجيل الخروج من بوابة الاطلاع.")
    return redirect("visits:readonly_login")


@_readonly_access_required
def readonly_dashboard_view(request: HttpRequest) -> HttpResponse:
    scope: ReadOnlyScope = request.readonly_scope  # type: ignore[attr-defined]
    week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
    week_obj = _resolve_week_or_404(week_no, allow_inactive=True)

    plans_qs = _plans_qs_for_scope(scope, week_obj=week_obj)
    assignment_qs = _assignments_qs_for_scope(scope).filter(is_active=True, school__is_active=True)

    plan_ids = list(plans_qs.values_list("id", flat=True))
    plan_rows = []
    for p in plans_qs[:6]:
        row = _plan_row(p)
        row["sector_names"] = _plan_sector_names_for_scope(p, scope)
        plan_rows.append(row)

    total_plans = len(plan_ids)
    approved_count = plans_qs.filter(status=Plan.STATUS_APPROVED).count()
    unlock_count = plans_qs.filter(status=Plan.STATUS_UNLOCK_REQUESTED).count()
    draft_count = max(total_plans - approved_count - unlock_count, 0)
    full_count = sum(1 for p in _plans_qs_for_scope(scope, week_obj=week_obj) if _plan_filled_count(p) == len(WEEKDAYS))

    context = {
        "scope": scope,
        "week": week_obj.week_no,
        "week_obj": week_obj,
        "week_choices": _build_week_choices(active_only=False),
        "kpi_total_plans": total_plans,
        "kpi_approved": approved_count,
        "kpi_draft": draft_count,
        "kpi_unlock": unlock_count,
        "kpi_full": full_count,
        "kpi_assignments": assignment_qs.count(),
        "kpi_schools": assignment_qs.values("school_id").distinct().count(),
        "kpi_supervisors": len(_allowed_supervisor_ids(scope)),
        "latest_plans": plan_rows,
        "sectors": _scope_sectors_qs(scope),
        "scope_sector_names": _readonly_scope_sector_names(scope),
    }
    return render(request, "visits/readonly_dashboard.html", context)


@_readonly_access_required
def readonly_plans_view(request: HttpRequest) -> HttpResponse:
    scope: ReadOnlyScope = request.readonly_scope  # type: ignore[attr-defined]
    week_no = _safe_int(request.GET.get("week") or _get_default_week_no(), default=_get_default_week_no())
    week_obj = _resolve_week_or_404(week_no, allow_inactive=True)

    q = _cell_str(request.GET.get("q"))
    status = _cell_str(request.GET.get("status") or "all")
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)

    qs = _plans_qs_for_scope(scope, week_obj=week_obj)

    if q:
        qs = qs.filter(
            Q(supervisor__full_name__icontains=q)
            | Q(supervisor__national_id__icontains=q)
            | Q(days__school__name__icontains=q)
        ).distinct()

    if status in ("approved", "draft", "unlock"):
        mapping = {
            "approved": Plan.STATUS_APPROVED,
            "draft": Plan.STATUS_DRAFT,
            "unlock": Plan.STATUS_UNLOCK_REQUESTED,
        }
        qs = qs.filter(status=mapping[status])
    elif status == "not_full":
        # تتم فلترة غير المكتملة بعد بناء الصفوف لأن اكتمال الخطة يعتمد على الأيام.
        pass

    if sector_id:
        if not scope.all_scope and sector_id not in scope.sector_ids:
            qs = qs.none()
        else:
            qs = qs.filter(
                Q(supervisor__sector_id=sector_id)
                | Q(supervisor__assignments__school__sector_id=sector_id, supervisor__assignments__is_active=True)
            ).distinct()

    rows = []
    for p in qs:
        row = _plan_row(p)
        row["sector_names"] = _plan_sector_names_for_scope(p, scope)
        rows.append(row)
    if status == "not_full":
        rows = [r for r in rows if not r["is_full"]]

    paginator = Paginator(rows, _safe_int(request.GET.get("ps") or 25, default=25))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/readonly_plans.html",
        {
            "scope": scope,
            "week": week_obj.week_no,
            "week_obj": week_obj,
            "week_choices": _build_week_choices(active_only=False),
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "status": status,
            "sector_id": sector_id,
            "sectors": _scope_sectors_qs(scope),
            "scope_sector_names": _readonly_scope_sector_names(scope),
        },
    )


@_readonly_access_required
def readonly_plan_detail_view(request: HttpRequest, plan_id: int) -> HttpResponse:
    scope: ReadOnlyScope = request.readonly_scope  # type: ignore[attr-defined]
    qs = _plans_qs_for_scope(scope)
    plan = get_object_or_404(qs, id=plan_id)
    row = _plan_row(plan)

    day_map = {d.weekday: d for d in plan.days.all()}
    days = []
    for wd, wd_name in WEEKDAYS:
        days.append({"weekday": wd, "weekday_name": wd_name, "day": day_map.get(wd)})

    assignments = Assignment.objects.filter(
        supervisor=plan.supervisor,
        is_active=True,
        school__is_active=True,
    ).select_related("school", "school__sector").order_by("school__name")

    if not scope.all_scope:
        assignments = assignments.filter(school__sector_id__in=scope.sector_ids)

    return render(
        request,
        "visits/readonly_plan_detail.html",
        {
            "scope": scope,
            "plan": plan,
            "row": row,
            "days": days,
            "assignments": assignments,
            "scope_sector_names": _readonly_scope_sector_names(scope),
        },
    )


@_readonly_access_required
def readonly_assignments_view(request: HttpRequest) -> HttpResponse:
    scope: ReadOnlyScope = request.readonly_scope  # type: ignore[attr-defined]
    q = _cell_str(request.GET.get("q"))
    sector_id = _safe_int(request.GET.get("sector") or 0, default=0)
    gender = _cell_str(request.GET.get("gender") or "")
    active = _cell_str(request.GET.get("active") or "1")

    qs = _assignments_qs_for_scope(scope)

    if active == "1":
        qs = qs.filter(is_active=True, school__is_active=True)
    elif active == "0":
        qs = qs.filter(Q(is_active=False) | Q(school__is_active=False))

    if q:
        qs = qs.filter(
            Q(supervisor__full_name__icontains=q)
            | Q(supervisor__national_id__icontains=q)
            | Q(school__name__icontains=q)
            | Q(school__stat_code__icontains=q)
        )

    if sector_id:
        if not scope.all_scope and sector_id not in scope.sector_ids:
            qs = qs.none()
        else:
            qs = qs.filter(school__sector_id=sector_id)

    if gender in ("boys", "girls"):
        qs = qs.filter(school__gender=gender)

    paginator = Paginator(qs, _safe_int(request.GET.get("ps") or 30, default=30))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/readonly_assignments.html",
        {
            "scope": scope,
            "rows": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "sector_id": sector_id,
            "gender": gender,
            "active": active,
            "sectors": _scope_sectors_qs(scope),
            "scope_sector_names": _readonly_scope_sector_names(scope),
        },
    )

# =============================================================================
# Admin - read-only viewer users management
# =============================================================================
def _readonly_access_groups():
    from django.contrib.auth.models import Group

    return Group.objects.filter(
        Q(name__in=DEPARTMENT_GROUPS)
        | Q(name__startswith="مدير وحدة:")
        | Q(name__startswith="مدير وحدة -")
        | Q(name__startswith="unit_manager:")
        | Q(name__startswith="readonly_unit:")
    ).order_by("name")


def _clear_readonly_access_groups(user) -> None:
    readonly_groups = _readonly_access_groups()
    if readonly_groups.exists():
        user.groups.remove(*list(readonly_groups))


def _apply_readonly_access_groups(user, *, role: str, sector_ids: list[int]) -> None:
    from django.contrib.auth.models import Group

    _clear_readonly_access_groups(user)

    if role == "department":
        group, _ = Group.objects.get_or_create(name="مدير القسم")
        user.groups.add(group)
        return

    if role == "unit":
        for sector_id in sector_ids:
            group, _ = Group.objects.get_or_create(name=f"مدير وحدة:{sector_id}")
            user.groups.add(group)


def _readonly_role_meta_for_user(user) -> dict[str, Any]:
    groups = _user_group_names(user)

    if groups & DEPARTMENT_GROUPS:
        return {
            "role": "department",
            "role_label": "مدير القسم",
            "sector_ids": [],
            "sector_names": ["جميع القطاعات"],
            "scope_label": "جميع الخطط والإسنادات",
        }

    sectors_by_name = {
        (s.name or "").strip(): s.id
        for s in Sector.objects.filter(is_active=True)
        if (s.name or "").strip()
    }
    sectors_by_id = {s.id: s for s in Sector.objects.all()}

    sector_ids: list[int] = []
    for group_name in groups:
        sid = _parse_sector_from_group_name(group_name, sectors_by_name)
        if sid and sid not in sector_ids:
            sector_ids.append(sid)

    if sector_ids:
        names = []
        for sid in sector_ids:
            sector = sectors_by_id.get(sid)
            names.append(getattr(sector, "name", "") or f"قطاع {sid}")
        return {
            "role": "unit",
            "role_label": "مدير وحدة",
            "sector_ids": sector_ids,
            "sector_names": names,
            "scope_label": "، ".join(names),
        }

    return {
        "role": "none",
        "role_label": "بدون صلاحية اطلاع",
        "sector_ids": [],
        "sector_names": [],
        "scope_label": "—",
    }


def _readonly_users_queryset():
    from django.contrib.auth import get_user_model

    User = get_user_model()
    group_ids = list(_readonly_access_groups().values_list("id", flat=True))
    if not group_ids:
        return User.objects.none()

    return (
        User.objects
        .filter(groups__id__in=group_ids)
        .distinct()
        .prefetch_related("groups")
        .order_by("first_name", "last_name", "username")
    )


def _viewer_user_display_name(user) -> str:
    full = f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()
    return full or getattr(user, "username", "")


def _parse_sector_ids_from_request(request: HttpRequest) -> list[int]:
    sector_ids: list[int] = []
    for value in request.POST.getlist("sectors"):
        sid = _safe_int(value, default=0)
        if sid and sid not in sector_ids:
            sector_ids.append(sid)
    return sector_ids


def _validate_viewer_user_form(request: HttpRequest, *, existing_user=None) -> tuple[dict[str, Any], list[str]]:
    username = _cell_str(request.POST.get("username"))
    first_name = _cell_str(request.POST.get("first_name"))
    last_name = _cell_str(request.POST.get("last_name"))
    email = _cell_str(request.POST.get("email"))
    role = _cell_str(request.POST.get("role") or "department")
    is_active = _cell_str(request.POST.get("is_active") or "") in ("1", "on", "true", "yes")
    password = request.POST.get("password") or ""
    password2 = request.POST.get("password2") or ""
    sector_ids = _parse_sector_ids_from_request(request)

    errors: list[str] = []

    if not username:
        errors.append("اسم المستخدم مطلوب.")

    if role not in ("department", "unit"):
        errors.append("نوع الصلاحية غير صحيح.")

    if role == "unit" and not sector_ids:
        errors.append("مدير الوحدة يجب أن يرتبط بقطاع واحد على الأقل.")

    valid_sector_ids = set(Sector.objects.filter(id__in=sector_ids).values_list("id", flat=True))
    for sid in sector_ids:
        if sid not in valid_sector_ids:
            errors.append(f"القطاع رقم {sid} غير موجود.")

    from django.contrib.auth import get_user_model
    User = get_user_model()

    if username:
        qs = User.objects.filter(username=username)
        if existing_user is not None:
            qs = qs.exclude(id=existing_user.id)
        if qs.exists():
            errors.append("اسم المستخدم مستخدم مسبقًا.")

    if existing_user is None:
        if not password:
            errors.append("كلمة المرور مطلوبة عند إنشاء الحساب.")
    if password or password2:
        if password != password2:
            errors.append("كلمتا المرور غير متطابقتين.")
        elif len(password) < 8:
            errors.append("كلمة المرور يجب ألا تقل عن 8 أحرف.")

    data = {
        "username": username,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "role": role,
        "is_active": is_active,
        "password": password,
        "sector_ids": sector_ids,
    }
    return data, errors


@admin_only_view
def admin_viewer_users_view(request: HttpRequest) -> HttpResponse:
    q = _cell_str(request.GET.get("q"))
    role = _cell_str(request.GET.get("role") or "all")
    active = _cell_str(request.GET.get("active") or "all")

    qs = _readonly_users_queryset()

    if q:
        qs = qs.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    if active == "active":
        qs = qs.filter(is_active=True)
    elif active == "inactive":
        qs = qs.filter(is_active=False)

    rows = []
    for user in qs:
        meta = _readonly_role_meta_for_user(user)
        if role in ("department", "unit") and meta["role"] != role:
            continue
        rows.append({"user": user, "meta": meta, "display_name": _viewer_user_display_name(user)})

    paginator = Paginator(rows, _safe_int(request.GET.get("ps") or 20, default=20))
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    context = {
        "rows": page_obj.object_list,
        "page_obj": page_obj,
        "q": q,
        "role": role,
        "active": active,
        "total_count": len(rows),
        "active_count": sum(1 for row in rows if row["user"].is_active),
        "department_count": sum(1 for row in rows if row["meta"]["role"] == "department"),
        "unit_count": sum(1 for row in rows if row["meta"]["role"] == "unit"),
    }
    return render(request, "visits/admin_viewer_users.html", context)


@admin_only_view
def admin_viewer_user_create_view(request: HttpRequest) -> HttpResponse:
    sectors = Sector.objects.filter(is_active=True).order_by("name")

    initial = {
        "username": "",
        "first_name": "",
        "last_name": "",
        "email": "",
        "role": "department",
        "is_active": True,
        "sector_ids": [],
    }

    if request.method == "POST":
        data, errors = _validate_viewer_user_form(request, existing_user=None)
        initial.update(data)

        if not errors:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.create(
                username=data["username"],
                first_name=data["first_name"],
                last_name=data["last_name"],
                email=data["email"],
                is_active=data["is_active"],
                is_staff=False,
                is_superuser=False,
            )
            user.set_password(data["password"])
            user.save()
            _apply_readonly_access_groups(user, role=data["role"], sector_ids=data["sector_ids"])
            messages.success(request, "تم إنشاء حساب الاطلاع بنجاح.")
            return redirect("visits:admin_viewer_users")

        for error in errors:
            messages.error(request, error)

    return render(
        request,
        "visits/admin_viewer_user_form.html",
        {
            "mode": "create",
            "form_data": initial,
            "sectors": sectors,
            "target_user": None,
        },
    )


@admin_only_view
def admin_viewer_user_edit_view(request: HttpRequest, user_id: int) -> HttpResponse:
    from django.contrib.auth import get_user_model

    User = get_user_model()
    target_user = get_object_or_404(User.objects.prefetch_related("groups"), id=user_id)

    if target_user.is_staff or target_user.is_superuser:
        messages.error(request, "لا يمكن تعديل مستخدم إداري من شاشة صلاحيات الاطلاع.")
        return redirect("visits:admin_viewer_users")

    meta = _readonly_role_meta_for_user(target_user)
    if meta["role"] == "none":
        messages.error(request, "هذا الحساب لا يملك صلاحية اطلاع.")
        return redirect("visits:admin_viewer_users")

    sectors = Sector.objects.filter(is_active=True).order_by("name")
    initial = {
        "username": target_user.username,
        "first_name": target_user.first_name,
        "last_name": target_user.last_name,
        "email": target_user.email,
        "role": meta["role"],
        "is_active": target_user.is_active,
        "sector_ids": meta["sector_ids"],
    }

    if request.method == "POST":
        data, errors = _validate_viewer_user_form(request, existing_user=target_user)
        initial.update(data)

        if not errors:
            target_user.username = data["username"]
            target_user.first_name = data["first_name"]
            target_user.last_name = data["last_name"]
            target_user.email = data["email"]
            target_user.is_active = data["is_active"]
            target_user.is_staff = False
            target_user.is_superuser = False
            if data["password"]:
                target_user.set_password(data["password"])
            target_user.save()
            _apply_readonly_access_groups(target_user, role=data["role"], sector_ids=data["sector_ids"])
            messages.success(request, "تم تحديث حساب الاطلاع بنجاح.")
            return redirect("visits:admin_viewer_users")

        for error in errors:
            messages.error(request, error)

    return render(
        request,
        "visits/admin_viewer_user_form.html",
        {
            "mode": "edit",
            "form_data": initial,
            "sectors": sectors,
            "target_user": target_user,
        },
    )


@admin_only_view
@require_POST
def admin_viewer_user_toggle_view(request: HttpRequest, user_id: int) -> HttpResponse:
    from django.contrib.auth import get_user_model

    User = get_user_model()
    target_user = get_object_or_404(User, id=user_id)

    if target_user.is_staff or target_user.is_superuser:
        messages.error(request, "لا يمكن تعطيل حساب إداري من شاشة صلاحيات الاطلاع.")
        return redirect("visits:admin_viewer_users")

    meta = _readonly_role_meta_for_user(target_user)
    if meta["role"] == "none":
        messages.error(request, "هذا الحساب لا يملك صلاحية اطلاع.")
        return redirect("visits:admin_viewer_users")

    target_user.is_active = not target_user.is_active
    target_user.save(update_fields=["is_active"])
    messages.success(request, "تم تحديث حالة حساب الاطلاع.")
    return redirect("visits:admin_viewer_users")


@admin_only_view
def admin_viewer_user_password_view(request: HttpRequest, user_id: int) -> HttpResponse:
    from django.contrib.auth import get_user_model

    User = get_user_model()
    target_user = get_object_or_404(User, id=user_id)

    if target_user.is_staff or target_user.is_superuser:
        messages.error(request, "لا يمكن تغيير كلمة مرور مستخدم إداري من هذه الشاشة.")
        return redirect("visits:admin_viewer_users")

    meta = _readonly_role_meta_for_user(target_user)
    if meta["role"] == "none":
        messages.error(request, "هذا الحساب لا يملك صلاحية اطلاع.")
        return redirect("visits:admin_viewer_users")

    if request.method == "POST":
        password = request.POST.get("password") or ""
        password2 = request.POST.get("password2") or ""

        if not password:
            messages.error(request, "كلمة المرور الجديدة مطلوبة.")
        elif password != password2:
            messages.error(request, "كلمتا المرور غير متطابقتين.")
        elif len(password) < 8:
            messages.error(request, "كلمة المرور يجب ألا تقل عن 8 أحرف.")
        else:
            target_user.set_password(password)
            target_user.save(update_fields=["password"])
            messages.success(request, "تم تغيير كلمة المرور بنجاح.")
            return redirect("visits:admin_viewer_users")

    return render(
        request,
        "visits/admin_viewer_user_password.html",
        {
            "target_user": target_user,
            "meta": meta,
            "display_name": _viewer_user_display_name(target_user),
        },
    )



# =============================================================================
# Read-only management portal refinements
# ضع هذا الجزء في آخر visits/views.py ليحل محل دوال بوابة الاطلاع السابقة إن وجدت.
# =============================================================================
from django.contrib.auth import authenticate as _ro_authenticate, login as _ro_login, logout as _ro_logout
from django.contrib.auth.models import Group
from django.core.paginator import Paginator as _RoPaginator
from django.db.models import Q as _RoQ
from django.shortcuts import redirect as _ro_redirect, render as _ro_render, get_object_or_404 as _ro_get_object_or_404
from django.contrib import messages as _ro_messages
from django.views.decorators.http import require_POST as _ro_require_POST

try:
    from .models import Assignment as _RoAssignment, Plan as _RoPlan, PlanDay as _RoPlanDay, PlanWeek as _RoPlanWeek, Sector as _RoSector, School as _RoSchool, Supervisor as _RoSupervisor
except Exception:  # عند وجود الأسماء مستوردة مسبقًا في views.py
    _RoAssignment = Assignment
    _RoPlan = Plan
    _RoPlanDay = PlanDay
    _RoPlanWeek = PlanWeek
    _RoSector = Sector
    _RoSchool = School
    _RoSupervisor = Supervisor

RO_DEPARTMENT_GROUP_NAMES = {"مدير القسم", "مدير قسم", "readonly_department", "readonly_department_manager", "department_manager"}
RO_UNIT_GROUP_PREFIXES = ("مدير وحدة:", "مدير وحدة -", "مدير وحدة ", "readonly_unit:", "unit_manager:")
RO_SUP_SESSION_KEY = globals().get("SESSION_SUP_ID", "visits_sup_id")


def _ro_digits(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _ro_safe_int(value: object, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _ro_active_weeks():
    return _RoPlanWeek.objects.filter(is_break=False).order_by("week_no")


def _ro_all_weeks():
    return _RoPlanWeek.objects.all().order_by("week_no")


def _ro_default_week_no() -> int:
    week = _ro_active_weeks().first() or _ro_all_weeks().first()
    return int(getattr(week, "week_no", 1) or 1)


def _ro_week_choices():
    out = []
    for week in _ro_all_weeks():
        label = f"الأسبوع {week.week_no}"
        if getattr(week, "title", ""):
            label += f" — {week.title}"
        if getattr(week, "is_break", False):
            label += " (إجازة)"
        out.append((week.week_no, label))
    return out


def _ro_plan_status_label(plan) -> str:
    status = getattr(plan, "status", "") or ""
    if status == getattr(_RoPlan, "STATUS_APPROVED", "approved"):
        return "معتمدة"
    if status == getattr(_RoPlan, "STATUS_UNLOCK_REQUESTED", "unlock_requested"):
        return "طلب فك اعتماد بانتظار الإدارة"
    return "مسودة"


def _ro_plan_status_code(plan) -> str:
    status = getattr(plan, "status", "") or ""
    if status == getattr(_RoPlan, "STATUS_APPROVED", "approved"):
        return "approved"
    if status == getattr(_RoPlan, "STATUS_UNLOCK_REQUESTED", "unlock_requested"):
        return "unlock"
    return "draft"


def _ro_day_is_filled(day) -> bool:
    if not day:
        return False
    if getattr(day, "school_id", None):
        return True
    return getattr(day, "visit_type", "") == getattr(_RoPlanDay, "VISIT_NONE", "none")


def _ro_plan_filled_count(plan) -> int:
    days = list(getattr(plan, "days").all()) if hasattr(plan, "days") else []
    day_map = {getattr(day, "weekday", None): day for day in days}
    return sum(1 for weekday in range(5) if _ro_day_is_filled(day_map.get(weekday)))


def _ro_last_saved(plan):
    for attr in ("updated_at", "saved_at", "modified_at", "created_at"):
        value = getattr(plan, attr, None)
        if value:
            return value
    return None


def _ro_group_names(user) -> set[str]:
    try:
        return set(user.groups.values_list("name", flat=True))
    except Exception:
        return set()


def _ro_is_department_manager(user) -> bool:
    if getattr(user, "is_staff", False):
        return True
    groups = _ro_group_names(user)
    return bool(groups & RO_DEPARTMENT_GROUP_NAMES)


def _ro_unit_sector_ids(user) -> list[int]:
    sector_ids: set[int] = set()
    for group_name in _ro_group_names(user):
        for prefix in RO_UNIT_GROUP_PREFIXES:
            if group_name.startswith(prefix):
                raw = group_name.split(":", 1)[1]
                sid = _ro_safe_int(raw, 0)
                if sid:
                    sector_ids.add(sid)
    return sorted(sector_ids)


def _ro_can_view(user) -> bool:
    return bool(
        user
        and user.is_authenticated
        and getattr(user, "is_active", False)
        and (_ro_is_department_manager(user) or bool(_ro_unit_sector_ids(user)))
    )


def _ro_access_profile(user) -> dict:
    if _ro_is_department_manager(user):
        sectors = list(_RoSector.objects.filter(is_active=True).order_by("name"))
        return {
            "role_code": "department",
            "role_label": "مدير القسم" if not getattr(user, "is_staff", False) else "مدير النظام",
            "scope_label": "جميع القطاعات",
            "allowed_sector_ids": None,
            "allowed_sectors": sectors,
            "is_department": True,
        }

    sector_ids = _ro_unit_sector_ids(user)
    sectors = list(_RoSector.objects.filter(id__in=sector_ids, is_active=True).order_by("name"))
    return {
        "role_code": "unit",
        "role_label": "مدير وحدة",
        "scope_label": "، ".join(s.name for s in sectors) if sectors else "لم يحدد نطاق الاطلاع",
        "allowed_sector_ids": sector_ids,
        "allowed_sectors": sectors,
        "is_department": False,
    }


def _ro_require_viewer(view_func):
    def _wrapped(request, *args, **kwargs):
        if not _ro_can_view(getattr(request, "user", None)):
            _ro_messages.warning(request, "يرجى تسجيل الدخول إلى بوابة الاطلاع.")
            return _ro_redirect("visits:viewer_login")
        # منع تداخل جلسة المشرف مع بوابة الاطلاع
        request.session.pop(RO_SUP_SESSION_KEY, None)
        return view_func(request, *args, **kwargs)
    return _wrapped


def _ro_scoped_schools_qs(profile: dict):
    qs = _RoSchool.objects.filter(is_active=True).select_related("sector")
    allowed = profile.get("allowed_sector_ids")
    if allowed is not None:
        qs = qs.filter(sector_id__in=allowed)
    return qs


def _ro_scoped_assignments_qs(profile: dict):
    qs = (
        _RoAssignment.objects
        .filter(is_active=True, school__is_active=True)
        .select_related("supervisor", "school", "school__sector")
        .order_by("school__sector__name", "school__name", "supervisor__full_name")
    )
    allowed = profile.get("allowed_sector_ids")
    if allowed is not None:
        qs = qs.filter(school__sector_id__in=allowed)
    return qs


def _ro_scoped_supervisor_ids(profile: dict) -> list[int]:
    return list(
        _ro_scoped_assignments_qs(profile)
        .values_list("supervisor_id", flat=True)
        .distinct()
    )


def _ro_supervisor_sector_map(supervisor_ids: list[int], profile: dict) -> dict[int, str]:
    """يعرض قطاعات المشرف من الإسنادات النشطة، وليس من الخطة فقط."""
    if not supervisor_ids:
        return {}

    qs = (
        _RoAssignment.objects
        .filter(is_active=True, supervisor_id__in=supervisor_ids, school__is_active=True)
        .select_related("school__sector")
        .order_by("school__sector__name")
    )
    allowed = profile.get("allowed_sector_ids")
    if allowed is not None:
        qs = qs.filter(school__sector_id__in=allowed)

    tmp: dict[int, list[str]] = {}
    for assignment in qs:
        sector_name = getattr(getattr(getattr(assignment, "school", None), "sector", None), "name", "") or "—"
        tmp.setdefault(assignment.supervisor_id, [])
        if sector_name not in tmp[assignment.supervisor_id]:
            tmp[assignment.supervisor_id].append(sector_name)

    return {sid: "، ".join(names) for sid, names in tmp.items()}


def _ro_plan_rows(plans, profile: dict) -> list[dict]:
    plans = list(plans)
    supervisor_ids = [p.supervisor_id for p in plans]
    sector_map = _ro_supervisor_sector_map(supervisor_ids, profile)
    rows = []
    for plan in plans:
        filled = _ro_plan_filled_count(plan)
        rows.append({
            "plan": plan,
            "supervisor": plan.supervisor,
            "sector_names": sector_map.get(plan.supervisor_id, "—"),
            "status_label": _ro_plan_status_label(plan),
            "status_code": _ro_plan_status_code(plan),
            "filled": filled,
            "filled_label": f"{filled}/5",
            "is_full": filled == 5,
            "last_saved": _ro_last_saved(plan),
        })
    return rows


def _ro_common_context(request, *, page_title: str = "بوابة الاطلاع") -> dict:
    profile = _ro_access_profile(request.user)
    return {
        "viewer_profile": profile,
        "viewer_role_label": profile["role_label"],
        "viewer_scope_label": profile["scope_label"],
        "viewer_allowed_sectors": profile["allowed_sectors"],
        "viewer_page_title": page_title,
    }


def readonly_login_view(request):
    # إزالة أي جلسة مشرف عند دخول بوابة الاطلاع
    request.session.pop(RO_SUP_SESSION_KEY, None)

    if request.user.is_authenticated and _ro_can_view(request.user):
        return _ro_redirect("visits:viewer_dashboard")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        user = _ro_authenticate(request, username=username, password=password)
        if user and _ro_can_view(user):
            _ro_login(request, user)
            request.session.pop(RO_SUP_SESSION_KEY, None)
            _ro_messages.success(request, "تم تسجيل الدخول إلى بوابة الاطلاع.")
            return _ro_redirect("visits:viewer_dashboard")
        _ro_messages.error(request, "بيانات الدخول غير صحيحة أو أن الحساب لا يملك صلاحية اطلاع.")

    return _ro_render(request, "visits/readonly_login.html")


def readonly_logout_view(request):
    _ro_logout(request)
    request.session.pop(RO_SUP_SESSION_KEY, None)
    _ro_messages.success(request, "تم تسجيل الخروج من بوابة الاطلاع.")
    return _ro_redirect("visits:viewer_login")


@_ro_require_viewer
def readonly_dashboard_view(request):
    profile = _ro_access_profile(request.user)
    week_no = _ro_safe_int(request.GET.get("week") or _ro_default_week_no(), _ro_default_week_no())
    week_obj = _ro_get_object_or_404(_RoPlanWeek, week_no=week_no)

    supervisor_ids = _ro_scoped_supervisor_ids(profile)
    plans = (
        _RoPlan.objects
        .filter(week=week_obj, supervisor_id__in=supervisor_ids)
        .select_related("supervisor", "week")
        .prefetch_related("days")
        .order_by("supervisor__full_name")
    )

    plan_rows = _ro_plan_rows(plans[:8], profile)
    assignments = _ro_scoped_assignments_qs(profile)
    schools_qs = _ro_scoped_schools_qs(profile)

    context = _ro_common_context(request, page_title="لوحة الاطلاع")
    context.update({
        "week": week_no,
        "week_obj": week_obj,
        "week_choices": _ro_week_choices(),
        "plans_count": plans.count(),
        "approved_count": plans.filter(status=getattr(_RoPlan, "STATUS_APPROVED", "approved")).count(),
        "draft_count": plans.filter(status=getattr(_RoPlan, "STATUS_DRAFT", "draft")).count(),
        "unlock_count": plans.filter(status=getattr(_RoPlan, "STATUS_UNLOCK_REQUESTED", "unlock_requested")).count(),
        "supervisors_count": len(supervisor_ids),
        "schools_count": schools_qs.count(),
        "assignments_count": assignments.count(),
        "rows": plan_rows,
    })
    return _ro_render(request, "visits/readonly_dashboard.html", context)


@_ro_require_viewer
def readonly_plans_view(request):
    profile = _ro_access_profile(request.user)
    week_no = _ro_safe_int(request.GET.get("week") or _ro_default_week_no(), _ro_default_week_no())
    status = (request.GET.get("status") or "all").strip()
    q = (request.GET.get("q") or "").strip()
    sector_id = _ro_safe_int(request.GET.get("sector") or 0, 0)

    week_obj = _ro_get_object_or_404(_RoPlanWeek, week_no=week_no)

    allowed = profile.get("allowed_sector_ids")
    effective_allowed = allowed
    if sector_id:
        if allowed is None or sector_id in allowed:
            effective_allowed = [sector_id]
        else:
            effective_allowed = []

    scoped_profile = dict(profile)
    scoped_profile["allowed_sector_ids"] = effective_allowed
    supervisor_ids = _ro_scoped_supervisor_ids(scoped_profile)

    plans_qs = (
        _RoPlan.objects
        .filter(week=week_obj, supervisor_id__in=supervisor_ids)
        .select_related("supervisor", "week")
        .prefetch_related("days")
        .order_by("supervisor__full_name")
    )

    if q:
        plans_qs = plans_qs.filter(
            _RoQ(supervisor__full_name__icontains=q)
            | _RoQ(supervisor__national_id__icontains=q)
        )

    approved_status = getattr(_RoPlan, "STATUS_APPROVED", "approved")
    draft_status = getattr(_RoPlan, "STATUS_DRAFT", "draft")
    unlock_status = getattr(_RoPlan, "STATUS_UNLOCK_REQUESTED", "unlock_requested")

    plans_list_for_counts = list(plans_qs)
    approved_count = sum(1 for p in plans_list_for_counts if p.status == approved_status)
    draft_count = sum(1 for p in plans_list_for_counts if p.status == draft_status)
    unlock_count = sum(1 for p in plans_list_for_counts if p.status == unlock_status)
    not_full_count = sum(1 for p in plans_list_for_counts if _ro_plan_filled_count(p) < 5)

    if status == "approved":
        plans_filtered = [p for p in plans_list_for_counts if p.status == approved_status]
    elif status == "draft":
        plans_filtered = [p for p in plans_list_for_counts if p.status == draft_status]
    elif status == "unlock":
        plans_filtered = [p for p in plans_list_for_counts if p.status == unlock_status]
    elif status == "not_full":
        plans_filtered = [p for p in plans_list_for_counts if _ro_plan_filled_count(p) < 5]
    else:
        plans_filtered = plans_list_for_counts

    paginator = _RoPaginator(plans_filtered, 20)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    rows = _ro_plan_rows(page_obj.object_list, profile)

    context = _ro_common_context(request, page_title="خطط الزيارات")
    context.update({
        "week": week_no,
        "week_obj": week_obj,
        "week_choices": _ro_week_choices(),
        "status": status,
        "q": q,
        "sector_id": sector_id,
        "sector_choices": profile["allowed_sectors"],
        "rows": rows,
        "page_obj": page_obj,
        "plans_count": len(plans_filtered),
        "approved_count": approved_count,
        "draft_count": draft_count,
        "unlock_count": unlock_count,
        "not_full_count": not_full_count,
    })
    return _ro_render(request, "visits/readonly_plans.html", context)

@_ro_require_viewer
def readonly_plan_detail_view(request, plan_id: int):
    profile = _ro_access_profile(request.user)
    supervisor_ids = _ro_scoped_supervisor_ids(profile)
    plan = _ro_get_object_or_404(
        _RoPlan.objects.select_related("supervisor", "week").prefetch_related("days", "days__school"),
        id=plan_id,
        supervisor_id__in=supervisor_ids,
    )

    sector_map = _ro_supervisor_sector_map([plan.supervisor_id], profile)
    filled = _ro_plan_filled_count(plan)
    day_map = {day.weekday: day for day in plan.days.all()}

    weekday_names = [(0, "الأحد"), (1, "الإثنين"), (2, "الثلاثاء"), (3, "الأربعاء"), (4, "الخميس")]
    day_rows = []
    for wd, wd_name in weekday_names:
        day = day_map.get(wd)
        visit_label = "—"
        school_or_reason = "—"
        note = ""
        if day:
            try:
                visit_label = day.get_visit_type_display()
            except Exception:
                visit_label = getattr(day, "visit_type", "") or "—"

            if getattr(day, "school_id", None) and getattr(day, "school", None):
                school_or_reason = day.school.name
            elif getattr(day, "no_visit_reason", None):
                try:
                    school_or_reason = day.get_no_visit_reason_display()
                except Exception:
                    school_or_reason = getattr(day, "no_visit_reason", "") or "—"
            note = getattr(day, "note", "") or ""

        day_rows.append({
            "weekday": wd_name,
            "day": day,
            "visit_label": visit_label,
            "school_or_reason": school_or_reason,
            "note": note,
        })

    context = _ro_common_context(request, page_title="تفاصيل الخطة")
    context.update({
        "plan": plan,
        "sector_names": sector_map.get(plan.supervisor_id, "—"),
        "status_label": _ro_plan_status_label(plan),
        "status_code": _ro_plan_status_code(plan),
        "filled": filled,
        "filled_label": f"{filled}/5",
        "is_full": filled == 5,
        "last_saved": _ro_last_saved(plan),
        "day_rows": day_rows,
    })
    return _ro_render(request, "visits/readonly_plan_detail.html", context)


@_ro_require_viewer
def readonly_assignments_view(request):
    profile = _ro_access_profile(request.user)
    q = (request.GET.get("q") or "").strip()
    sector_id = _ro_safe_int(request.GET.get("sector") or 0, 0)

    allowed = profile.get("allowed_sector_ids")
    scoped_profile = dict(profile)
    if sector_id:
        if allowed is None or sector_id in allowed:
            scoped_profile["allowed_sector_ids"] = [sector_id]
        else:
            scoped_profile["allowed_sector_ids"] = []

    assignments = _ro_scoped_assignments_qs(scoped_profile)
    if q:
        assignments = assignments.filter(
            _RoQ(supervisor__full_name__icontains=q)
            | _RoQ(supervisor__national_id__icontains=q)
            | _RoQ(school__name__icontains=q)
            | _RoQ(school__stat_code__icontains=q)
        )

    paginator = _RoPaginator(assignments, 25)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    context = _ro_common_context(request, page_title="الإسنادات")
    context.update({
        "q": q,
        "sector_id": sector_id,
        "sector_choices": profile["allowed_sectors"],
        "page_obj": page_obj,
        "rows": page_obj.object_list,
        "assignments_count": assignments.count(),
        "schools_count": assignments.values("school_id").distinct().count(),
        "supervisors_count": assignments.values("supervisor_id").distinct().count(),
    })
    return _ro_render(request, "visits/readonly_assignments.html", context)

# =============================================================================
# Read-only portal sector display fix
# يعالج ظهور القطاع في بوابة الاطلاع بالاعتماد أولًا على قطاع المشرف،
# ثم قطاع المدرسة من الإسناد عند الحاجة.
# =============================================================================
def _ro_sector_name_from_obj(obj) -> str:
    try:
        name = getattr(getattr(obj, "sector", None), "name", "") or ""
        return str(name).strip()
    except Exception:
        return ""


def _ro_unique_sector_names(names) -> list[str]:
    out: list[str] = []
    for name in names:
        name = str(name or "").strip()
        if not name or name == "—":
            continue
        if name not in out:
            out.append(name)
    return out


def _ro_allowed_sector_ids(profile: dict):
    allowed = profile.get("allowed_sector_ids")
    if allowed is None:
        return None
    return [int(x) for x in allowed if str(x).isdigit()]


def _ro_scoped_schools_qs(profile: dict):
    """
    المدارس ضمن نطاق الاطلاع.
    لمدير الوحدة نعتمد على الإسنادات التي تقع ضمن قطاع المشرف أو قطاع المدرسة،
    لأن بعض المدارس قد لا يكون قطاعها مكتملًا بينما قطاع المشرف موجود في صفحة المشرفين.
    """
    qs = _RoSchool.objects.filter(is_active=True).select_related("sector")
    allowed = _ro_allowed_sector_ids(profile)
    if allowed is None:
        return qs.order_by("name")

    school_ids = (
        _RoAssignment.objects
        .filter(is_active=True, school__is_active=True)
        .filter(
            _RoQ(supervisor__sector_id__in=allowed)
            | _RoQ(school__sector_id__in=allowed)
        )
        .values_list("school_id", flat=True)
        .distinct()
    )
    return qs.filter(id__in=school_ids).order_by("name")


def _ro_scoped_assignments_qs(profile: dict):
    """
    إسنادات بوابة الاطلاع.
    عند تحديد نطاق قطاعي يتم اعتبار الإسناد ضمن النطاق إذا كان:
    - قطاع المشرف ضمن النطاق، أو
    - قطاع المدرسة ضمن النطاق.
    هذا يمنع ظهور القطاع فارغًا عندما يكون موجودًا في بطاقة المشرف فقط.
    """
    qs = (
        _RoAssignment.objects
        .filter(is_active=True, school__is_active=True)
        .select_related("supervisor", "supervisor__sector", "school", "school__sector")
        .order_by("supervisor__sector__name", "school__sector__name", "school__name", "supervisor__full_name")
    )
    allowed = _ro_allowed_sector_ids(profile)
    if allowed is not None:
        qs = qs.filter(
            _RoQ(supervisor__sector_id__in=allowed)
            | _RoQ(school__sector_id__in=allowed)
        )
    return qs


def _ro_scoped_supervisor_ids(profile: dict) -> list[int]:
    """
    المشرفون ضمن نطاق الاطلاع.
    مدير القسم يرى جميع المشرفين النشطين.
    مدير الوحدة يرى المشرفين المرتبطين بقطاعه مباشرة أو عبر إسناد مدرسة داخل قطاعه.
    """
    allowed = _ro_allowed_sector_ids(profile)

    if allowed is None:
        return list(
            _RoSupervisor.objects
            .filter(is_active=True)
            .values_list("id", flat=True)
            .distinct()
        )

    by_supervisor_sector = set(
        _RoSupervisor.objects
        .filter(is_active=True, sector_id__in=allowed)
        .values_list("id", flat=True)
    )

    by_assignment_school_sector = set(
        _RoAssignment.objects
        .filter(is_active=True, school__is_active=True, school__sector_id__in=allowed)
        .values_list("supervisor_id", flat=True)
        .distinct()
    )

    return sorted(by_supervisor_sector | by_assignment_school_sector)


def _ro_supervisor_sector_map(supervisor_ids: list[int], profile: dict) -> dict[int, str]:
    """
    يعرض قطاع المشرف في بوابة الاطلاع.
    الأولوية:
    1) القطاع المسجل في صفحة المشرفين Supervisor.sector
    2) قطاعات المدارس المسندة له Assignment.school.sector
    """
    if not supervisor_ids:
        return {}

    allowed = _ro_allowed_sector_ids(profile)
    result: dict[int, list[str]] = {int(sid): [] for sid in supervisor_ids}

    supervisors_qs = (
        _RoSupervisor.objects
        .filter(id__in=supervisor_ids)
        .select_related("sector")
    )
    if allowed is not None:
        supervisors_qs = supervisors_qs.filter(
            _RoQ(sector_id__in=allowed)
            | _RoQ(assignments__school__sector_id__in=allowed, assignments__is_active=True)
        ).distinct()

    for supervisor in supervisors_qs:
        sid = int(supervisor.id)
        name = _ro_sector_name_from_obj(supervisor)
        if name:
            result.setdefault(sid, [])
            result[sid].append(name)

    assignments_qs = (
        _RoAssignment.objects
        .filter(is_active=True, supervisor_id__in=supervisor_ids, school__is_active=True)
        .select_related("supervisor__sector", "school__sector")
    )
    if allowed is not None:
        assignments_qs = assignments_qs.filter(
            _RoQ(supervisor__sector_id__in=allowed)
            | _RoQ(school__sector_id__in=allowed)
        )

    for assignment in assignments_qs:
        sid = int(assignment.supervisor_id)
        result.setdefault(sid, [])

        sup_sector = _ro_sector_name_from_obj(getattr(assignment, "supervisor", None))
        school_sector = _ro_sector_name_from_obj(getattr(assignment, "school", None))

        if sup_sector:
            result[sid].append(sup_sector)
        if school_sector:
            result[sid].append(school_sector)

    return {
        sid: "، ".join(_ro_unique_sector_names(names)) or "—"
        for sid, names in result.items()
    }

