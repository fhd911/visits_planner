# visits/importers.py
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Assignment, Principal, School, Supervisor


# =============================================================================
# Result container (متوافق مع views_import.py)
# =============================================================================
@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    rejected_rows: list[dict[str, Any]] = field(default_factory=list)

    @property
    def stats_dict(self) -> dict[str, int]:
        return {"created": self.created, "updated": self.updated, "skipped": self.skipped}

    def reject(self, source: str, reason: str, row: dict[str, Any] | None = None) -> None:
        """سجل صف متجاهل مع السبب + معاينة بسيطة"""
        self.skipped += 1
        item: dict[str, Any] = {"source": source, "reason": reason}

        if row:
            # نلتقط أهم الحقول لتسهيل التصحيح
            pick = [
                "اسم المدرسة", "name",
                "الرقم الإحصائي", "stat_code", "school_stat_code",
                "السجل المدني", "national_id",
                "اسم المشرف", "supervisor_name", "المشرف",
                "اسم المدير", "full_name",
                "رقم الجوال", "mobile",
                "القسم", "department",
            ]
            mini = {}
            for k in pick:
                if k in row and row.get(k) not in (None, "", " "):
                    mini[k] = str(row.get(k)).strip()

            if mini:
                item.update(mini)
                item["_row_preview"] = " | ".join([f"{k}={v}" for k, v in list(mini.items())[:12]])
            else:
                item["_row_preview"] = "no-preview"

        self.rejected_rows.append(item)


# =============================================================================
# Helpers
# =============================================================================
def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _digits(v: Any) -> str:
    """أرقام فقط (هوية/جوال)"""
    s = _cell(v)
    if not s:
        return ""
    s = s.replace(".0", "").strip()
    return re.sub(r"\D+", "", s)


def _code(v: Any) -> str:
    """
    ✅ الرقم الإحصائي كـ code (حروف + أرقام)
    يدعم:
      70228
      M3964353
      70308-1  -> 703081
    """
    s = _cell(v)
    if not s:
        return ""
    s = s.replace(".0", "").strip()
    s = re.sub(r"[^A-Za-z0-9]+", "", s)
    return s.upper()


def _to_bool(v: Any) -> bool:
    s = _cell(v).lower()
    if s in {"1", "true", "yes", "y", "نعم"}:
        return True
    if s in {"0", "false", "no", "n", "لا"}:
        return False
    return True


def _canon_header(h: str) -> str:
    """
    توحيد أسماء الأعمدة (عربي/إنجليزي) لمفاتيح قياسية
    """
    x = _cell(h).lower().replace("_", " ").strip()

    # مدارس
    if x in {"اسم المدرسة", "name", "المدرسة"}:
        return "name"
    if x in {"الرقم الإحصائي", "الرقم الاحصائي", "stat code", "stat_code"}:
        return "stat_code"
    if x in {"نوع التعليم", "education type", "education_type"}:
        return "education_type"
    if x in {"المرحلة", "stage"}:
        return "stage"

    # مدير
    if x in {"اسم المدير", "اسم المديرة", "full_name", "الاسم"}:
        return "full_name"
    if x in {"رقم الجوال", "الجوال", "mobile"}:
        return "mobile"
    if x in {"قطاع المدرسة", "القطاع", "sector"}:
        return "sector"
    if x in {"السجل المدني", "رقم الهوية", "national id", "national_id", "الهوية"}:
        return "national_id"
    if x in {"رقم احصائي المدرسة", "school stat code", "school_stat_code"}:
        return "school_stat_code"

    # مشرف
    if x in {"اسم المشرف", "supervisor_name", "المشرف"}:
        return "supervisor_name"
    if x in {"القسم", "department"}:
        return "department"
    if x in {"رقم هوية المشرف", "supervisor_national_id"}:
        return "supervisor_national_id"

    return _cell(h)


def _load_rows(file) -> list[dict[str, Any]]:
    """
    يقرأ ملف Excel ويرجع list[dict]
    - مفاتيح عربية أصلية
    - + مفاتيح canonical لتسهيل الاستيراد
    """
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers_raw = [_cell(h) for h in rows[0]]
    out: list[dict[str, Any]] = []

    for r in rows[1:]:
        if not any(x is not None and _cell(x) != "" for x in r):
            continue

        d: dict[str, Any] = {}

        # الأصلية
        for i, h in enumerate(headers_raw):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None

        # القياسية
        for i, h in enumerate(headers_raw):
            canon = _canon_header(h)
            if canon and canon not in d:
                d[canon] = r[i] if i < len(r) else None

        out.append(d)

    return out


# =============================================================================
# Import Functions
# =============================================================================
@transaction.atomic
def import_schools(file, gender: str) -> ImportResult:
    """
    المدارس (غالبًا عندك):
    A: اسم المدرسة
    B: الرقم الإحصائي
    C: نوع التعليم
    D: المرحلة
    """
    res = ImportResult()
    rows = _load_rows(file)
    source = f"schools_{gender}"

    for row in rows:
        name = _cell(row.get("name") or row.get("اسم المدرسة"))
        stat_code = _code(row.get("stat_code") or row.get("الرقم الإحصائي"))
        edu = _cell(row.get("education_type") or row.get("نوع التعليم"))
        stage = _cell(row.get("stage") or row.get("المرحلة"))

        if not stat_code or not name:
            res.reject(source, "نقص بيانات: اسم المدرسة أو الرقم الإحصائي", row)
            continue

        obj, created = School.objects.update_or_create(
            stat_code=stat_code,
            defaults={
                "name": name,
                "gender": gender,
                "education_type": edu,
                "stage": stage,
                "is_active": True,
            },
        )
        if created:
            res.created += 1
        else:
            res.updated += 1

    return res


@transaction.atomic
def import_principals(file) -> ImportResult:
    """
    ملف المدراء:
    A: السجل المدني
    B: اسم المدير
    C: رقم الجوال
    D: الرقم الإحصائي
    E: قطاع المدرسة
    """
    res = ImportResult()
    rows = _load_rows(file)
    source = "principals"

    for row in rows:
        stat_code = _code(row.get("stat_code") or row.get("school_stat_code") or row.get("الرقم الإحصائي"))
        if not stat_code:
            res.reject(source, "نقص بيانات: الرقم الإحصائي", row)
            continue

        school = School.objects.filter(stat_code=stat_code).first()
        if not school:
            res.reject(source, f"مدرسة غير موجودة: stat_code={stat_code}", row)
            continue

        national_id = _digits(row.get("national_id") or row.get("السجل المدني"))
        full_name = _cell(row.get("full_name") or row.get("اسم المدير") or row.get("اسم المديرة"))
        mobile = _digits(row.get("mobile") or row.get("رقم الجوال"))
        sector = _cell(row.get("sector") or row.get("قطاع المدرسة"))

        if not full_name:
            res.reject(source, "نقص بيانات: اسم المدير/المديرة", row)
            continue

        obj, created = Principal.objects.update_or_create(
            school=school,
            defaults={
                "national_id": national_id,
                "full_name": full_name,
                "mobile": mobile,
                "sector": sector,
            },
        )
        if created:
            res.created += 1
        else:
            res.updated += 1

    return res


@transaction.atomic
def import_supervisors(file) -> ImportResult:
    """
    ملف المشرفين:
    A: السجل المدني
    B: اسم المشرف
    C: القسم
    """
    res = ImportResult()
    rows = _load_rows(file)
    source = "supervisors"

    for row in rows:
        national_id = _digits(row.get("national_id") or row.get("السجل المدني"))
        full_name = _cell(row.get("supervisor_name") or row.get("اسم المشرف") or row.get("full_name"))
        department = _cell(row.get("department") or row.get("القسم"))

        if not national_id or not full_name:
            res.reject(source, "نقص بيانات: هوية المشرف أو الاسم", row)
            continue

        obj, created = Supervisor.objects.update_or_create(
            national_id=national_id,
            defaults={
                "full_name": full_name,
                "department": department,
                "is_active": True,
            },
        )
        if created:
            res.created += 1
        else:
            res.updated += 1

    return res


@transaction.atomic
def import_assignments(file) -> ImportResult:
    """
    ملف الإسناد:
    - يلتقط هوية المشرف من:
      السجل المدني / supervisor_national_id / حتى من اسم المشرف إذا كان فيه أرقام
    - ويلتقط الرقم الإحصائي من:
      الرقم الإحصائي / stat_code / school_stat_code
    """
    res = ImportResult()
    rows = _load_rows(file)
    source = "assignments"

    for row in rows:
        sup_id = (
            _digits(row.get("supervisor_national_id"))
            or _digits(row.get("national_id"))
            or _digits(row.get("السجل المدني"))
        )

        # fallback: بعض ملفاتك يكون رقم الهوية داخل اسم المشرف
        if not sup_id:
            sup_id = _digits(row.get("supervisor_name") or row.get("اسم المشرف") or row.get("المشرف"))

        stat_code = _code(row.get("school_stat_code") or row.get("stat_code") or row.get("الرقم الإحصائي"))

        if not sup_id or not stat_code:
            res.reject(source, "نقص بيانات: هوية المشرف أو الرقم الإحصائي", row)
            continue

        supervisor = Supervisor.objects.filter(national_id=sup_id).first()
        if not supervisor:
            res.reject(source, f"مشرف غير موجود: national_id={sup_id}", row)
            continue

        school = School.objects.filter(stat_code=stat_code).first()
        if not school:
            res.reject(source, f"مدرسة غير موجودة: stat_code={stat_code}", row)
            continue

        obj, created = Assignment.objects.update_or_create(
            supervisor=supervisor,
            school=school,
            defaults={"is_active": True},
        )
        if created:
            res.created += 1
        else:
            res.updated += 1

    return res


# =============================================================================
# Build rejected.xlsx bytes (لتحميل المتجاهل)
# =============================================================================
def build_rejected_excel_bytes(rejected_rows: list[dict[str, Any]]) -> bytes:
    """
    يبني ملف rejected.xlsx فيه شيت لكل مصدر (schools/principals/supervisors/assignments)
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # تجميع حسب source
    groups: dict[str, list[dict[str, Any]]] = {}
    for r in rejected_rows:
        src = str(r.get("source") or "rejected")
        groups.setdefault(src, []).append(r)

    def safe_sheet(name: str) -> str:
        name = re.sub(r"[\[\]\*\?/\\:]", "-", name) or "rejected"
        return name[:31]

    for src, rows in groups.items():
        ws = wb.create_sheet(title=safe_sheet(src))

        # الأعمدة
        cols_set = set()
        for rr in rows:
            cols_set.update(rr.keys())

        preferred = [
            "source",
            "reason",
            "stat_code",
            "school_stat_code",
            "national_id",
            "supervisor_national_id",
            "supervisor_name",
            "full_name",
            "name",
            "_row_preview",
        ]
        cols = [c for c in preferred if c in cols_set] + sorted([c for c in cols_set if c not in preferred])

        # Header
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Data
        for i, rr in enumerate(rows, start=2):
            for c, col in enumerate(cols, start=1):
                v = rr.get(col, "")
                ws.cell(row=i, column=c, value=str(v) if v is not None else "")

        # widths
        for c, col in enumerate(cols, start=1):
            w = max(12, min(42, len(str(col)) + 2))
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.freeze_panes = "A2"

    from io import BytesIO

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
