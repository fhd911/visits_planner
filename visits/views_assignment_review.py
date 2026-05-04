from __future__ import annotations

from io import BytesIO

from django.apps import apps
from typing import Any

from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpRequest, HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Assignment, School, Supervisor
from .views import admin_only_view, _safe_int


ISSUE_ALL = "all"
ISSUE_UNASSIGNED = "unassigned_schools"
ISSUE_DUPLICATE = "duplicate_assignments"
ISSUE_INACTIVE_SUPERVISOR = "inactive_supervisor_assignments"
ISSUE_INACTIVE_SCHOOL = "inactive_school_assignments"
ISSUE_SUPERVISOR_WITHOUT_SCHOOLS = "supervisors_without_schools"
ISSUE_OVERLOADED = "overloaded_supervisors"

PRIORITY_CRITICAL = "critical"
PRIORITY_ACTION = "action"
PRIORITY_REVIEW = "review"

PRIORITY_LABELS = {
    PRIORITY_CRITICAL: "حرجة",
    PRIORITY_ACTION: "تحتاج معالجة",
    PRIORITY_REVIEW: "للمراجعة",
}

PRIORITY_ORDER = {
    PRIORITY_CRITICAL: 1,
    PRIORITY_ACTION: 2,
    PRIORITY_REVIEW: 3,
}

ISSUE_CHOICES = (
    (ISSUE_ALL, "كل الملاحظات"),
    (ISSUE_UNASSIGNED, "مدارس غير مسندة"),
    (ISSUE_DUPLICATE, "إسناد مكرر"),
    (ISSUE_INACTIVE_SUPERVISOR, "مشرف غير نشط لديه مدارس"),
    (ISSUE_INACTIVE_SCHOOL, "مدرسة معطلة لها إسناد"),
    (ISSUE_SUPERVISOR_WITHOUT_SCHOOLS, "مشرفون بلا مدارس"),
    (ISSUE_OVERLOADED, "مشرفون بحمل مرتفع"),
)


def _gender_label(value: str) -> str:
    value = (value or "").strip().lower()
    if value in ("boys", "male", "m", "بنين"):
        return "بنين"
    if value in ("girls", "female", "f", "بنات"):
        return "بنات"
    return "—"


def _school_label(school: School | None) -> str:
    if not school:
        return "—"
    stat_code = getattr(school, "stat_code", "") or "—"
    return f"{school.name} — {stat_code}"


def _supervisor_label(supervisor: Supervisor | None) -> str:
    if not supervisor:
        return "—"
    national_id = getattr(supervisor, "national_id", "") or "—"
    return f"{supervisor.full_name} — {national_id}"


def _row_matches_query(row: dict[str, Any], q: str) -> bool:
    if not q:
        return True

    haystack = " ".join(
        str(row.get(key, "") or "")
        for key in (
            "title",
            "detail",
            "school_name",
            "school_stat_code",
            "supervisor_name",
            "supervisor_national_id",
            "sector_name",
        )
    ).lower()

    return q.lower() in haystack




def _apply_priority(row: dict[str, Any]) -> dict[str, Any]:
    """
    يضيف مستوى الأولوية التشغيلي لكل ملاحظة:
    - حرجة: تعارض أو خلل مباشر.
    - تحتاج معالجة: حالة تؤثر على اكتمال الإسناد.
    - للمراجعة: حالة معلوماتية أو حمل مرتفع.
    """
    issue = row.get("issue")

    if issue in (ISSUE_DUPLICATE, ISSUE_INACTIVE_SUPERVISOR):
        priority = PRIORITY_CRITICAL
    elif issue in (ISSUE_UNASSIGNED, ISSUE_INACTIVE_SCHOOL):
        priority = PRIORITY_ACTION
    else:
        priority = PRIORITY_REVIEW

    row["priority"] = priority
    row["priority_label"] = PRIORITY_LABELS.get(priority, "للمراجعة")
    row["priority_order"] = PRIORITY_ORDER.get(priority, 99)
    return row


def _priority_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    return {
        PRIORITY_CRITICAL: sum(1 for row in rows if row.get("priority") == PRIORITY_CRITICAL),
        PRIORITY_ACTION: sum(1 for row in rows if row.get("priority") == PRIORITY_ACTION),
        PRIORITY_REVIEW: sum(1 for row in rows if row.get("priority") == PRIORITY_REVIEW),
    }


def _build_assignment_review_data(*, overload_threshold: int = 25) -> dict[str, Any]:
    """
    يبني بيانات مراجعة الإسناد من النماذج الحالية فقط:
    School / Supervisor / Assignment
    """
    active_assignments = Assignment.objects.filter(is_active=True)
    active_school_assignments = active_assignments.filter(school__is_active=True)
    active_supervisor_assignments = active_assignments.filter(supervisor__is_active=True)

    assigned_active_school_ids = set(
        active_assignments.filter(school__is_active=True)
        .values_list("school_id", flat=True)
        .distinct()
    )

    # 1) مدارس نشطة غير مسندة
    unassigned_schools = (
        School.objects.filter(is_active=True)
        .exclude(id__in=assigned_active_school_ids)
        .select_related("sector")
        .order_by("name")
    )

    # 2) مدارس لها أكثر من إسناد نشط
    duplicate_items = (
        active_assignments.values("school_id")
        .annotate(total=Count("id"))
        .filter(total__gt=1, school_id__isnull=False)
        .order_by("-total")
    )
    duplicate_school_ids = [item["school_id"] for item in duplicate_items]
    duplicate_total_by_school = {item["school_id"]: item["total"] for item in duplicate_items}

    duplicate_schools = (
        School.objects.filter(id__in=duplicate_school_ids)
        .select_related("sector")
        .order_by("name")
    )
    duplicate_assignments = (
        active_assignments.filter(school_id__in=duplicate_school_ids)
        .select_related("school", "supervisor", "school__sector")
        .order_by("school__name", "supervisor__full_name")
    )

    duplicate_supervisors_by_school: dict[int, list[str]] = {}
    for assignment in duplicate_assignments:
        duplicate_supervisors_by_school.setdefault(assignment.school_id, []).append(
            _supervisor_label(assignment.supervisor)
        )

    # 3) إسنادات نشطة لمشرفين غير نشطين
    inactive_supervisor_assignments = (
        active_assignments.filter(supervisor__is_active=False)
        .select_related("school", "supervisor", "school__sector")
        .order_by("supervisor__full_name", "school__name")
    )

    # 4) إسنادات نشطة لمدارس معطلة
    inactive_school_assignments = (
        active_assignments.filter(school__is_active=False)
        .select_related("school", "supervisor", "school__sector")
        .order_by("school__name")
    )

    # 5) مشرفون نشطون بلا مدارس
    supervisor_ids_with_active_assignments = set(
        active_assignments.filter(supervisor__is_active=True, school__is_active=True)
        .values_list("supervisor_id", flat=True)
        .distinct()
    )
    supervisors_without_schools = (
        Supervisor.objects.filter(is_active=True)
        .exclude(id__in=supervisor_ids_with_active_assignments)
        .order_by("full_name")
    )

    # 6) مشرفون بحمل مرتفع
    overloaded_items = (
        active_assignments.filter(supervisor__is_active=True, school__is_active=True)
        .values("supervisor_id")
        .annotate(total=Count("school_id", distinct=True))
        .filter(total__gte=overload_threshold)
        .order_by("-total")
    )
    overloaded_supervisor_ids = [item["supervisor_id"] for item in overloaded_items]
    overload_total_by_supervisor = {
        item["supervisor_id"]: item["total"] for item in overloaded_items
    }
    overloaded_supervisors = (
        Supervisor.objects.filter(id__in=overloaded_supervisor_ids)
        .order_by("full_name")
    )

    rows: list[dict[str, Any]] = []

    for school in unassigned_schools:
        rows.append(
            {
                "issue": ISSUE_UNASSIGNED,
                "issue_label": "مدرسة غير مسندة",
                "severity": "warning",
                "title": school.name,
                "detail": "مدرسة نشطة لا يوجد لها إسناد نشط لمشرف.",
                "school_name": school.name,
                "school_stat_code": school.stat_code or "",
                "school_gender": _gender_label(getattr(school, "gender", "")),
                "sector_name": getattr(school.sector, "name", "") or "—",
                "supervisor_name": "",
                "supervisor_national_id": "",
                "count": "",
                "action_label": "استيراد الإسناد",
                "action_url": reverse("visits:admin_schools_with_supervisors_import"),
            }
        )

    for school in duplicate_schools:
        supervisors = duplicate_supervisors_by_school.get(school.id, [])
        rows.append(
            {
                "issue": ISSUE_DUPLICATE,
                "issue_label": "إسناد مكرر",
                "severity": "danger",
                "title": school.name,
                "detail": "يوجد أكثر من إسناد نشط لهذه المدرسة: " + " / ".join(supervisors),
                "school_name": school.name,
                "school_stat_code": school.stat_code or "",
                "school_gender": _gender_label(getattr(school, "gender", "")),
                "sector_name": getattr(school.sector, "name", "") or "—",
                "supervisor_name": " / ".join(supervisors),
                "supervisor_national_id": "",
                "count": duplicate_total_by_school.get(school.id, ""),
                "action_label": "معالجة التكرار",
                "action_url": reverse("visits:admin_assignment_duplicate_resolve", args=[school.id]),
            }
        )

    for assignment in inactive_supervisor_assignments:
        rows.append(
            {
                "issue": ISSUE_INACTIVE_SUPERVISOR,
                "issue_label": "مشرف غير نشط",
                "severity": "danger",
                "title": assignment.supervisor.full_name,
                "detail": f"مشرف غير نشط لديه إسناد نشط على مدرسة: {_school_label(assignment.school)}",
                "school_name": assignment.school.name,
                "school_stat_code": assignment.school.stat_code or "",
                "school_gender": _gender_label(getattr(assignment.school, "gender", "")),
                "sector_name": getattr(assignment.school.sector, "name", "") or "—",
                "supervisor_name": assignment.supervisor.full_name,
                "supervisor_national_id": getattr(assignment.supervisor, "national_id", "") or "",
                "count": "",
                "action_label": "معالجة المشرف",
                "action_url": reverse("visits:admin_assignment_inactive_supervisor_resolve", args=[assignment.supervisor_id]),
            }
        )

    for assignment in inactive_school_assignments:
        rows.append(
            {
                "issue": ISSUE_INACTIVE_SCHOOL,
                "issue_label": "مدرسة معطلة",
                "severity": "warning",
                "title": assignment.school.name,
                "detail": f"مدرسة معطلة لا يزال لها إسناد نشط مع المشرف: {_supervisor_label(assignment.supervisor)}",
                "school_name": assignment.school.name,
                "school_stat_code": assignment.school.stat_code or "",
                "school_gender": _gender_label(getattr(assignment.school, "gender", "")),
                "sector_name": getattr(assignment.school.sector, "name", "") or "—",
                "supervisor_name": assignment.supervisor.full_name,
                "supervisor_national_id": getattr(assignment.supervisor, "national_id", "") or "",
                "count": "",
                "action_label": "معالجة المدرسة",
                "action_url": reverse("visits:admin_assignment_inactive_school_resolve", args=[assignment.school_id]),
            }
        )

    for supervisor in supervisors_without_schools:
        rows.append(
            {
                "issue": ISSUE_SUPERVISOR_WITHOUT_SCHOOLS,
                "issue_label": "مشرف بلا مدارس",
                "severity": "info",
                "title": supervisor.full_name,
                "detail": "مشرف نشط لا توجد له مدارس مسندة حاليًا.",
                "school_name": "",
                "school_stat_code": "",
                "school_gender": "",
                "sector_name": "",
                "supervisor_name": supervisor.full_name,
                "supervisor_national_id": getattr(supervisor, "national_id", "") or "",
                "count": 0,
                "action_label": "مدارس المشرف",
                "action_url": reverse("visits:admin_supervisor_assignments", args=[supervisor.id]),
            }
        )

    for supervisor in overloaded_supervisors:
        count = overload_total_by_supervisor.get(supervisor.id, 0)
        rows.append(
            {
                "issue": ISSUE_OVERLOADED,
                "issue_label": "حمل مرتفع",
                "severity": "warning",
                "title": supervisor.full_name,
                "detail": f"المشرف لديه {count} مدرسة نشطة مسندة، والحد المحدد للمراجعة هو {overload_threshold}.",
                "school_name": "",
                "school_stat_code": "",
                "school_gender": "",
                "sector_name": "",
                "supervisor_name": supervisor.full_name,
                "supervisor_national_id": getattr(supervisor, "national_id", "") or "",
                "count": count,
                "action_label": "مدارس المشرف",
                "action_url": reverse("visits:admin_supervisor_assignments", args=[supervisor.id]),
            }
        )

    rows = [_apply_priority(row) for row in rows]
    rows.sort(
        key=lambda row: (
            row.get("priority_order", 99),
            row.get("issue_label", ""),
            row.get("title", ""),
        )
    )

    priority_counts = _priority_counts(rows)

    return {
        "rows": rows,
        "counts": {
            ISSUE_UNASSIGNED: unassigned_schools.count(),
            ISSUE_DUPLICATE: len(duplicate_school_ids),
            ISSUE_INACTIVE_SUPERVISOR: inactive_supervisor_assignments.count(),
            ISSUE_INACTIVE_SCHOOL: inactive_school_assignments.count(),
            ISSUE_SUPERVISOR_WITHOUT_SCHOOLS: supervisors_without_schools.count(),
            ISSUE_OVERLOADED: overloaded_supervisors.count(),
            "total": len(rows),
            "critical": priority_counts[PRIORITY_CRITICAL],
            "action": priority_counts[PRIORITY_ACTION],
            "review": priority_counts[PRIORITY_REVIEW],
        },
    }


def _filter_assignment_review_rows(
    rows: list[dict[str, Any]],
    *,
    issue: str,
    q: str,
) -> list[dict[str, Any]]:
    if issue and issue != ISSUE_ALL:
        rows = [row for row in rows if row.get("issue") == issue]

    if q:
        rows = [row for row in rows if _row_matches_query(row, q)]

    return rows




LOG_ACTION_DUPLICATE_RESOLVED = "duplicate_resolved"
LOG_ACTION_INACTIVE_SUPERVISOR_DISABLED = "inactive_supervisor_disabled"
LOG_ACTION_INACTIVE_SCHOOL_DISABLED = "inactive_school_disabled"

LOG_TARGET_SCHOOL = "school"
LOG_TARGET_SUPERVISOR = "supervisor"

LOG_ACTION_CHOICES = (
    (LOG_ACTION_DUPLICATE_RESOLVED, "معالجة إسناد مكرر"),
    (LOG_ACTION_INACTIVE_SUPERVISOR_DISABLED, "تعطيل إسنادات مشرف غير نشط"),
    (LOG_ACTION_INACTIVE_SCHOOL_DISABLED, "تعطيل إسنادات مدرسة معطلة"),
)


def _get_assignment_review_log_model():
    """
    يرجع Model سجل معالجة الإسناد إذا كان موجودًا ومهاجرًا.
    وجوده اختياري حتى لا تتعطل الصفحات أثناء مرحلة التركيب.
    """
    try:
        return apps.get_model("visits", "AssignmentReviewLog")
    except Exception:
        return None


def _request_user_or_none(request: HttpRequest):
    user = getattr(request, "user", None)
    if user and getattr(user, "is_authenticated", False):
        return user
    return None


def _create_assignment_review_log(
    *,
    request: HttpRequest,
    action_type: str,
    target_type: str,
    title: str,
    details: str = "",
    school: School | None = None,
    supervisor: Supervisor | None = None,
    disabled_count: int = 0,
    metadata: dict[str, Any] | None = None,
):
    """
    يسجل عمليات المعالجة الحساسة إن كان AssignmentReviewLog موجودًا.
    إذا لم تتم إضافة الموديل أو الهجرة بعد، يتجاهل التسجيل ولا يعطل العملية.
    """
    LogModel = _get_assignment_review_log_model()
    if LogModel is None:
        return None

    try:
        return LogModel.objects.create(
            user=_request_user_or_none(request),
            action_type=action_type,
            target_type=target_type,
            school=school,
            supervisor=supervisor,
            title=title,
            details=details or "",
            disabled_count=disabled_count or 0,
            metadata=metadata or {},
        )
    except Exception:
        return None

def _excel_response(wb: Workbook, filename: str) -> HttpResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_assignment_review_workbook(rows: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "مراجعة الإسناد"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    header_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="E8F5E9")
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "الأولوية",
        "نوع الملاحظة",
        "الخطورة",
        "العنوان",
        "التفاصيل",
        "الرقم الإحصائي",
        "المدرسة",
        "الجنس",
        "القطاع",
        "المشرف",
        "سجل المشرف",
        "العدد",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="تقرير مراجعة الإسناد").font = title_font
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=f"تاريخ التصدير: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=2, column=1).font = normal_font
    ws.cell(row=2, column=1).alignment = center

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for row in rows:
        values = [
            row.get("priority_label", ""),
            row.get("issue_label", ""),
            row.get("severity", ""),
            row.get("title", ""),
            row.get("detail", ""),
            row.get("school_stat_code", ""),
            row.get("school_name", ""),
            row.get("school_gender", ""),
            row.get("sector_name", ""),
            row.get("supervisor_name", ""),
            row.get("supervisor_national_id", ""),
            row.get("count", ""),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 5, 7, 10, 11) else right

        row_idx += 1

    widths = {
        1: 16,
        2: 22,
        3: 14,
        4: 32,
        5: 70,
        6: 18,
        7: 36,
        8: 12,
        9: 22,
        10: 32,
        11: 18,
        12: 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    return wb


@admin_only_view
def admin_assignment_review_view(request: HttpRequest) -> HttpResponse:
    issue = (request.GET.get("issue") or ISSUE_ALL).strip()
    if issue not in dict(ISSUE_CHOICES):
        issue = ISSUE_ALL

    q = (request.GET.get("q") or "").strip()
    overload_threshold = _safe_int(request.GET.get("threshold") or 25, default=25)
    if overload_threshold < 1:
        overload_threshold = 25

    data = _build_assignment_review_data(overload_threshold=overload_threshold)
    filtered_rows = _filter_assignment_review_rows(
        data["rows"],
        issue=issue,
        q=q,
    )

    paginator = Paginator(filtered_rows, 30)
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/admin_assignment_review.html",
        {
            "rows": list(page_obj.object_list),
            "page_obj": page_obj,
            "q": q,
            "issue": issue,
            "issue_choices": ISSUE_CHOICES,
            "counts": data["counts"],
            "overload_threshold": overload_threshold,
            "priority_labels": PRIORITY_LABELS,
        },
    )


@admin_only_view
def admin_assignment_review_export_view(request: HttpRequest) -> HttpResponse:
    issue = (request.GET.get("issue") or ISSUE_ALL).strip()
    if issue not in dict(ISSUE_CHOICES):
        issue = ISSUE_ALL

    q = (request.GET.get("q") or "").strip()
    overload_threshold = _safe_int(request.GET.get("threshold") or 25, default=25)
    if overload_threshold < 1:
        overload_threshold = 25

    data = _build_assignment_review_data(overload_threshold=overload_threshold)
    rows = _filter_assignment_review_rows(
        data["rows"],
        issue=issue,
        q=q,
    )

    wb = _build_assignment_review_workbook(rows)
    filename = f"assignment_review_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, filename)



@admin_only_view
def admin_assignment_duplicate_resolve_view(request: HttpRequest, school_id: int) -> HttpResponse:
    """
    صفحة معالجة الإسناد المكرر لمدرسة محددة.

    تعرض كل الإسنادات النشطة للمدرسة، وتسمح باختيار مشرف واحد للإبقاء عليه
    ثم تعطيل بقية الإسنادات النشطة لنفس المدرسة.
    """
    school = get_object_or_404(
        School.objects.select_related("sector"),
        id=school_id,
    )

    assignments = list(
        Assignment.objects.filter(
            school=school,
            is_active=True,
        )
        .select_related("supervisor")
        .order_by("supervisor__full_name", "id")
    )

    inactive_assignments = list(
        Assignment.objects.filter(
            school=school,
            is_active=False,
        )
        .select_related("supervisor")
        .order_by("-id")[:10]
    )

    return render(
        request,
        "visits/admin_assignment_duplicate_resolve.html",
        {
            "school": school,
            "assignments": assignments,
            "inactive_assignments": inactive_assignments,
            "active_count": len(assignments),
        },
    )


@admin_only_view
@require_POST
def admin_assignment_duplicate_keep_view(request: HttpRequest, school_id: int) -> HttpResponse:
    """
    تنفيذ معالجة التكرار:
    - يبقى الإسناد المحدد نشطًا.
    - تعطل بقية الإسنادات النشطة للمدرسة نفسها.
    """
    school = get_object_or_404(School, id=school_id)

    keep_assignment_id = _safe_int(request.POST.get("keep_assignment_id") or 0, default=0)

    assignments_qs = Assignment.objects.filter(
        school=school,
        is_active=True,
    ).select_related("supervisor")

    keep_assignment = assignments_qs.filter(id=keep_assignment_id).first()

    if not keep_assignment:
        messages.error(request, "لم يتم تحديد إسناد صحيح للإبقاء عليه.")
        return redirect("visits:admin_assignment_duplicate_resolve", school_id=school.id)

    disabled_assignment_ids = list(
        Assignment.objects.filter(
            school=school,
            is_active=True,
        )
        .exclude(id=keep_assignment.id)
        .values_list("id", flat=True)
    )

    disabled_count = (
        Assignment.objects.filter(id__in=disabled_assignment_ids)
        .update(is_active=False)
    )

    _create_assignment_review_log(
        request=request,
        action_type=LOG_ACTION_DUPLICATE_RESOLVED,
        target_type=LOG_TARGET_SCHOOL,
        school=school,
        supervisor=keep_assignment.supervisor,
        title=f"معالجة إسناد مكرر للمدرسة: {school.name}",
        details=(
            f"تم الإبقاء على إسناد المشرف/ {keep_assignment.supervisor.full_name} "
            f"وتعطيل {disabled_count} إسناد آخر للمدرسة."
        ),
        disabled_count=disabled_count,
        metadata={
            "school_id": school.id,
            "school_name": school.name,
            "school_stat_code": getattr(school, "stat_code", "") or "",
            "kept_assignment_id": keep_assignment.id,
            "kept_supervisor_id": keep_assignment.supervisor_id,
            "kept_supervisor_name": keep_assignment.supervisor.full_name,
            "disabled_assignment_ids": disabled_assignment_ids,
        },
    )

    messages.success(
        request,
        f"تم الإبقاء على إسناد المشرف/ {keep_assignment.supervisor.full_name} وتعطيل {disabled_count} إسناد آخر للمدرسة.",
    )

    return redirect("visits:admin_assignment_duplicate_resolve", school_id=school.id)



def _build_inactive_supervisor_assignments_workbook(
    *,
    supervisor: Supervisor,
    assignments,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "مدارس المشرف غير النشط"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    header_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="FEE2E2")
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "م",
        "الرقم الإحصائي",
        "اسم المدرسة",
        "الجنس",
        "القطاع",
        "حالة المدرسة",
        "حالة الإسناد",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="مدارس مسندة لمشرف غير نشط")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    info_cell = ws.cell(
        row=2,
        column=1,
        value=f"المشرف: {supervisor.full_name} — السجل المدني: {getattr(supervisor, 'national_id', '') or '—'}",
    )
    info_cell.font = normal_font
    info_cell.alignment = center

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for index, assignment in enumerate(assignments, start=1):
        school = assignment.school
        values = [
            index,
            getattr(school, "stat_code", "") or "—",
            getattr(school, "name", "") or "—",
            _gender_label(getattr(school, "gender", "")),
            getattr(getattr(school, "sector", None), "name", "") or "—",
            "نشطة" if getattr(school, "is_active", False) else "معطلة",
            "نشط" if getattr(assignment, "is_active", False) else "غير نشط",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 6, 7) else right

        row_idx += 1

    widths = {
        1: 8,
        2: 18,
        3: 42,
        4: 12,
        5: 24,
        6: 14,
        7: 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    return wb


@admin_only_view
def admin_assignment_inactive_supervisor_resolve_view(
    request: HttpRequest,
    supervisor_id: int,
) -> HttpResponse:
    """
    صفحة معالجة حالة: مشرف غير نشط لديه مدارس مسندة.

    تعرض المدارس المسندة للمشرف غير النشط، وتتيح:
    - تصدير مدارسه.
    - فتح صفحة إسناداته.
    - تعطيل جميع إسناداته النشطة بعد التأكيد.
    """
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)

    assignments = list(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
        )
        .select_related("school", "school__sector")
        .order_by("school__name")
    )

    active_school_count = sum(
        1 for assignment in assignments
        if getattr(assignment.school, "is_active", False)
    )
    inactive_school_count = len(assignments) - active_school_count

    return render(
        request,
        "visits/admin_assignment_inactive_supervisor_resolve.html",
        {
            "supervisor": supervisor,
            "assignments": assignments,
            "assignment_count": len(assignments),
            "active_school_count": active_school_count,
            "inactive_school_count": inactive_school_count,
        },
    )


@admin_only_view
def admin_assignment_inactive_supervisor_export_view(
    request: HttpRequest,
    supervisor_id: int,
) -> HttpResponse:
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)

    assignments = list(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
        )
        .select_related("school", "school__sector")
        .order_by("school__name")
    )

    wb = _build_inactive_supervisor_assignments_workbook(
        supervisor=supervisor,
        assignments=assignments,
    )
    filename = f"inactive_supervisor_assignments_{supervisor_id}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, filename)


@admin_only_view
@require_POST
def admin_assignment_inactive_supervisor_disable_view(
    request: HttpRequest,
    supervisor_id: int,
) -> HttpResponse:
    """
    يعطل جميع الإسنادات النشطة للمشرف غير النشط.
    لا يحذف الإسنادات من قاعدة البيانات.
    """
    supervisor = get_object_or_404(Supervisor, id=supervisor_id)

    confirm = (request.POST.get("confirm") or "").strip()
    if confirm != "1":
        messages.error(request, "يجب تأكيد العملية قبل تعطيل الإسنادات.")
        return redirect("visits:admin_assignment_inactive_supervisor_resolve", supervisor_id=supervisor.id)

    disabled_assignment_ids = list(
        Assignment.objects.filter(
            supervisor=supervisor,
            is_active=True,
        ).values_list("id", flat=True)
    )

    disabled_count = Assignment.objects.filter(
        id__in=disabled_assignment_ids,
    ).update(is_active=False)

    _create_assignment_review_log(
        request=request,
        action_type=LOG_ACTION_INACTIVE_SUPERVISOR_DISABLED,
        target_type=LOG_TARGET_SUPERVISOR,
        supervisor=supervisor,
        title=f"تعطيل إسنادات مشرف غير نشط: {supervisor.full_name}",
        details=f"تم تعطيل {disabled_count} إسناد نشط للمشرف غير النشط.",
        disabled_count=disabled_count,
        metadata={
            "supervisor_id": supervisor.id,
            "supervisor_name": supervisor.full_name,
            "supervisor_national_id": getattr(supervisor, "national_id", "") or "",
            "disabled_assignment_ids": disabled_assignment_ids,
        },
    )

    messages.success(
        request,
        f"تم تعطيل {disabled_count} إسناد نشط للمشرف/ {supervisor.full_name}.",
    )

    return redirect("visits:admin_assignment_inactive_supervisor_resolve", supervisor_id=supervisor.id)



def _build_inactive_school_assignments_workbook(
    *,
    school: School,
    assignments,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "إسنادات مدرسة معطلة"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    header_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="FEF3C7")
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "م",
        "اسم المشرف",
        "السجل المدني",
        "الجوال",
        "البريد",
        "حالة المشرف",
        "حالة الإسناد",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="إسنادات نشطة لمدرسة معطلة")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    info_cell = ws.cell(
        row=2,
        column=1,
        value=f"المدرسة: {school.name} — الرقم الإحصائي: {getattr(school, 'stat_code', '') or '—'}",
    )
    info_cell.font = normal_font
    info_cell.alignment = center

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for index, assignment in enumerate(assignments, start=1):
        supervisor = assignment.supervisor
        values = [
            index,
            getattr(supervisor, "full_name", "") or "—",
            getattr(supervisor, "national_id", "") or "—",
            getattr(supervisor, "mobile", "") or "—",
            getattr(supervisor, "email", "") or "—",
            "نشط" if getattr(supervisor, "is_active", False) else "غير نشط",
            "نشط" if getattr(assignment, "is_active", False) else "غير نشط",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 3, 4, 6, 7) else right

        row_idx += 1

    widths = {
        1: 8,
        2: 34,
        3: 18,
        4: 18,
        5: 32,
        6: 14,
        7: 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    return wb


@admin_only_view
def admin_assignment_inactive_school_resolve_view(
    request: HttpRequest,
    school_id: int,
) -> HttpResponse:
    """
    صفحة معالجة حالة: مدرسة معطلة لها إسناد نشط.

    تعرض المشرفين المسندين للمدرسة المعطلة، وتتيح:
    - تصدير الإسنادات النشطة.
    - تعطيل جميع إسنادات المدرسة النشطة بعد التأكيد.
    """
    school = get_object_or_404(
        School.objects.select_related("sector"),
        id=school_id,
    )

    assignments = list(
        Assignment.objects.filter(
            school=school,
            is_active=True,
        )
        .select_related("supervisor")
        .order_by("supervisor__full_name")
    )

    active_supervisor_count = sum(
        1 for assignment in assignments
        if getattr(assignment.supervisor, "is_active", False)
    )
    inactive_supervisor_count = len(assignments) - active_supervisor_count

    return render(
        request,
        "visits/admin_assignment_inactive_school_resolve.html",
        {
            "school": school,
            "assignments": assignments,
            "assignment_count": len(assignments),
            "active_supervisor_count": active_supervisor_count,
            "inactive_supervisor_count": inactive_supervisor_count,
        },
    )


@admin_only_view
def admin_assignment_inactive_school_export_view(
    request: HttpRequest,
    school_id: int,
) -> HttpResponse:
    school = get_object_or_404(
        School.objects.select_related("sector"),
        id=school_id,
    )

    assignments = list(
        Assignment.objects.filter(
            school=school,
            is_active=True,
        )
        .select_related("supervisor")
        .order_by("supervisor__full_name")
    )

    wb = _build_inactive_school_assignments_workbook(
        school=school,
        assignments=assignments,
    )
    filename = f"inactive_school_assignments_{school_id}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, filename)


@admin_only_view
@require_POST
def admin_assignment_inactive_school_disable_view(
    request: HttpRequest,
    school_id: int,
) -> HttpResponse:
    """
    يعطل جميع الإسنادات النشطة للمدرسة المعطلة.
    لا يحذف الإسنادات من قاعدة البيانات.
    """
    school = get_object_or_404(School, id=school_id)

    confirm = (request.POST.get("confirm") or "").strip()
    if confirm != "1":
        messages.error(request, "يجب تأكيد العملية قبل تعطيل الإسنادات.")
        return redirect("visits:admin_assignment_inactive_school_resolve", school_id=school.id)

    disabled_assignment_ids = list(
        Assignment.objects.filter(
            school=school,
            is_active=True,
        ).values_list("id", flat=True)
    )

    disabled_count = Assignment.objects.filter(
        id__in=disabled_assignment_ids,
    ).update(is_active=False)

    _create_assignment_review_log(
        request=request,
        action_type=LOG_ACTION_INACTIVE_SCHOOL_DISABLED,
        target_type=LOG_TARGET_SCHOOL,
        school=school,
        title=f"تعطيل إسنادات مدرسة معطلة: {school.name}",
        details=f"تم تعطيل {disabled_count} إسناد نشط للمدرسة المعطلة.",
        disabled_count=disabled_count,
        metadata={
            "school_id": school.id,
            "school_name": school.name,
            "school_stat_code": getattr(school, "stat_code", "") or "",
            "disabled_assignment_ids": disabled_assignment_ids,
        },
    )

    messages.success(
        request,
        f"تم تعطيل {disabled_count} إسناد نشط للمدرسة/ {school.name}.",
    )

    return redirect("visits:admin_assignment_inactive_school_resolve", school_id=school.id)



def _log_action_display(value: str) -> str:
    return dict(LOG_ACTION_CHOICES).get(value, value or "—")


def _log_target_display(value: str) -> str:
    return {
        LOG_TARGET_SCHOOL: "مدرسة",
        LOG_TARGET_SUPERVISOR: "مشرف",
    }.get(value, value or "—")


def _build_assignment_review_logs_workbook(logs) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "سجل معالجة الإسناد"
    ws.sheet_view.rightToLeft = True

    title_font = Font(name="Cairo", bold=True, size=14)
    header_font = Font(name="Cairo", bold=True, size=11)
    normal_font = Font(name="Cairo", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="E8F5E9")
    header_fill = PatternFill("solid", fgColor="F1F5F9")

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "م",
        "تاريخ التنفيذ",
        "نوع العملية",
        "المستهدف",
        "المدرسة",
        "المشرف",
        "عدد الإسنادات المعطلة",
        "المنفذ",
        "التفاصيل",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="سجل عمليات معالجة الإسناد").font = title_font
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).alignment = center

    header_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = header_row + 1
    for index, log in enumerate(logs, start=1):
        created_at = timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M")
        user_label = getattr(log.user, "username", "") or "—"
        school_label = getattr(log.school, "name", "") or "—"
        supervisor_label = getattr(log.supervisor, "full_name", "") or "—"

        values = [
            index,
            created_at,
            log.get_action_type_display() if hasattr(log, "get_action_type_display") else _log_action_display(getattr(log, "action_type", "")),
            log.get_target_type_display() if hasattr(log, "get_target_type_display") else _log_target_display(getattr(log, "target_type", "")),
            school_label,
            supervisor_label,
            getattr(log, "disabled_count", 0),
            user_label,
            getattr(log, "details", "") or getattr(log, "title", ""),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center if col in (1, 2, 3, 4, 7, 8) else right

        row_idx += 1

    widths = {
        1: 8,
        2: 20,
        3: 30,
        4: 14,
        5: 36,
        6: 32,
        7: 18,
        8: 18,
        9: 70,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"
    return wb


def _assignment_review_logs_queryset(request: HttpRequest):
    q = (request.GET.get("q") or "").strip()
    action_type = (request.GET.get("action_type") or "all").strip()

    LogModel = _get_assignment_review_log_model()
    if LogModel is None:
        return [], q, action_type

    logs = LogModel.objects.select_related(
        "user",
        "school",
        "supervisor",
    ).order_by("-created_at", "-id")

    valid_action_types = {
        LOG_ACTION_DUPLICATE_RESOLVED,
        LOG_ACTION_INACTIVE_SUPERVISOR_DISABLED,
        LOG_ACTION_INACTIVE_SCHOOL_DISABLED,
    }

    if action_type in valid_action_types:
        logs = logs.filter(action_type=action_type)

    if q:
        logs = logs.filter(
            Q(title__icontains=q)
            | Q(details__icontains=q)
            | Q(school__name__icontains=q)
            | Q(school__stat_code__icontains=q)
            | Q(supervisor__full_name__icontains=q)
            | Q(supervisor__national_id__icontains=q)
            | Q(user__username__icontains=q)
        )

    return logs, q, action_type


@admin_only_view
def admin_assignment_review_logs_view(request: HttpRequest) -> HttpResponse:
    logs_qs, q, action_type = _assignment_review_logs_queryset(request)

    paginator = Paginator(logs_qs, 30)
    page_obj = paginator.get_page(_safe_int(request.GET.get("page") or 1, default=1))

    return render(
        request,
        "visits/admin_assignment_review_logs.html",
        {
            "rows": list(page_obj.object_list),
            "page_obj": page_obj,
            "q": q,
            "action_type": action_type,
            "action_choices": LOG_ACTION_CHOICES,
        },
    )


@admin_only_view
def admin_assignment_review_logs_export_view(request: HttpRequest) -> HttpResponse:
    logs_qs, _q, _action_type = _assignment_review_logs_queryset(request)
    logs = list(logs_qs[:5000])

    wb = _build_assignment_review_logs_workbook(logs)
    filename = f"assignment_review_logs_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, filename)
