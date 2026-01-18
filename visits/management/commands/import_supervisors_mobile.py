from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from visits.models import Supervisor


def digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


class Command(BaseCommand):
    help = "Import/Update supervisors mobile from Excel (by national_id)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to xlsx file")
        parser.add_argument("--sheet", type=str, default=None, help="Sheet name (optional)")
        parser.add_argument("--start-row", type=int, default=2, help="Data start row (default 2)")
        parser.add_argument("--col-nid", type=str, default="B", help="National ID column (default B)")
        parser.add_argument("--col-mobile", type=str, default="C", help="Mobile column (default C)")

    def handle(self, *args, **opts):
        xlsx_path = Path(opts["xlsx_path"])
        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        if opts["sheet"]:
            if opts["sheet"] not in wb.sheetnames:
                raise CommandError(f"Sheet not found: {opts['sheet']}. Available: {wb.sheetnames}")
            ws = wb[opts["sheet"]]
        else:
            ws = wb.active

        start_row = int(opts["start_row"])
        col_nid = (opts["col_nid"] or "B").upper().strip()
        col_mobile = (opts["col_mobile"] or "C").upper().strip()

        updated = 0
        skipped = 0
        not_found = 0

        for r in range(start_row, ws.max_row + 1):
            nid_raw = ws[f"{col_nid}{r}"].value
            mob_raw = ws[f"{col_mobile}{r}"].value

            nid = digits(str(nid_raw or "")).strip()
            mob = digits(str(mob_raw or "")).strip()

            if not nid:
                skipped += 1
                continue

            sup = Supervisor.objects.filter(national_id=nid).first()
            if not sup:
                not_found += 1
                continue

            # اقبل 9-12 رقم تقريبًا (05xxxxxxxx أو 9665xxxxxxxx)
            if mob and len(mob) >= 9:
                sup.mobile = mob
                sup.save(update_fields=["mobile"])
                updated += 1
            else:
                # لو فاضي لا نحذفه
                skipped += 1

        self.stdout.write(self.style.SUCCESS(f"Updated mobiles: {updated}"))
        self.stdout.write(self.style.WARNING(f"Not found (national_id not in DB): {not_found}"))
        self.stdout.write(self.style.WARNING(f"Skipped rows: {skipped}"))
